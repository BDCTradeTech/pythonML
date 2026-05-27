"""
Fase 3 — tabs/compras.py
Pestaña Invoices (Compras): vista de invoices QuickBooks del cliente.
"""
from __future__ import annotations

import os
import tempfile
from datetime import datetime
from typing import Any, Dict, List, Optional

from nicegui import app, background_tasks, run, ui
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from db import (
    get_despachantes,
    get_invoice_extras,
    get_qb_app_credentials,
    get_qb_tokens,
    get_user_qb_customer,
    upsert_invoice_extra,
)
from qb_api import (
    fetch_qb_customer_detail,
    fetch_qb_invoice_detail,
    fetch_qb_invoice_pdf,
    fetch_qb_invoices,
    patch_invoice_pdf_line_items,
)


# ---------------------------------------------------------------------------
# Helpers de sesión (patrón unificado de todos los tabs extraídos)
# ---------------------------------------------------------------------------

def _require_login() -> Optional[Dict[str, Any]]:
    user = app.storage.user.get("user")
    if not user:
        ui.notify("Debes iniciar sesión para continuar", color="negative")
    return user


# ---------------------------------------------------------------------------
# Helper compartido con main.py (build_tab_historicos / build_tab_stock).
# Se unificará en qb_api.py cuando esos tabs también se extraigan.
# ---------------------------------------------------------------------------

def _qb_invoice_pdf_download_basename(doc: Any) -> str:
    """Nombre de archivo sugerido para PDF de invoice: `{doc}.pdf` (sin prefijo invoice_)."""
    base = str(doc or "invoice").strip().replace(" ", "_")
    for c in '<>:"/\\|?*':
        base = base.replace(c, "_")
    return f"{base[:80]}.pdf"


# ---------------------------------------------------------------------------
# Tab principal
# ---------------------------------------------------------------------------

def build_tab_compras(container) -> None:
    """Pestaña Invoices: conexión a QuickBooks para mostrar Invoices del cliente."""
    user = _require_login()
    if not user:
        return

    qb_creds = get_qb_app_credentials(user["id"])
    qb_tokens = get_qb_tokens(user["id"])

    with container:
        if not qb_creds:
            ui.label(
                "Configurá QuickBooks en Configuración (Client ID, Client Secret, Redirect URI) y conectá tu cuenta."
            ).classes("text-gray-600")
            return

        if not qb_tokens:
            ui.label(
                "Credenciales configuradas. Andá a Configuración → QuickBooks y hacé clic en 'Conectar cuenta' para autorizar."
            ).classes("text-warning")
            return

        header_card = ui.column().classes("w-full mb-2")
        filtro_row = ui.row().classes("w-full mb-2 items-center gap-4")
        result_area = ui.column().classes("w-full gap-2")

        with result_area:
            with ui.card().classes("w-full p-8 items-center gap-4"):
                ui.spinner(size="xl")
                ui.label("Cargando invoices...").classes("text-xl text-gray-700")

        invoices_ref: Dict[str, List[Dict[str, Any]]] = {"data": []}
        header_data_ref: Dict[str, Any] = {}
        sort_col_compras: Dict[str, str] = {"val": "txn_date"}
        sort_asc_compras: Dict[str, bool] = {"val": False}
        filtro_status_ref: Dict[str, str] = {"val": "Abierta+Vencida"}
        filtro_estado_ref: Dict[str, str] = {"val": "Todos"}
        filtro_courier_ref: Dict[str, str] = {"val": "Todas"}
        invoice_bdc_qsel_css_done: Dict[str, bool] = {"done": False}
        _desp_hex_palette = [
            "#1d4ed8",
            "#7e22ce",
            "#b45309",
            "#0e7490",
            "#a21caf",
            "#4338ca",
            "#c2410c",
            "#0369a1",
            "#6d28d9",
            "#be185d",
        ]

        def _color_despachante_hex(nombre: str) -> str:
            n = (nombre or "").strip()
            if not n:
                return "#4b5563"
            key = n.lower()
            if key == "sixtar":
                return "#ea580c"  # naranja
            if key == "lhs":
                return "#38bdf8"  # celeste
            return _desp_hex_palette[sum(ord(c) for c in n) % len(_desp_hex_palette)]

        def _norm_factura_courier(v: Any) -> str:
            s = str(v or "").strip().lower()
            return "Pagada" if s == "pagada" else "Impaga"

        def _hex_factura_courier(val: Any) -> str:
            return "#16a34a" if _norm_factura_courier(val) == "Pagada" else "#dc2626"

        def _ensure_invoice_bdc_qsel_css() -> None:
            if invoice_bdc_qsel_css_done["done"]:
                return
            invoice_bdc_qsel_css_done["done"] = True
            ui.add_head_html(
                """
<style>
.invoice-bdc-qsel .q-field__native,
.invoice-bdc-qsel .q-field__native span {
  color: var(--qsel-color, #374151) !important;
  font-weight: 600;
}
</style>
"""
            )

        def _fmt_fecha(s: str) -> str:
            """Convierte YYYY-MM-DD a dd-mm-yyyy."""
            if not s or len(str(s)) < 10:
                return str(s) if s else "—"
            p = str(s)[:10].split("-")
            return f"{p[2]}-{p[1]}-{p[0]}" if len(p) == 3 else str(s)

        def _mostrar_detalle_invoice(inv: Dict[str, Any]) -> None:
            """Abre un popup con el detalle de la factura, cargando desde QuickBooks."""
            invoice_popup_ctx: Dict[str, Any] = {}
            dlg = ui.dialog()
            with dlg:
                with ui.card().classes("p-6 min-w-[650px] max-w-[90vw] max-h-[70vh] overflow-hidden flex flex-col"):
                    with ui.row().classes("items-center gap-2 shrink-0"):
                        ui.label("Detalle de la factura").classes("text-lg font-semibold")
                        ui.label("Invoice nro").classes("text-base text-gray-600")
                        ui.label(str(inv.get("doc", "—"))).classes("text-base font-medium")
                    ui.separator().classes("mb-3")
                    cont = ui.column().classes("gap-2 overflow-y-auto min-h-0 flex-1")

            async def _cargar_y_mostrar() -> None:
                with cont:
                    ui.spinner(size="md")
                    ui.label("Cargando detalle...").classes("text-gray-600")
                detail, err = await run.io_bound(fetch_qb_invoice_detail, user["id"], inv.get("id", ""))
                cont.clear()
                with cont:
                    if err:
                        ui.label(f"Error: {err}").classes("text-negative")
                    else:
                        inv_obj = detail or {}
                        invoice_popup_ctx["inv_obj"] = inv_obj
                        def _fmt_dd_mm_aaaa(s: str) -> str:
                            if not s or len(str(s)) < 10:
                                return str(s) if s else "—"
                            parts = str(s)[:10].split("-")
                            return f"{parts[2]}-{parts[1]}-{parts[0]}" if len(parts) == 3 else str(s)
                        doc = inv_obj.get("DocNumber", inv.get("doc", "—"))
                        txn = _fmt_dd_mm_aaaa(str(inv_obj.get("TxnDate", inv.get("txn_date", ""))))
                        due = _fmt_dd_mm_aaaa(str(inv_obj.get("DueDate", inv.get("due_date", ""))))
                        total = inv_obj.get("TotalAmt", inv.get("amount_num"))
                        bal = inv_obj.get("Balance", inv.get("balance"))
                        try:
                            total_fmt = f"{float(total):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if total is not None else "—"
                        except (TypeError, ValueError):
                            total_fmt = str(total) or "—"
                        try:
                            bal_fmt = f"{float(bal):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if bal is not None else "—"
                        except (TypeError, ValueError):
                            bal_fmt = str(bal) or "—"
                        _label_w = "w-28"
                        with ui.element("table").classes("w-full text-sm table-fixed"):
                            with ui.element("colgroup"):
                                ui.element("col").classes(_label_w)
                                ui.element("col")
                            for lbl, val in [("Fecha", txn), ("Vencimiento", due)]:
                                with ui.element("tr"):
                                    with ui.element("td").classes(f"font-semibold pr-4 py-1 {_label_w}"):
                                        ui.label(lbl)
                                    with ui.element("td").classes("py-1"):
                                        ui.label(str(val))
                        ui.element("div").classes("border-t border-gray-300 my-2")
                        with ui.element("table").classes("w-full text-sm table-fixed"):
                            with ui.element("colgroup"):
                                ui.element("col").classes(_label_w)
                                ui.element("col")
                            for lbl, val in [
                                ("Total", f"u$ {total_fmt}" if total_fmt != "—" else "—"),
                                ("Saldo", f"u$ {bal_fmt}" if bal_fmt != "—" else "—"),
                            ]:
                                with ui.element("tr"):
                                    with ui.element("td").classes(f"font-semibold pr-4 py-1 {_label_w}"):
                                        ui.label(lbl)
                                    with ui.element("td").classes("py-1"):
                                        ui.label(str(val))
                        ui.separator().classes("my-2")
                        lines = inv_obj.get("Line") or []
                        if isinstance(lines, dict):
                            lines = [lines]
                        if lines:
                            ui.label("Ítems").classes("font-semibold mt-3 mb-1")
                            with ui.element("table").classes("w-full text-sm border"):
                                with ui.element("thead"):
                                    with ui.element("tr").classes("bg-gray-100"):
                                        with ui.element("th").classes("px-2 py-1 text-left"):
                                            ui.label("Descripción")
                                        with ui.element("th").classes("px-2 py-1 text-right"):
                                            ui.label("Cant.")
                                        with ui.element("th").classes("px-2 py-1 text-right"):
                                            ui.label("Importe")
                                with ui.element("tbody"):
                                    for idx, lin in enumerate(lines):
                                        sales = lin.get("SalesItemLineDetail") or {}
                                        desc = lin.get("Description", sales.get("ItemRef", {}).get("name", "—") if isinstance(sales, dict) else "—")
                                        if idx == len(lines) - 1 and (str(desc).strip() in ("-", "—", "")):
                                            desc = "Total"
                                        qty = sales.get("Qty", 1) if isinstance(sales, dict) else 1
                                        amt = lin.get("Amount", 0)
                                        try:
                                            amt_str = f"{float(amt):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                                        except (TypeError, ValueError):
                                            amt_str = str(amt)
                                        qty_display = "" if desc == "Total" else str(qty)
                                        with ui.element("tr").classes("border-t"):
                                            with ui.element("td").classes("px-2 py-1"):
                                                ui.label(str(desc)[:80])
                                            with ui.element("td").classes("px-2 py-1 text-right"):
                                                ui.label(qty_display)
                                            with ui.element("td").classes("px-2 py-1 text-right"):
                                                ui.label(f"u$ {amt_str}")

                    with ui.row().classes("mt-3 gap-2"):
                        def _descargar_pdf() -> None:
                            pdf_bytes, err = fetch_qb_invoice_pdf(user["id"], inv.get("id", ""))
                            if err:
                                ui.notify(f"Error al descargar PDF: {err}", color="negative")
                                return
                            if not pdf_bytes:
                                ui.notify("No se pudo obtener el PDF", color="warning")
                                return
                            try:
                                fd, path = tempfile.mkstemp(suffix=".pdf")
                                os.write(fd, pdf_bytes)
                                os.close(fd)
                                nombre = _qb_invoice_pdf_download_basename(inv.get("doc", "invoice"))
                                with dlg:
                                    ui.download(path, nombre)
                                ui.notify("Descarga iniciada", color="positive")
                            except Exception as ex:
                                ui.notify(f"Error: {ex}", color="negative")

                        def _descargar_pdf_parcheado(reemplazo: str) -> None:
                            inv_obj = invoice_popup_ctx.get("inv_obj") or {}
                            pdf_bytes, err = fetch_qb_invoice_pdf(user["id"], inv.get("id", ""))
                            if err:
                                ui.notify(f"Error al descargar PDF: {err}", color="negative")
                                return
                            if not pdf_bytes:
                                ui.notify("No se pudo obtener el PDF", color="warning")
                                return
                            patched, perr = patch_invoice_pdf_line_items(
                                pdf_bytes,
                                inv_obj,
                                reemplazo,
                                user_id=user["id"],
                                sku_interleaved_display=True,
                            )
                            if not patched:
                                ui.notify(perr or "No se pudo modificar el PDF", color="negative")
                                return
                            try:
                                fd, path = tempfile.mkstemp(suffix=".pdf")
                                os.write(fd, patched)
                                os.close(fd)
                                base_fn = _qb_invoice_pdf_download_basename(inv.get("doc", "invoice"))
                                stem = base_fn[:-4] if base_fn.lower().endswith(".pdf") else base_fn
                                nombre = f"BDC_{stem}.pdf"
                                with dlg:
                                    ui.download(path, nombre)
                                if perr:
                                    ui.notify(perr, color="warning")
                                else:
                                    ui.notify("PDF modificado (todas las descripciones)", color="positive")
                            except Exception as ex:
                                ui.notify(f"Error: {ex}", color="negative")

                        ui.button("Cerrar popup", on_click=dlg.close).props("dense no-caps")
                        ui.button("Invoice", on_click=_descargar_pdf, color="secondary").props("dense no-caps icon=download")
                        ui.button(
                            "Mouse",
                            on_click=lambda: _descargar_pdf_parcheado("MOUSE"),
                            color="secondary",
                        ).props("dense no-caps icon=download")
                        ui.button(
                            "Smartwatch",
                            on_click=lambda: _descargar_pdf_parcheado("SMARTWATCH"),
                            color="secondary",
                        ).props("dense no-caps icon=download")

            dlg.open()
            background_tasks.create(_cargar_y_mostrar(), name="invoice_detail")

        def _generar_excel_invoices() -> tuple[Optional[str], Optional[str], Optional[str]]:
            """Genera el Excel en un hilo. Retorna (path, nombre_archivo, None) si OK, o (None, None, error_msg)."""
            try:
                invs = invoices_ref.get("data", [])
                filtro_val = filtro_status_ref.get("val", "Abierta+Vencida")
                if filtro_val == "Abierta+Vencida":
                    invs = [i for i in invs if (i.get("status") or "").lower() in ("abierta", "vencida")]
                elif filtro_val != "Todas":
                    invs = [i for i in invs if (i.get("status") or "").lower() == filtro_val.lower()]
                if filtro_estado_ref.get("val", "Todos") != "Todos":
                    invs = [i for i in invs if (i.get("estado") or "En USA") == filtro_estado_ref["val"]]
                fc_filt = filtro_courier_ref.get("val", "Todas")
                if fc_filt == "Pagas":
                    invs = [i for i in invs if _norm_factura_courier(i.get("factura_courier")) == "Pagada"]
                elif fc_filt == "Impagas":
                    invs = [i for i in invs if _norm_factura_courier(i.get("factura_courier")) == "Impaga"]
                sc = sort_col_compras.get("val", "txn_date")
                asc = sort_asc_compras.get("val", False)

                def _sk(x: Dict, col: str):
                    if col == "importe_factura":
                        s = str(x.get("importe_factura") or "").replace(" ", "").lstrip("$").strip()
                        if not s:
                            return 0.0
                        try:
                            if "," in s:
                                s = s.replace(".", "").replace(",", ".")
                            return float(s)
                        except (ValueError, TypeError):
                            return 0.0
                    if col == "factura_courier":
                        return _norm_factura_courier(x.get("factura_courier"))
                    v = x.get(col) or ""
                    if col in ("amount", "amount_num"):
                        try:
                            return float(x.get("amount_num") or 0)
                        except (ValueError, TypeError):
                            return 0
                    return str(v).lower()

                invs_exp = sorted(invs, key=lambda x: _sk(x, sc), reverse=not asc)
                if not invs_exp:
                    return (None, None, "No hay invoices para exportar")
                cust_name = (header_data_ref.get("cust_name") or "Cliente").replace("/", "-").replace("\\", "-")[:50]
                ahora = datetime.now()
                nombre_archivo = f"{cust_name}-{ahora.year:04d}-{ahora.month:02d}-{ahora.day:02d}-{ahora.hour:02d}{ahora.minute:02d}.xlsx"
                sheet_name = f"Invoices {ahora.day:02d}-{ahora.month:02d}-{ahora.year % 100:02d}"[:31]

                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                thin_side = Side(border_style="thin")
                all_borders = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

                headers = ["Fecha", "Días", "Invoice", "Importe", "Status"]
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=1, column=col, value=h)
                    c.fill = black_fill
                    c.font = header_font
                    c.border = all_borders
                    c.alignment = Alignment(horizontal="center", vertical="center")

                today_date = datetime.now().date()
                for idx, inv in enumerate(invs_exp, 2):
                    txn = inv.get("txn_date", "") or ""
                    fecha_ddmm = _fmt_fecha(txn) if txn else "—"
                    try:
                        dt = datetime.strptime(str(txn)[:10], "%Y-%m-%d").date() if len(str(txn)) >= 10 else None
                        dias_val = (today_date - dt).days if dt else None
                    except (ValueError, TypeError):
                        dias_val = None
                    doc_raw = inv.get("doc", "") or ""
                    try:
                        invoice_num = int(str(doc_raw).strip()) if str(doc_raw).strip().isdigit() else None
                    except (ValueError, TypeError):
                        invoice_num = None
                    amt = inv.get("amount_num", 0) or 0
                    try:
                        importe_num = float(amt)
                    except (ValueError, TypeError):
                        importe_num = 0.0
                    status = inv.get("status", "—")

                    c_fecha = ws.cell(row=idx, column=1, value=fecha_ddmm)
                    c_fecha.fill = white_fill
                    c_fecha.border = all_borders

                    c_dias = ws.cell(row=idx, column=2, value=dias_val)
                    c_dias.fill = white_fill
                    c_dias.border = all_borders
                    c_dias.number_format = "0"

                    c_inv = ws.cell(row=idx, column=3, value=invoice_num if invoice_num is not None else doc_raw or "—")
                    c_inv.fill = white_fill
                    c_inv.border = all_borders
                    if invoice_num is not None:
                        c_inv.number_format = "0"

                    c_imp = ws.cell(row=idx, column=4, value=importe_num)
                    c_imp.fill = white_fill
                    c_imp.border = all_borders
                    c_imp.number_format = '$#,##0.00'

                    c_status = ws.cell(row=idx, column=5, value=status)
                    c_status.fill = white_fill
                    c_status.border = all_borders

                ws.column_dimensions["A"].width = 12
                ws.column_dimensions["B"].width = 8
                ws.column_dimensions["C"].width = 14
                ws.column_dimensions["D"].width = 14
                ws.column_dimensions["E"].width = 12

                fd, path = tempfile.mkstemp(suffix=".xlsx")
                os.close(fd)
                wb.save(path)
                return (path, nombre_archivo, None)
            except Exception as e:
                return (None, None, str(e))

        def _pintar_compras() -> None:
            header_card.clear()
            filtro_row.clear()
            result_area.clear()
            invs = invoices_ref.get("data", [])
            filtro_val = filtro_status_ref.get("val", "Abierta+Vencida")
            if filtro_val == "Abierta+Vencida":
                invs = [i for i in invs if (i.get("status") or "").lower() in ("abierta", "vencida")]
            elif filtro_val != "Todas":
                invs = [i for i in invs if (i.get("status") or "").lower() == filtro_val.lower()]
            filtro_estado_val = filtro_estado_ref.get("val", "Todos")
            if filtro_estado_val != "Todos":
                invs = [i for i in invs if (i.get("estado") or "En USA") == filtro_estado_val]
            fc_filt = filtro_courier_ref.get("val", "Todas")
            if fc_filt == "Pagas":
                invs = [i for i in invs if _norm_factura_courier(i.get("factura_courier")) == "Pagada"]
            elif fc_filt == "Impagas":
                invs = [i for i in invs if _norm_factura_courier(i.get("factura_courier")) == "Impaga"]
            sc = sort_col_compras.get("val", "txn_date")
            asc = sort_asc_compras.get("val", False)

            def _sort_key(x: Dict, col: str):
                if col == "importe_factura":
                    s = str(x.get("importe_factura") or "").replace(" ", "").lstrip("$").strip()
                    if not s:
                        return 0.0
                    try:
                        if "," in s:
                            s = s.replace(".", "").replace(",", ".")
                        return float(s)
                    except (ValueError, TypeError):
                        return 0.0
                if col == "factura_courier":
                    return _norm_factura_courier(x.get("factura_courier"))
                v = x.get(col) or ""
                if col in ("amount", "amount_num"):
                    n = x.get("amount_num")
                    try:
                        return float(n) if n is not None else 0
                    except (ValueError, TypeError):
                        return 0
                return str(v).lower()

            invs_sorted = sorted(invs, key=lambda x: _sort_key(x, sc), reverse=not asc)
            total_importe = sum(i.get("amount_num", 0) for i in invs_sorted)
            total_fmt = f"{total_importe:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            header_data_ref["total_importe"] = f"u$ {total_fmt}"

            # Calcular días por invoice (para métricas de deuda)
            today_date = datetime.now().date()
            dias_list: List[int] = []
            for inv in invs_sorted:
                txn = inv.get("txn_date", "") or ""
                try:
                    if len(str(txn)) >= 10:
                        dt = datetime.strptime(str(txn)[:10], "%Y-%m-%d").date()
                        dias_list.append((today_date - dt).days)
                except (ValueError, TypeError):
                    pass
            deuda_promedio_dias = f"{sum(dias_list) / len(dias_list):.1f}" if dias_list else "—"
            deuda_mas_antigua = str(max(dias_list)) if dias_list else "—"
            cantidad_ordenes = len(invs_sorted)

            with header_card:
                if not header_data_ref:
                    ui.label("Cargando...").classes("text-gray-600")
                    return
                ui.label("Invoices").classes("text-xl font-semibold text-gray-800 mb-2")
                open_balance = header_data_ref.get("open_balance", "—")
                overdue = header_data_ref.get("overdue", "0.00")
                with ui.card().classes("w-full p-4 bg-grey-2"):
                    with ui.row().classes("w-full gap-6 flex-wrap items-center"):
                        with ui.column().classes("gap-0"):
                            ui.label("Total Deuda").classes("text-base font-semibold text-gray-800")
                            ui.label(open_balance).classes("text-sm text-gray-600")
                        ui.element("div").classes("w-px h-8 bg-gray-400 shrink-0")
                        with ui.column().classes("gap-0"):
                            ui.label("Deuda Vencida").classes("text-base font-semibold text-gray-800")
                            ui.label(overdue).classes("text-sm text-gray-600")
                        ui.element("div").classes("w-px h-8 bg-gray-400 shrink-0")
                        with ui.column().classes("gap-0"):
                            ui.label("Deuda Seleccionada").classes("text-base font-semibold text-gray-800")
                            ui.label(header_data_ref.get("total_importe", "u$ 0,00")).classes("text-sm text-gray-600")
                        ui.element("div").classes("w-px h-8 bg-gray-400 shrink-0")
                        with ui.column().classes("gap-0"):
                            ui.label("Deuda promedio días").classes("text-base font-semibold text-gray-800")
                            ui.label(deuda_promedio_dias).classes("text-sm text-gray-600")
                        ui.element("div").classes("w-px h-8 bg-gray-400 shrink-0")
                        with ui.column().classes("gap-0"):
                            ui.label("Deuda más antigua").classes("text-base font-semibold text-gray-800")
                            ui.label(deuda_mas_antigua).classes("text-sm text-gray-600")
                        ui.element("div").classes("w-px h-8 bg-gray-400 shrink-0")
                        with ui.column().classes("gap-0"):
                            ui.label("Órdenes Impagas").classes("text-base font-semibold text-gray-800")
                            ui.label(str(cantidad_ordenes)).classes("text-sm text-gray-600")
            with filtro_row:
                ui.label("Pago Courier:").classes("text-sm font-semibold text-gray-800")
                filtro_courier = ui.select(
                    {"Todas": "Todas", "Pagas": "Pagas", "Impagas": "Impagas"},
                    value=filtro_courier_ref.get("val", "Todas"),
                    label="",
                ).classes("w-40").props("dense outlined")

                def _on_filtro_courier_change(e):
                    val = getattr(e, "args", None) or getattr(e, "value", None)
                    filtro_courier_ref["val"] = str(val) if val is not None else "Todas"
                    _pintar_compras()

                filtro_courier.on_value_change(_on_filtro_courier_change)
                filtro_status_val = filtro_status_ref.get("val", "Abierta+Vencida")
                ui.label("Status:").classes("text-sm font-semibold text-gray-800 ml-4")
                filtro_status = ui.select(
                    {"Abierta+Vencida": "Abierta+Vencida", "Todas": "Todas", "Abierta": "Abierta", "Vencida": "Vencida", "Pagada": "Pagada"},
                    value=filtro_status_val,
                    label="",
                ).classes("w-44").props("dense outlined")
                def _on_filtro_change(e):
                    val = getattr(e, "args", None) or getattr(e, "value", None)
                    filtro_status_ref["val"] = str(val) if val is not None else "Abierta+Vencida"
                    _pintar_compras()
                filtro_status.on_value_change(_on_filtro_change)
                ui.label("Estado:").classes("text-sm font-semibold text-gray-800 ml-4")
                filtro_estado = ui.select(
                    {"Todos": "Todos", "En USA": "En USA", "Viajando": "Viajando", "Recibida": "Recibida"},
                    value=filtro_estado_ref.get("val", "Todos"),
                    label="",
                ).classes("w-40").props("dense outlined")
                def _on_filtro_estado_change(e):
                    val = getattr(e, "args", None) or getattr(e, "value", None)
                    filtro_estado_ref["val"] = str(val) if val is not None else "Todos"
                    _pintar_compras()
                filtro_estado.on_value_change(_on_filtro_estado_change)
                async def _imprimir_invoices_async() -> None:
                    path, nombre_archivo, err = await run.io_bound(_generar_excel_invoices)
                    if err:
                        ui.notify(err, color="warning" if "No hay" in (err or "") else "negative")
                        return
                    if path and nombre_archivo:
                        ui.download(path, nombre_archivo)
                        ui.notify(f"Exportado: {nombre_archivo}", color="positive")
                        def _cleanup() -> None:
                            try:
                                if path and os.path.exists(path):
                                    os.unlink(path)
                            except Exception:
                                pass
                        ui.timer(5.0, _cleanup, once=True)
                ui.button("Imprimir invoices", on_click=_imprimir_invoices_async, color="primary").props("dense no-caps icon=print").classes("ml-4")

            with result_area:
                _ensure_invoice_bdc_qsel_css()

                def _on_sort(c: str) -> None:
                    if sort_col_compras.get("val") == c:
                        sort_asc_compras["val"] = not sort_asc_compras.get("val", False)
                    else:
                        sort_col_compras["val"] = c
                        sort_asc_compras["val"] = False
                    _pintar_compras()

                def _save_inv_extra(inv: Dict, **kwargs) -> None:
                    qid = inv.get("id", "")
                    if not qid:
                        return
                    upsert_invoice_extra(user["id"], qid, **kwargs)
                    for k, v in kwargs.items():
                        inv[k] = str(v) if v is not None else ""
                    ui.notify("Guardado", color="positive")

                def _on_desp_change(e: Any, inv: Dict[str, Any], wrap: Any) -> None:
                    _save_inv_extra(inv, despachante=e.value or "")
                    wrap.style(f"--qsel-color: {_color_despachante_hex(e.value or '')}")

                def _on_fc_change(e: Any, inv: Dict[str, Any], wrap: Any) -> None:
                    v = _norm_factura_courier(e.value)
                    _save_inv_extra(inv, factura_courier=v)
                    wrap.style(f"--qsel-color: {_hex_factura_courier(v)}")

                if not invs_sorted:
                    ui.label("No hay facturas." if not invs else "No hay facturas con ese Status.").classes("text-gray-500")
                else:
                    with ui.element("div").classes("w-full overflow-x-auto"):
                        with ui.element("table").classes("w-full border-collapse text-sm"):
                            with ui.element("thead"):
                                with ui.element("tr").classes("bg-primary text-white font-semibold"):
                                    estado_opts = {"En USA": "En USA", "Viajando": "Viajando", "Recibida": "Recibida"}
                                    cols = [
                                        ("numero", "Numero"),
                                        ("txn_date", "Fecha"),
                                        ("dias", "Días"),
                                        ("tipo", "Tipo"),
                                        ("doc", "Nº"),
                                        ("guia", "Guía"),
                                        ("despachante", "Despachante"),
                                        ("importe_factura", "Importe factura"),
                                        ("factura_courier", "Factura Courier"),
                                        ("pa", "PA"),
                                        ("estado", "Estado"),
                                        ("amount", "Importe"),
                                        ("status", "Status"),
                                    ]
                                    for col_key, h in cols:
                                        with ui.element("th").classes("px-2 py-2 border text-center"):
                                            if col_key in ("numero", "dias"):
                                                ui.label(h)
                                            else:
                                                ui.button(h, on_click=lambda c=col_key: _on_sort(c)).props("flat dense no-caps").classes("text-white hover:bg-white/20 cursor-pointer font-semibold")
                            with ui.element("tbody"):
                                for idx, inv in enumerate(invs_sorted, 1):
                                    row_el = ui.element("tr").classes("border-t border-gray-200 hover:bg-gray-50")
                                    with row_el:
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            ui.label(str(idx))
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            ui.label(_fmt_fecha(inv.get("txn_date", "—")))
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            txn = inv.get("txn_date", "") or ""
                                            try:
                                                dt = datetime.strptime(str(txn)[:10], "%Y-%m-%d").date() if len(str(txn)) >= 10 else None
                                                dias = (datetime.now().date() - dt).days if dt else None
                                            except (ValueError, TypeError):
                                                dias = None
                                            dias_val = str(dias) if dias is not None else "—"
                                            dias_cls = "text-red-600 font-semibold" if dias is not None and dias > 30 else "text-gray-900"
                                            ui.label(dias_val).classes(dias_cls)
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            ui.label(inv.get("tipo", "Factura"))
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            with ui.row().classes("gap-1 items-center"):
                                                ui.label(inv.get("doc", "—"))
                                                ui.button("Ver", on_click=lambda inv=inv: _mostrar_detalle_invoice(inv)).props("flat dense size=sm")
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            inp_guia = ui.input(value=inv.get("guia", "")).classes("w-40").props("dense")
                                            inp_guia.on("blur", lambda evt, inv=inv, inp=inp_guia: _save_inv_extra(inv, guia=inp.value))
                                            inp_guia.on("keydown.enter", lambda evt, inv=inv, inp=inp_guia: _save_inv_extra(inv, guia=inp.value))
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            despachantes_list = get_despachantes()
                                            desp_opts = {"": "(otro)"}
                                            desp_opts.update({d["nombre"]: d["nombre"] for d in despachantes_list})
                                            desp_actual = inv.get("despachante", "") or ""
                                            if desp_actual and desp_actual not in desp_opts:
                                                desp_opts[desp_actual] = desp_actual
                                            desp_wrap = (
                                                ui.element("div")
                                                .classes("invoice-bdc-qsel w-40")
                                                .style(f"--qsel-color: {_color_despachante_hex(desp_actual)}")
                                            )
                                            with desp_wrap:
                                                ui.select(
                                                    desp_opts,
                                                    value=desp_actual or None,
                                                    on_change=lambda e, inv=inv, w=desp_wrap: _on_desp_change(e, inv, w),
                                                ).classes("w-full min-w-0").props("dense outlined")
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            def _fmt_importe_factura(val):
                                                if not val:
                                                    return ""
                                                try:
                                                    s = str(val).replace(" ", "").lstrip("$").strip()
                                                    if "," in s:
                                                        s = s.replace(".", "").replace(",", ".")
                                                    n = float(s)
                                                    return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                                                except (ValueError, TypeError):
                                                    return str(val).lstrip("$").strip()
                                            def _parse_imp(s):
                                                if not s:
                                                    return ""
                                                return str(s).replace(" ", "").lstrip("$").replace(".", "").replace(",", ".").strip()
                                            _imp_val = _fmt_importe_factura(inv.get("importe_factura", ""))
                                            inp_imp = ui.input(value=_imp_val).classes("w-32").props('dense outlined prefix="$"')
                                            inp_imp.on("blur", lambda evt, inv=inv, inp=inp_imp: _save_inv_extra(inv, importe_factura=_parse_imp(inp.value)))
                                            inp_imp.on("keydown.enter", lambda evt, inv=inv, inp=inp_imp: _save_inv_extra(inv, importe_factura=_parse_imp(inp.value)))
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            fc_disp = _norm_factura_courier(inv.get("factura_courier"))
                                            fc_wrap = (
                                                ui.element("div")
                                                .classes("invoice-bdc-qsel w-36")
                                                .style(f"--qsel-color: {_hex_factura_courier(fc_disp)}")
                                            )
                                            with fc_wrap:
                                                ui.select(
                                                    {"Impaga": "Impaga", "Pagada": "Pagada"},
                                                    value=fc_disp,
                                                    on_change=lambda e, inv=inv, w=fc_wrap: _on_fc_change(e, inv, w),
                                                ).classes("w-full min-w-0").props("dense outlined")
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            def _fmt_pa(val):
                                                if not val:
                                                    return ""
                                                try:
                                                    s = str(val).replace(" ", "").strip()
                                                    if s.lower().startswith("u$"):
                                                        s = s[2:].strip()
                                                    if "," in s:
                                                        s = s.replace(".", "").replace(",", ".")
                                                    n = float(s)
                                                    return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                                                except (ValueError, TypeError):
                                                    u = str(val).replace(" ", "").strip()
                                                    if u.lower().startswith("u$"):
                                                        u = u[2:].strip()
                                                    return u
                                            def _parse_pa(s):
                                                if not s:
                                                    return ""
                                                t = str(s).replace(" ", "").strip()
                                                if t.lower().startswith("u$"):
                                                    t = t[2:].strip()
                                                return t.replace(".", "").replace(",", ".").strip()
                                            _pa_val = _fmt_pa(inv.get("pa", ""))
                                            inp_pa = ui.input(value=_pa_val).classes("w-32").props('dense outlined prefix="u$"')
                                            inp_pa.on("blur", lambda evt, inv=inv, inp=inp_pa: _save_inv_extra(inv, pa=_parse_pa(inp.value)))
                                            inp_pa.on("keydown.enter", lambda evt, inv=inv, inp=inp_pa: _save_inv_extra(inv, pa=_parse_pa(inp.value)))
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            with ui.row().classes("justify-center w-full"):
                                                est_display = inv.get("estado", "") or "En USA"
                                                ui.select(estado_opts, value=est_display or "En USA", on_change=lambda e, inv=inv: _save_inv_extra(inv, estado=e.value)).classes("w-28").props("dense")
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-right"):
                                            amt_num = inv.get("amount_num", 0) or 0
                                            amt_str = f"{float(amt_num):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                                            ui.label(f"u$ {amt_str}")
                                        with ui.element("td").classes("px-2 py-1 border-b border-gray-100 text-center"):
                                            ui.label(inv.get("status", "—"))

        def _cargar_compras() -> None:
            qb_cust = get_user_qb_customer(user["id"])
            cust_id = qb_cust["id"] if qb_cust else None
            cust_name = qb_cust["name"] if qb_cust else None
            if not cust_id:
                header_data_ref.clear()
                header_card.clear()
                with header_card:
                    ui.label("⚠️ Seleccioná tu cliente en Configuración → QuickBooks → 'Soy el cliente'.").classes("text-warning p-4")
                return
            cust_detail, _ = fetch_qb_customer_detail(user["id"], cust_id)
            inv_result, err_inv = fetch_qb_invoices(user["id"], cust_id)
            if err_inv:
                header_data_ref.clear()
                header_card.clear()
                with header_card:
                    ui.label(f"Error al cargar facturas: {err_inv}").classes("text-negative p-4")
                    if "403" in str(err_inv) and ("3100" in str(err_inv) or "AuthorizationFailed" in str(err_inv)):
                        ui.label(
                            "Sugerencia (error 403/3100):\n"
                            "• Verificá que las credenciales (Client ID, Secret) sean de Producción en developer.intuit.com y que la app esté autorizada.\n"
                            "• Desvincular → Conectar cuenta para obtener nuevos tokens."
                        ).classes("text-sm text-gray-700 mt-2 p-3 bg-gray-100 rounded whitespace-pre-line")
                return
            invoices, overdue_total = inv_result
            open_balance = ""
            if cust_detail:
                bal = cust_detail.get("Balance") or cust_detail.get("BalanceWithJobs")
                if bal is not None:
                    try:
                        open_balance = f"{float(bal):,.2f}"
                    except (TypeError, ValueError):
                        open_balance = str(bal)
            if not open_balance and invoices:
                total_bal = sum(float(inv.get("balance", 0) or 0) for inv in invoices)
                open_balance = f"{total_bal:,.2f}"
            overdue = f"{overdue_total:,.2f}" if overdue_total else "0.00"
            qb_cust_display = (cust_detail.get("DisplayName") or cust_detail.get("FullyQualifiedName") or cust_detail.get("CompanyName") or "").strip() if cust_detail else ""
            header_data_ref["cust_name"] = qb_cust_display or cust_name or cust_id
            header_data_ref["open_balance"] = open_balance or "—"
            header_data_ref["overdue"] = overdue
            inv_list = [
                {"id": inv.get("id", ""), "txn_date": inv.get("txn_date", ""), "due_date": inv.get("due_date", ""), "tipo": inv.get("tipo", "Factura"), "doc": inv.get("doc", ""), "amount": inv.get("amount", ""), "amount_num": inv.get("amount_num", 0), "balance": inv.get("balance", 0), "status": inv.get("status", "")}
                for inv in invoices
            ]
            qb_ids = [inv["id"] for inv in inv_list if inv.get("id")]
            extras = get_invoice_extras(user["id"], qb_ids) if qb_ids else {}
            for inv in inv_list:
                qid = inv.get("id", "")
                ex = extras.get(qid, {})
                inv["guia"] = ex.get("guia") or ""
                inv["despachante"] = ex.get("despachante") or ""
                inv["importe_factura"] = ex.get("importe_factura") or ""
                inv["pa"] = ex.get("pa") or ""
                inv["factura_courier"] = ex.get("factura_courier") or "Impaga"
                est = ex.get("estado") or ""
                status_low = (inv.get("status") or "").lower()
                if status_low == "pagada":
                    inv["estado"] = est if est else "Recibida"
                    if not est:
                        upsert_invoice_extra(user["id"], qid, estado="Recibida")
                else:
                    inv["estado"] = est if est else "En USA"
                    if not est:
                        upsert_invoice_extra(user["id"], qid, estado="En USA")
            invoices_ref["data"] = inv_list
            _pintar_compras()

        ui.timer(0.3, _cargar_compras, once=True)
