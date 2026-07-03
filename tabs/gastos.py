"""
tabs/gastos.py — Gestión de documentos impositivos mensuales por sección.
Funciones exportadas: build_tab_gastos, procesar_archivo_con_gemini
"""
from __future__ import annotations

import base64
import json
import os
import re
import tempfile
import unicodedata
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

import requests
from nicegui import app, run, ui

from db import (
    delete_gastos_archivo,
    get_app_config,
    get_connection,
    get_cotizador_param,
    get_gastos_archivos,
    get_gastos_consolidado,
    get_gastos_prompt,
    insert_gastos_archivo,
    mark_gastos_procesado,
    save_gastos_consolidado,
    update_gastos_extraccion,
    upsert_gastos_prompt,
    user_can_access_tab,
)
from ml_api import get_ml_access_token, ml_get_orders, ml_get_user_id, ml_get_user_profile

_BLUE       = "#2A7AC7"
_BLUE_BG    = "#EEF6FD"
_HDR_BG     = "#EEF6FD"
_HDR_COLOR  = "#185FA5"
_HDR_BORDER = "#85B7EB"
_GREEN      = "#3B6D11"
_YELLOW     = "#E2A93B"
_GRAY       = "#9E9E9E"
_RED        = "#A32D2D"

FUENTES_CONSOLIDADO = {
    "arca":     {"label": "ARCA (VEPs)",           "icon": "ti-building-bank", "color": "#185FA5", "bg": "#E8F1FA", "border": "#85B7EB"},
    "fact":     {"label": "Facturas ML",           "icon": "ti-file-invoice",  "color": "#8B4513", "bg": "#FBF0E4", "border": "#D4A574"},
    "perc":     {"label": "Reportes Percepciones", "icon": "ti-receipt",       "color": "#5B2B8F", "bg": "#F0E4FB", "border": "#B48FE0"},
    "reten":    {"label": "Reportes Retenciones",  "icon": "ti-receipt-tax",   "color": "#8B0000", "bg": "#FDE4E4", "border": "#E08F8F"},
    "repo":     {"label": "Reportes ML",           "icon": "ti-report",        "color": "#0F5F2B", "bg": "#E4FBF0", "border": "#7FCFA0"},
    "analisis": {"label": "Análisis ML",           "icon": "ti-sparkles",      "color": "#7A5A0E", "bg": "#FBF8DC", "border": "#D4B860"},
    "calc":     {
        "label": "Cálculo interno", "icon": "ti-calculator",
        "color": "var(--color-text-secondary)", "bg": "var(--color-background-secondary)",
        "border": "var(--color-border-secondary)",
    },
    "ventas_db": {"label": "Ventas propias (BD)", "icon": "ti-database", "color": "#0E6B6B", "bg": "#E3F7F7", "border": "#7FCFCF"},
}


def render_fuente_badge(fuente_key: str, with_label: bool = False) -> str:
    """Badge HTML de una fuente de datos. Por default solo ícono (el label completo
    queda disponible como tooltip vía title) — con_label=True se usa en la leyenda."""
    f = FUENTES_CONSOLIDADO.get(fuente_key)
    if not f:
        return ""
    texto = f'<span>{f["label"]}</span>' if with_label else ""
    return (
        f'<span title="{f["label"]}" style="display:inline-flex;align-items:center;gap:5px;'
        f'padding:3px 8px;border-radius:4px;font-size:11px;font-weight:500;line-height:1;'
        f'color:{f["color"]};background:{f["bg"]};border:0.5px solid {f["border"]}">'
        f'<i class="ti {f["icon"]}" style="font-size:12px"></i>{texto}</span>'
    )


def render_fuente_badges(fuentes: Optional[list], with_label: bool = False) -> str:
    """Concatena badges para una línea alimentada por una o varias fuentes."""
    if not fuentes:
        return ""
    return "".join(render_fuente_badge(fk, with_label=with_label) for fk in fuentes)

_DOT = "display:inline-block;width:12px;height:12px;border-radius:9999px;flex-shrink:0;background:{}"

_PROMPT_PRE_STYLE = (
    "font-size:11px;line-height:1.5;white-space:pre-wrap;word-break:break-word;"
    "min-height:120px;max-height:none;height:auto;overflow-y:auto;"
    "padding:8px 10px;"
    "border:0.5px dashed #ccc;background:#f9f9f9;border-radius:4px;"
    "margin:0;font-family:inherit"
)


def _escape_prompt_html(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

_SECCIONES: List[tuple] = [
    ("facturas_ml",  "Facturas MercadoLibre",  ".pdf",  True,  "ti-file-invoice"),
    ("retenciones",  "Retenciones",             ".xlsx", True,  "ti-file-spreadsheet"),
    ("percepciones", "Percepciones",            ".xlsx", True,  "ti-file-spreadsheet"),
    ("pagos_arca",   "Pagos ARCA",              ".pdf",  True,  "ti-file-invoice"),
    ("reportes_ml",  "Reportes MercadoLibre",   ".xlsx", True,  "ti-file-spreadsheet"),
    ("analisis_ml",  "Análisis MercadoLibre",   ".pdf",  False, "ti-report-analytics"),
]

_MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

_BASE_PATH = Path(__file__).parent.parent / "gastos"

_PROMPTS_DEFAULT: Dict[str, str] = {
    "facturas_ml": (
        "Analizá esta factura de MercadoLibre y extraé en JSON puro (sin markdown ni bloques de código). "
        "Campos obligatorios: tipo_documento (Factura A/B/C), emisor (razón social), cuit_emisor, "
        "receptor (razón social), cuit_receptor, fecha (DD/MM/AAAA), punto_venta, nro_comprobante, "
        "lineas_antes_subtotal (lista con TODOS los conceptos que aparezcan ANTES del subtotal — "
        "cargos, bonificaciones, servicios, comisiones, etc. — cada uno con concepto y monto numérico; "
        "si una bonificación tiene signo negativo en el documento, devolvé el monto negativo; "
        "si no hay ningún concepto antes del subtotal, devolvé array vacío []), "
        "subtotal (número), "
        "lineas_intermedias (lista con TODOS los conceptos entre Subtotal y Total — "
        "IVA 21%, IVA 10.5%, Percepciones IIBB, Percepción IVA, otros impuestos, descuentos, etc. — "
        "extraé TODAS sin omitir ninguna, cada una con nombre exacto y monto numérico), "
        "total (número), cae, cae_vto (DD/MM/AAAA). "
        "CRÍTICO: Es OBLIGATORIO extraer TODAS las líneas del documento, tanto antes como después del subtotal. "
        "Las líneas antes del subtotal pueden incluir cargos por uso de plataforma, bonificaciones, servicios, etc. "
        "Las líneas entre Subtotal y Total pueden ser 1 sola o hasta 15 líneas distintas, incluyendo: "
        "- Todas las percepciones de IIBB de distintas provincias (Corrientes, Catamarca, CABA, Tucumán, "
        "La Pampa, Neuquén, Buenos Aires, etc.) "
        "- Percepciones IVA de distintas jurisdicciones "
        "- IVA 21% e IVA 10,5% "
        "- Retenciones si las hubiere "
        "- Cualquier otro concepto que aparezca listado entre Subtotal y Total. "
        "NO OMITAS NINGUNA. Si tenés dudas sobre si una línea corresponde, INCLUÍLA. "
        "Es preferible incluir de más que omitir alguna. "
        "Extraé el nombre EXACTAMENTE como aparece en el PDF, sin abreviar ni traducir. "
        'Formato esperado: {"tipo_documento":"Factura A","emisor":"MercadoLibre S.R.L.",'
        '"cuit_emisor":"...","receptor":"...","cuit_receptor":"...","fecha":"DD/MM/AAAA",'
        '"punto_venta":"...","nro_comprobante":"...",'
        '"lineas_antes_subtotal":[{"concepto":"Cargos por uso de la plataforma Mercado Libre","monto":23976326.91},'
        '{"concepto":"Bonificaciones uso plataforma Mercado Libre","monto":-1171322.59}],'
        '"subtotal":0.0,'
        '"lineas_intermedias":[{"concepto":"IVA 21%","monto":0.0},{"concepto":"PERCEPCION IIBB CORRIENTES","monto":0.0}],'
        '"total":0.0,"cae":"...","cae_vto":"DD/MM/AAAA"}'
    ),
    "retenciones": (
        "Analizá este Excel de retenciones. Extraé en JSON puro (sin markdown ni bloques de código). "
        "De la CABECERA del Excel extraé: "
        "usuario (ej: 'NORTHTECHNOLOGY'), "
        "impuesto (ej: 'Impuesto a los IIBB Corrientes'), "
        "fecha_desde (del campo 'Intervalo de fechas consultadas', formato DD/MM/AAAA), "
        "fecha_hasta (del mismo campo, formato DD/MM/AAAA). "
        "De la TABLA de detalle, mirá la columna 'Alícuota': "
        "si el valor es el MISMO en todas las filas, extraé ese valor único. "
        "Si varía, extraé el rango como 'min-max'. "
        "Del BLOQUE DE TOTALES extraé: "
        "base_imponible (número), "
        "importe_retenido (número), "
        "importe_devuelto (número). "
        "IMPORTANTE: NO extraer las filas individuales de la tabla — solo los totales agregados y la alícuota. "
        "El JSON debe ser ESTRICTAMENTE válido. Usá SOLO comillas dobles. "
        "Responde ÚNICAMENTE con el JSON. "
        'Ejemplo: {"usuario":"NORTHTECHNOLOGY","impuesto":"Impuesto a los IIBB Corrientes",'
        '"fecha_desde":"01/04/2026","fecha_hasta":"01/05/2026","alicuota":"2,00 %",'
        '"base_imponible":869973.24,"importe_retenido":17399.46,"importe_devuelto":3777.96}'
    ),
    "percepciones": (
        "Analizá este Excel de percepciones. Extraé en JSON puro (sin markdown ni bloques de código). "
        "De la CABECERA del Excel extraé: "
        "usuario (ej: 'NORTHTECHNOLOGY'), "
        "impuesto (ej: 'Impuesto a los IIBB Corrientes - Percepciones'), "
        "condicion_fiscal (ej: 'Responsable Inscripto', tal como aparece en la cabecera), "
        "fecha_desde (del campo 'Intervalo de fechas consultadas', formato DD/MM/AAAA), "
        "fecha_hasta (del mismo campo, formato DD/MM/AAAA). "
        "De la TABLA de detalle, mirá la columna 'Alícuota': "
        "si el valor es el MISMO en todas las filas, extraé ese valor único. "
        "Si varía, extraé el rango como 'min-max'. "
        "CRÍTICO: base_imponible es la SUMA de TODOS los valores de la columna 'Base imponible' de la tabla "
        "de detalle. NO uses un valor individual — sumá todas las filas. Lo mismo para monto_percibido: "
        "es la SUMA de TODOS los valores de la columna 'Monto percibido' (o 'Importe percibido', "
        "según el nombre que aparezca en el Excel). "
        "IMPORTANTE: NO extraer las filas individuales de la tabla — solo los agregados sumados y la alícuota. "
        "El JSON debe ser ESTRICTAMENTE válido. Usá SOLO comillas dobles. "
        "Responde ÚNICAMENTE con el JSON. "
        'Ejemplo: {"usuario":"NORTHTECHNOLOGY","impuesto":"Impuesto a los IIBB Corrientes",'
        '"condicion_fiscal":"Responsable Inscripto",'
        '"fecha_desde":"01/04/2026","fecha_hasta":"01/05/2026","alicuota":"2,00 %",'
        '"base_imponible":869973.24,"monto_percibido":17399.46}'
    ),
    "pagos_arca": (
        "Analizá este comprobante de pago ARCA/AFIP. Identificá el tipo de documento y extraé en JSON puro "
        "(sin markdown ni bloques de código) TODOS los conceptos financieros con nombre EXACTO y monto numérico. "
        "Campos comunes (siempre extraer): "
        "tipo (IVA | SIFERE Convenio Multilateral | texto descriptivo del documento), "
        "periodo (AAAA-MM), "
        "numero_vep. "
        "Si el documento es SIFERE Convenio Multilateral: "
        "lineas_convenio_multilateral (lista con TODAS las líneas CM que aparezcan antes del importe total — "
        "cada una con jurisdiccion (nombre de la provincia sin el prefijo CM), "
        "codigo (número entre paréntesis) y monto numérico; si no hay, devolvé []). "
        "Si el documento es IVA: "
        "determinacion_del_impuesto (lista de {concepto, monto} con TODOS los ítems de la sección "
        "Determinación del Impuesto, en el orden en que aparecen), "
        "determinacion_posicion_mensual (lista de {concepto, monto} con TODOS los ítems de la sección "
        "Determinación de la Posición Mensual, en el orden en que aparecen). "
        "En cualquier caso, si aparece explícitamente en el documento: "
        "importe_total_a_pagar (número). "
        "CRÍTICO: extraé los nombres EXACTAMENTE como aparecen en el documento, sin abreviar ni traducir. "
        "No omitas ningún concepto. Si un grupo no aparece en el documento, omití esa clave. "
        "El JSON debe ser ESTRICTAMENTE válido — solo comillas dobles, sin trailing commas ni comentarios. "
        "Responde ÚNICAMENTE con el JSON. "
        'Ejemplo SIFERE: {"tipo":"SIFERE Convenio Multilateral","periodo":"2026-04",'
        '"numero_vep":"1629093052",'
        '"lineas_convenio_multilateral":['
        '{"jurisdiccion":"PCIA BS AS","codigo":"5802","monto":2244403.49},'
        '{"jurisdiccion":"CHACO","codigo":"5806","monto":7087.99}],'
        '"importe_total_a_pagar":3383401.10} '
        'Ejemplo IVA: {"tipo":"IVA","periodo":"2026-04","numero_vep":"1172911737",'
        '"determinacion_del_impuesto":['
        '{"concepto":"Total del Débito Fiscal del Período","monto":19292005.21},'
        '{"concepto":"Total del Crédito Fiscal del Período","monto":21863249.64},'
        '{"concepto":"Saldo técnico a favor del contribuyente del período anterior","monto":12105024.60},'
        '{"concepto":"Saldo técnico a favor del contribuyente","monto":14676269.03}],'
        '"determinacion_posicion_mensual":['
        '{"concepto":"Saldo técnico a favor de ARCA","monto":0.00},'
        '{"concepto":"Saldo técnico a favor del contribuyente","monto":14676269.03},'
        '{"concepto":"Saldo a favor de libre disponibilidad del período anterior neto de usos","monto":1727.75},'
        '{"concepto":"Total de retenciones, percepciones y pagos a cuenta neto de restituciones","monto":1534.81},'
        '{"concepto":"Saldo de libre disponibilidad a favor del contribuyente del período","monto":3262.56}],'
        '"importe_total_a_pagar":0.0}'
    ),
    "reportes_ml": (
        "Analizá este reporte de operaciones de MercadoLibre. Identificá las columnas principales y "
        "resumí las primeras 5 filas. Devolvé JSON puro con keys: columnas (lista), resumen_filas (lista de dicts)."
    ),
    "analisis_ml": (
        "Extraé TODO el texto del documento tal cual aparece. No lo interpretes, "
        "no lo resumas. Solo transcribí el contenido completo del PDF, respetando "
        "el orden y la estructura visual (párrafos, tablas, listas). Devolvé el "
        "texto en formato plano."
    ),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _require_login() -> Optional[Dict[str, Any]]:
    user = app.storage.user.get("user")
    if not user:
        ui.notify("Debes iniciar sesión", color="negative")
    return user


def _fmt_size(n: int) -> str:
    if n >= 1_048_576:
        return f"{n / 1_048_576:.1f} MB"
    if n >= 1024:
        return f"{n / 1024:.0f} KB"
    return f"{n} B"


def _semaforo_color(archivos: list) -> str:
    if not archivos:
        return _GRAY
    statuses = [f.get("extraction_status", "pendiente") for f in archivos]
    if all(s == "procesado" for s in statuses):
        return _GREEN
    return _YELLOW


def _badge_html(status: str) -> str:
    _STYLES = {
        "pendiente": f"background:#FBF1DC;color:#7A5A0E;border:1px solid {_YELLOW}",
        "procesado": f"background:#EAF3DE;color:#27500A;border:1px solid {_GREEN}",
        "error":     f"background:#FBE9E9;color:#7A1414;border:1px solid {_RED}",
    }
    _LABELS = {"pendiente": "Pendiente", "procesado": "Procesado", "error": "Error"}
    sty = _STYLES.get(status, _STYLES["pendiente"])
    lbl = _LABELS.get(status, status)
    return (
        f'<span style="font-size:9px;padding:1px 5px;border-radius:3px;'
        f'white-space:nowrap;{sty}">{lbl}</span>'
    )


def _pdf_first_page_b64(path: Path) -> Optional[str]:
    try:
        import fitz
        doc = fitz.open(str(path))
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
        data = pix.tobytes("png")
        doc.close()
        return base64.b64encode(data).decode()
    except Exception:
        return None


def _pdf_all_pages_b64(path: Path) -> List[str]:
    try:
        import fitz
        doc = fitz.open(str(path))
        pages = [
            base64.b64encode(page.get_pixmap(matrix=fitz.Matrix(1.5, 1.5)).tobytes("png")).decode()
            for page in doc
        ]
        doc.close()
        return pages
    except Exception:
        return []


_EXCEL_PREVIEW_HEADERS_DETALLE = (
    "número de venta", "n° de venta", "referencia externa", "detalle de movimientos",
    "fecha del cargo",
)


def _cortar_en_detalle(filas: list) -> int:
    """Índice de la primera fila que pertenece a la tabla de detalle (headers de columnas o de
    agrupamiento tipo "Información sobre..."), para no incluirla ni lo que sigue."""
    for i, linea in enumerate(filas):
        celdas = linea.split("\t")
        primera_celda = celdas[0].strip().lower()
        if any(primera_celda.startswith(h) for h in _EXCEL_PREVIEW_HEADERS_DETALLE):
            return i
        info_count = sum(1 for c in celdas if "información sobre" in c.strip().lower())
        if info_count >= 2:
            return i
    return len(filas)


def _excel_preview_html(path: Path) -> str:
    """Preview acotado: solo metadatos + totales, sin la tabla de detalle completa."""
    try:
        filas = [f for f in leer_excel_completo(path) if not f.startswith("=== HOJA:")]
        if not filas:
            return "<p style='font-size:11px;color:#9e9e9e'>Excel vacío</p>"

        corte = _cortar_en_detalle(filas)
        preview = []
        for linea in filas[:corte]:
            celdas = linea.split("\t")
            # Filas banner/párrafo (una sola celda con texto largo, p.ej. leyendas legales) → ruido, saltar
            no_vacias = [c for c in celdas if c.strip()]
            if len(no_vacias) <= 1 and (not no_vacias or len(no_vacias[0]) > 40):
                continue
            preview.append(linea)
            if len(preview) >= 12:
                break

        titulo = path.stem.replace("_", " ").replace("-", " ").strip()
        titulo_html = (
            '<div style="font-size:12px;font-weight:700;color:#333;'
            'white-space:normal;word-break:break-word;line-height:1.3;padding:8px 10px">'
            f"{titulo}</div>"
        )

        body = ""
        for linea in preview:
            tds = "".join(
                f'<td style="border:1px solid #e0e0e0;padding:2px 6px;white-space:nowrap">{c}</td>'
                for c in linea.split("\t")
            )
            body += f"<tr>{tds}</tr>"
        tabla_html = (
            '<table style="border-collapse:collapse;font-size:11px;width:100%">'
            f"<tbody>{body}</tbody></table>"
        )
        return titulo_html + tabla_html
    except Exception as exc:
        return f"<p style='font-size:11px;color:#a32d2d'>Error al leer Excel: {exc}</p>"


def leer_excel_completo(path: Path) -> list[str]:
    """Lee TODAS las hojas de un Excel y devuelve las filas como líneas tab-separated.

    wb.active solo apunta a la hoja marcada como activa en el workbook, y los
    reportes de retenciones de MercadoPago suelen tener los datos reales en
    una hoja distinta a la activa (que a veces solo trae el título).
    """
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=False)
    filas_texto = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        filas_texto.append(f"=== HOJA: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            filas_texto.append("\t".join(str(c) if c is not None else "" for c in row))
    wb.close()
    return filas_texto


def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _concepto_norm(s: str) -> str:
    """Normaliza un concepto de factura para matching por substring: sin acentos, minúsculas,
    sin puntuación — así "I.V.A. INSC.21,00%" matchea "iva" igual que "IVA 21%"."""
    s = _strip_accents(str(s or "").lower())
    return "".join(c for c in s if c.isalnum())


_MESES_ABBR_ARCHIVO = {"ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"}


def _extraer_impuesto_percepciones(filename: str) -> Optional[str]:
    """Deduce 'Percepciones IIBB {provincia}' del NOMBRE del archivo.

    El Excel de percepciones no trae el nombre del impuesto explícito en ninguna celda
    (a diferencia de Retenciones, que sí lo trae) — el nombre de archivo es la única fuente confiable.
    """
    palabras = Path(filename).stem.split("-")
    idx = next(
        (i for i, p in enumerate(palabras) if _strip_accents(p.lower()) == "percepcion"),
        None,
    )
    if idx is None or idx + 2 >= len(palabras):
        return None
    provincia_palabras = []
    for p in palabras[idx + 2:]:
        pl = _strip_accents(p.lower())
        if pl in _MESES_ABBR_ARCHIVO or re.fullmatch(r"20\d{2}", p):
            break
        provincia_palabras.append(p)
    if not provincia_palabras:
        return None
    return f"Percepciones IIBB {' '.join(provincia_palabras)}"


def _calcular_totales_percepciones(filas: list) -> Optional[dict]:
    """Suma en Python (100% preciso, sin límite de filas ni de tokens) TODA la tabla de
    detalle de un Excel de percepciones: base_imponible y monto_percibido."""
    def _norm(s: str) -> str:
        return _strip_accents(s.strip().lower())

    idx_base = idx_alic = idx_monto = idx_fecha = header_idx = None
    for i, linea in enumerate(filas):
        celdas_norm = [_norm(c) for c in linea.split("\t")]
        if "base imponible" in celdas_norm:
            idx_base = celdas_norm.index("base imponible")
            if "alicuota" in celdas_norm:
                idx_alic = celdas_norm.index("alicuota")
            for cand in ("monto percibido", "importe percibido"):
                if cand in celdas_norm:
                    idx_monto = celdas_norm.index(cand)
                    break
            for cand in ("fecha del cargo", "fecha de la venta", "fecha de operacion"):
                if cand in celdas_norm:
                    idx_fecha = celdas_norm.index(cand)
                    break
            header_idx = i
            break

    if header_idx is None or idx_base is None or idx_monto is None:
        return None

    total_base = total_monto = 0.0
    alicuota_val = None
    fecha_min = fecha_max = None
    n_filas = 0
    for linea in filas[header_idx + 1:]:
        celdas = linea.split("\t")
        if len(celdas) <= max(idx_base, idx_monto):
            continue
        try:
            base_v = float(celdas[idx_base])
            monto_v = float(celdas[idx_monto])
        except ValueError:
            continue
        total_base += base_v
        total_monto += monto_v
        n_filas += 1
        if alicuota_val is None and idx_alic is not None and idx_alic < len(celdas):
            try:
                alicuota_val = float(celdas[idx_alic])
            except ValueError:
                pass
        if idx_fecha is not None and idx_fecha < len(celdas):
            try:
                dt = datetime.strptime(celdas[idx_fecha].strip().split(" ")[0], "%Y-%m-%d")
                if fecha_min is None or dt < fecha_min:
                    fecha_min = dt
                if fecha_max is None or dt > fecha_max:
                    fecha_max = dt
            except ValueError:
                pass

    if n_filas == 0:
        return None

    alicuota_str = None
    if alicuota_val is not None:
        alicuota_str = f"{alicuota_val * 100:.2f}".replace(".", ",") + " %"

    return {
        "filas_tabla": n_filas,
        "base_imponible": round(total_base, 2),
        "monto_percibido": round(total_monto, 2),
        "alicuota": alicuota_str,
        "fecha_desde": fecha_min.strftime("%d/%m/%Y") if fecha_min else None,
        "fecha_hasta": fecha_max.strftime("%d/%m/%Y") if fecha_max else None,
    }


def _es_notas_credito(filename: str) -> bool:
    """Detecta el 'Reporte notas de crédito' de MercadoLibre por el nombre del archivo
    (p.ej. 'Reporte_Notas_Credito_MercadoLibre_Abr2026.xlsx'), sin importar acentos/mayúsculas."""
    norm = _strip_accents(filename.lower())
    return "notas" in norm and "credito" in norm


def _es_notas_debito(filename: str) -> bool:
    """Detecta el 'Reporte notas de débito' de MercadoLibre por el nombre del archivo
    (p.ej. 'Reporte_Notas_Debito_EnviosFlex_Abr2026.xlsx'), sin importar acentos/mayúsculas."""
    norm = _strip_accents(filename.lower())
    return "notas" in norm and "debito" in norm


def _es_notas_credito_ml(filename: str) -> bool:
    """Detecta específicamente el 'Reporte notas de crédito' de MercadoLibre (cargos —
    p.ej. 'Reporte_Notas_Credito_MercadoLibre_Abr2026.xlsx'), distinto del de Envíos Flex
    (bonificaciones), que tiene otra estructura de columnas ('Valor del cargo' en vez de
    'Valor de la bonificación')."""
    norm = _strip_accents(filename.lower())
    return _es_notas_credito(filename) and "flex" not in norm


def _calcular_notas_credito_ml(filas: list) -> Optional[dict]:
    """Calcula en Python (sin Gemini) los 3 campos de 'Notas de Crédito MercadoLibre':
    fecha_desde/fecha_hasta (rango de fechas de la tabla de detalle) y total (SUMA en valor
    absoluto de la columna 'Valor del cargo'). Al igual que en los reportes de Envíos Flex, los
    valores vienen en NEGATIVO (son anulaciones de cargos previos) — se usa abs() porque 'Total'
    representa el monto total acreditado, no un saldo con signo."""
    def _norm(s: str) -> str:
        return _strip_accents(s.strip().lower())

    idx_fecha = idx_valor = header_idx = None
    for i, linea in enumerate(filas):
        celdas_norm = [_norm(c) for c in linea.split("\t")]
        cand_fecha = next((j for j, c in enumerate(celdas_norm) if "fecha" in c), None)
        cand_valor = next((j for j, c in enumerate(celdas_norm) if c == "valor del cargo"), None)
        if cand_fecha is not None and cand_valor is not None:
            idx_fecha, idx_valor, header_idx = cand_fecha, cand_valor, i
            break

    if header_idx is None:
        return None

    fechas = []
    total = 0.0
    n_filas = 0
    for linea in filas[header_idx + 1:]:
        celdas = linea.split("\t")
        if len(celdas) <= max(idx_fecha, idx_valor):
            continue
        try:
            v = float(celdas[idx_valor])
        except ValueError:
            continue
        total += abs(v)
        n_filas += 1
        if celdas[idx_fecha].strip():
            try:
                fechas.append(datetime.strptime(celdas[idx_fecha].strip().split(" ")[0], "%Y-%m-%d"))
            except ValueError:
                pass

    if n_filas == 0:
        return None

    return {
        "fecha_desde": fechas and min(fechas).strftime("%d/%m/%Y") or None,
        "fecha_hasta": fechas and max(fechas).strftime("%d/%m/%Y") or None,
        "total": round(total, 2),
    }


def analizar_facturacion_ml(path: Path) -> Optional[dict]:
    """Calcula en Python (100% determinista, sin Gemini) el análisis completo del
    'Reporte Facturación MercadoLibre': lee la hoja REPORT del Excel con openpyxl
    (fila 1 = fecha de actualización, filas 7-8 = headers, fila 9+ = detalle) y devuelve
    totales generales, margen neto estimado, desglose por tipo de cargo, ventas por día
    de semana, tops de provincias/categorías/productos/clientes, concentración y
    percepciones — sin límite de filas ni de tokens."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=False)
    sheet_name = next(
        (s for s in wb.sheetnames if _strip_accents(s.strip().lower()) == "report"), None
    )
    if sheet_name is None:
        wb.close()
        return None
    ws = wb[sheet_name]

    def _norm(s) -> str:
        return _strip_accents(str(s or "").strip().lower())

    def _num(v) -> float:
        if v is None or v == "":
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def _fecha(v) -> Optional[datetime]:
        if v is None or v == "":
            return None
        if isinstance(v, datetime):
            return v
        try:
            return datetime.strptime(str(v).strip().split(" ")[0], "%Y-%m-%d")
        except ValueError:
            return None

    def _col(row, idx1: int):
        j = idx1 - 1
        return row[j] if j < len(row) else None

    _row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    fecha_actualizacion = None
    if _row1:
        _v = _col(_row1, 3)
        _fv = _fecha(_v)
        fecha_actualizacion = _fv.strftime("%d/%m/%Y") if _fv else (str(_v).strip() if _v else None)

    filas = [
        row for row in ws.iter_rows(min_row=9, values_only=True)
        if not all(c is None for c in row)
    ]
    wb.close()
    if not filas:
        return None

    # --- ventas únicas (dedup por Número de venta) ---
    ventas: Dict[str, dict] = {}
    for row in filas:
        nro = _col(row, 13)
        if nro is None or str(nro).strip() == "":
            continue
        nro = str(nro).strip()
        v = ventas.setdefault(nro, {
            "total_venta": None, "fecha_venta": None, "cliente": None,
            "provincia": None, "categoria": None, "titulo": None,
            "cantidad_vendida": None,
        })
        _tv = _num(_col(row, 23))
        if _tv and not v["total_venta"]:
            v["total_venta"] = _tv
        _fv = _fecha(_col(row, 15))
        if _fv and not v["fecha_venta"]:
            v["fecha_venta"] = _fv
        for campo, idx in (("cliente", 18), ("provincia", 19), ("categoria", 31), ("titulo", 29)):
            _val = _col(row, idx)
            if _val and not v[campo]:
                v[campo] = str(_val).strip()
        _cant = _num(_col(row, 20))
        if _cant and not v["cantidad_vendida"]:
            v["cantidad_vendida"] = _cant

    fechas_venta = [v["fecha_venta"] for v in ventas.values() if v["fecha_venta"]]
    fecha_desde = min(fechas_venta).strftime("%d/%m/%Y") if fechas_venta else None
    fecha_hasta = max(fechas_venta).strftime("%d/%m/%Y") if fechas_venta else None

    total_ingresos = round(sum(v["total_venta"] or 0.0 for v in ventas.values()), 2)
    cantidad_operaciones = len(ventas)
    ticket_promedio = round(total_ingresos / cantidad_operaciones, 2) if cantidad_operaciones else 0.0

    # --- totales generales (sobre TODAS las filas, sin dedup) ---
    suma_cargo = 0.0
    suma_descuento_total = 0.0
    suma_descuento_pos = 0.0
    cantidad_descuentos = 0
    envios_pagados_comprador = 0.0
    comisiones_venta_monto = 0.0
    envios_neto_monto = 0.0
    impuestos_monto = 0.0
    detalle_agg: Dict[str, dict] = defaultdict(lambda: {"total": 0.0, "cantidad": 0})
    percepciones_agg: Dict[str, dict] = defaultdict(lambda: {"monto": 0.0, "cantidad": 0})

    _SET_COMISIONES = {_norm(s) for s in (
        "Cargo por vender", "Anulación del cargo por vender",
        "Costo por unidad vendida", "Costo por ofrecer cuotas",
        "Cargo por devolución", "Anulación del cargo por devolución",
        "Cargo por mantenimiento de Mi página",
    )}
    _SET_ENVIOS = {_norm(s) for s in (
        "Cargo por envíos de Mercado Libre",
        "Anulación del cargo por envíos de Mercado Libre",
    )}

    for row in filas:
        detalle_raw = _col(row, 4)
        detalle = str(detalle_raw).strip() if detalle_raw else ""
        dn = _norm(detalle)
        valor_cargo = _num(_col(row, 8))
        valor_descuento = _num(_col(row, 11))

        suma_cargo += valor_cargo
        suma_descuento_total += valor_descuento
        if valor_descuento > 0:
            suma_descuento_pos += valor_descuento
            cantidad_descuentos += 1
        envios_pagados_comprador += _num(_col(row, 26))

        if dn in _SET_COMISIONES:
            comisiones_venta_monto += valor_cargo
        if dn in _SET_ENVIOS:
            envios_neto_monto += valor_cargo
        if "percep" in dn:
            impuestos_monto += valor_cargo
            percepciones_agg[detalle]["monto"] += valor_cargo
            percepciones_agg[detalle]["cantidad"] += 1
            continue
        if detalle:
            detalle_agg[detalle]["total"] += valor_cargo
            detalle_agg[detalle]["cantidad"] += 1

    facturacion_bruta = round(suma_cargo + suma_descuento_total, 2)
    facturacion_neta = round(suma_cargo, 2)
    total_descuentos = round(suma_descuento_pos, 2)

    base = total_ingresos
    comisiones_pct = round(comisiones_venta_monto / base * 100, 2) if base else 0.0
    envios_pct = round(envios_neto_monto / base * 100, 2) if base else 0.0
    impuestos_pct = round(impuestos_monto / base * 100, 2) if base else 0.0
    margen_neto_pct = round(100 - comisiones_pct - envios_pct - impuestos_pct, 2)

    # --- desglose por tipo de cargo (empareja "X" con "Anulación del X") ---
    bases: Dict[str, tuple] = {}
    anulaciones: Dict[str, tuple] = {}
    for label, info in detalle_agg.items():
        n = _norm(label)
        resto = None
        if n.startswith("anulacion del "):
            resto = n[len("anulacion del "):]
        elif n.startswith("anulacion de "):
            resto = n[len("anulacion de "):]
        if resto is not None:
            anulaciones[resto] = (label, info["total"], info["cantidad"])
        else:
            bases[n] = (label, info["total"], info["cantidad"])

    desglose_cargos = {}
    for n, (label, total, cantidad) in bases.items():
        entry = {
            "total": round(total, 2), "cantidad": cantidad,
            "anulacion_monto": 0.0, "anulacion_cantidad": 0, "proporcion_anulaciones": 0.0,
        }
        if n in anulaciones:
            anu_label, anu_total, anu_cant = anulaciones[n]
            entry["anulacion_monto"] = round(anu_total, 2)
            entry["anulacion_cantidad"] = anu_cant
            entry["proporcion_anulaciones"] = round(anu_cant / cantidad * 100, 2) if cantidad else 0.0
        desglose_cargos[label] = entry
    for n, (label, total, cantidad) in anulaciones.items():
        if n not in bases:
            desglose_cargos[label] = {
                "total": round(total, 2), "cantidad": cantidad,
                "anulacion_monto": 0.0, "anulacion_cantidad": 0, "proporcion_anulaciones": 0.0,
            }

    # --- ventas por día de la semana ---
    _DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
    por_dia = {d: {"monto": 0.0, "cantidad": 0} for d in _DIAS}
    for v in ventas.values():
        if not v["fecha_venta"]:
            continue
        dia = _DIAS[v["fecha_venta"].weekday()]
        por_dia[dia]["monto"] += v["total_venta"] or 0.0
        por_dia[dia]["cantidad"] += 1
    ventas_por_dia = {}
    for d in _DIAS:
        m = round(por_dia[d]["monto"], 2)
        ventas_por_dia[d] = {
            "monto": m, "cantidad_ventas": por_dia[d]["cantidad"],
            "porcentaje": round(m / total_ingresos * 100, 2) if total_ingresos else 0.0,
        }

    # --- provincias (normalizadas) ---
    def _norm_provincia(p) -> str:
        if not p:
            return "Sin Provincia"
        n = _strip_accents(p.strip().lower())
        if n == "buenos aires":
            return "Buenos Aires"
        if n in ("ciudad autonoma de buenos aires", "ciudad autonoma buenos aires", "caba"):
            return "CABA"
        return p.strip().title()

    prov_agg: Dict[str, dict] = defaultdict(lambda: {"monto": 0.0, "cantidad": 0})
    for v in ventas.values():
        p = _norm_provincia(v["provincia"])
        prov_agg[p]["monto"] += v["total_venta"] or 0.0
        prov_agg[p]["cantidad"] += 1
    prov_ordenadas = sorted(prov_agg.items(), key=lambda kv: kv[1]["monto"], reverse=True)
    top_provincias = [
        {
            "nombre": nombre, "monto": round(info["monto"], 2), "cantidad_ventas": info["cantidad"],
            "porcentaje": round(info["monto"] / total_ingresos * 100, 2) if total_ingresos else 0.0,
        }
        for nombre, info in prov_ordenadas[:5]
    ]

    # --- provincias — TODAS, sin normalizar (para agrupar aguas abajo con
    # normalizar_jurisdiccion_percepcion en la sección "Facturación por Provincia" del consolidado) ---
    prov_raw_agg: Dict[str, dict] = defaultdict(lambda: {"monto": 0.0, "cantidad": 0})
    for v in ventas.values():
        p = (v["provincia"] or "Sin Provincia").strip()
        prov_raw_agg[p]["monto"] += v["total_venta"] or 0.0
        prov_raw_agg[p]["cantidad"] += 1
    provincias_completas = [
        {"nombre": nombre, "monto": round(info["monto"], 2), "cantidad_ventas": info["cantidad"]}
        for nombre, info in sorted(prov_raw_agg.items(), key=lambda kv: kv[1]["monto"], reverse=True)
    ]

    # --- categorías ---
    cat_agg: Dict[str, dict] = defaultdict(lambda: {"monto": 0.0, "cantidad": 0})
    for v in ventas.values():
        c = v["categoria"] or "Sin Categoría"
        cat_agg[c]["monto"] += v["total_venta"] or 0.0
        cat_agg[c]["cantidad"] += 1
    cat_por_monto = sorted(cat_agg.items(), key=lambda kv: kv[1]["monto"], reverse=True)
    cat_por_cantidad = sorted(cat_agg.items(), key=lambda kv: kv[1]["cantidad"], reverse=True)
    top_categorias = [
        {
            "nombre": nombre, "monto": round(info["monto"], 2), "cantidad_ventas": info["cantidad"],
            "porcentaje": round(info["monto"] / total_ingresos * 100, 2) if total_ingresos else 0.0,
        }
        for nombre, info in cat_por_monto[:10]
    ]

    # --- productos ---
    prod_agg: Dict[str, dict] = defaultdict(lambda: {"monto": 0.0, "unidades": 0.0})
    for v in ventas.values():
        t = v["titulo"] or "Sin Título"
        prod_agg[t]["monto"] += v["total_venta"] or 0.0
        prod_agg[t]["unidades"] += v["cantidad_vendida"] or 0.0
    prod_ordenados = sorted(prod_agg.items(), key=lambda kv: kv[1]["monto"], reverse=True)
    top_productos = [
        {
            "nombre": nombre, "monto": round(info["monto"], 2), "unidades": round(info["unidades"], 2),
            "porcentaje": round(info["monto"] / total_ingresos * 100, 2) if total_ingresos else 0.0,
        }
        for nombre, info in prod_ordenados[:10]
    ]

    # --- clientes ---
    cli_agg: Dict[str, dict] = defaultdict(lambda: {"monto": 0.0, "cantidad": 0})
    for v in ventas.values():
        c = v["cliente"] or "Sin Cliente"
        cli_agg[c]["monto"] += v["total_venta"] or 0.0
        cli_agg[c]["cantidad"] += 1
    cli_por_monto = sorted(cli_agg.items(), key=lambda kv: kv[1]["monto"], reverse=True)
    cli_por_cantidad = sorted(cli_agg.items(), key=lambda kv: kv[1]["cantidad"], reverse=True)
    top_clientes_facturacion = [
        {
            "nombre": nombre, "monto": round(info["monto"], 2), "cantidad_ventas": info["cantidad"],
            "porcentaje": round(info["monto"] / total_ingresos * 100, 2) if total_ingresos else 0.0,
        }
        for nombre, info in cli_por_monto[:10]
    ]
    top_clientes_frecuentes = [
        {"nombre": nombre, "cantidad_ventas": info["cantidad"], "monto": round(info["monto"], 2)}
        for nombre, info in cli_por_cantidad[:10]
    ]

    # --- concentración ---
    total_categorias_unicas = len(cat_agg)
    total_clientes_unicos = len(cli_agg)
    concentracion_top10_categorias_facturacion = round(
        sum(info["monto"] for _, info in cat_por_monto[:10]) / total_ingresos * 100, 2
    ) if total_ingresos else 0.0
    concentracion_top10_categorias_cantidad = round(
        sum(info["cantidad"] for _, info in cat_por_cantidad[:10]) / cantidad_operaciones * 100, 2
    ) if cantidad_operaciones else 0.0
    concentracion_top10_clientes_facturacion = round(
        sum(info["monto"] for _, info in cli_por_monto[:10]) / total_ingresos * 100, 2
    ) if total_ingresos else 0.0
    concentracion_top10_clientes_cantidad = round(
        sum(info["cantidad"] for _, info in cli_por_cantidad[:10]) / cantidad_operaciones * 100, 2
    ) if cantidad_operaciones else 0.0

    # --- percepciones detalladas ---
    percepciones = {
        label: {"monto": round(info["monto"], 2), "cantidad": info["cantidad"]}
        for label, info in percepciones_agg.items()
    }
    total_percepciones = round(sum(p["monto"] for p in percepciones.values()), 2)

    return {
        "fecha_actualizacion": fecha_actualizacion,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "total_ingresos": total_ingresos,
        "cantidad_operaciones": cantidad_operaciones,
        "ticket_promedio": ticket_promedio,
        "facturacion_bruta": facturacion_bruta,
        "facturacion_neta": facturacion_neta,
        "total_descuentos": total_descuentos,
        "cantidad_descuentos": cantidad_descuentos,
        "envios_pagados_comprador": round(envios_pagados_comprador, 2),
        "margen": {
            "comisiones_pct": comisiones_pct,
            "envios_pct": envios_pct,
            "impuestos_pct": impuestos_pct,
            "margen_neto_pct": margen_neto_pct,
        },
        "desglose_cargos": desglose_cargos,
        "ventas_por_dia": ventas_por_dia,
        "top_provincias": top_provincias,
        "provincias_completas": provincias_completas,
        "top_categorias": top_categorias,
        "top_productos": top_productos,
        "top_clientes_facturacion": top_clientes_facturacion,
        "top_clientes_frecuentes": top_clientes_frecuentes,
        "concentracion": {
            "total_categorias_unicas": total_categorias_unicas,
            "concentracion_top10_categorias_facturacion": concentracion_top10_categorias_facturacion,
            "concentracion_top10_categorias_cantidad": concentracion_top10_categorias_cantidad,
            "total_clientes_unicos": total_clientes_unicos,
            "concentracion_top10_clientes_facturacion": concentracion_top10_clientes_facturacion,
            "concentracion_top10_clientes_cantidad": concentracion_top10_clientes_cantidad,
        },
        "percepciones": percepciones,
        "total_percepciones": total_percepciones,
        # % sobre el ingreso bruto por venta (misma base que "Impuestos y Retenciones" del margen)
        "total_percepciones_pct": round(total_percepciones / base * 100, 2) if base else 0.0,
    }


def _titulo_cargo_ml(s: str) -> str:
    """Title-case 'suave' para labels de cargos ML: capitaliza cada palabra salvo los
    conectores (por/de/del/...), igual que el resto del visor los muestra."""
    _STOP = {"por", "de", "del", "la", "el", "los", "las", "en", "y", "a", "con", "al"}
    palabras = s.split()
    out = []
    for i, w in enumerate(palabras):
        wl = w.lower()
        if i > 0 and _strip_accents(wl) in _STOP:
            out.append(wl)
        else:
            out.append(wl[:1].upper() + wl[1:] if wl else wl)
    return " ".join(out)


def _label_cargo_ml(raw: str) -> str:
    _RENAME = {"cargo por envios de mercado libre": "Envíos"}
    key = _strip_accents(raw.strip().lower())
    return _RENAME.get(key, _titulo_cargo_ml(raw))


def _render_facturacion_ml_html(d: dict) -> str:
    """Arma el HTML del panel derecho de 'facturacion_ml': secciones con separador
    uppercase (ícono + título) y filas label/valor alineadas, a partir del dict devuelto
    por analizar_facturacion_ml. d nunca se muta (viene de extracted_data, persistido tal cual)."""

    def _money(v) -> str:
        try:
            n = float(v)
        except (TypeError, ValueError):
            return "—"
        s = f"{abs(n):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"$ {'-' if n < 0 else ''}{s}"

    def _pct(v) -> str:
        try:
            n = float(v)
        except (TypeError, ValueError):
            return "—"
        return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + " %"

    def _entero(v) -> str:
        try:
            n = int(v)
        except (TypeError, ValueError):
            return "—"
        return f"{n:,}".replace(",", ".")

    _ROW_CSS = (
        "display:flex;justify-content:space-between;align-items:baseline;"
        "padding:3px 12px;border-bottom:1px solid #f5f5f5;font-size:11px;gap:12px"
    )
    _STYLE = f"<style>.fml-row{{{_ROW_CSS}}}.fml-row:hover{{background:#fafafa}}</style>"
    _SEP = (
        f"font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.06em;"
        f"color:{_HDR_COLOR};background:{_HDR_BG};padding:8px 12px;"
        f"border-bottom:1px solid {_HDR_BORDER};margin-top:14px;display:flex;"
        f"align-items:center;gap:6px"
    )

    def _sec(icon: str, titulo: str) -> str:
        return f'<div style="{_SEP}"><i class="ti {icon}"></i><span>{titulo}</span></div>'

    def _row(label, value, bold: bool = False, indent: bool = False) -> str:
        style_label = "color:#555"
        if indent:
            style_label += ";padding-left:16px"
        if bold:
            style_label += ";font-weight:700"
        style_value = f"color:{_BLUE};font-variant-numeric:tabular-nums;white-space:nowrap;text-align:right"
        if bold:
            style_value += ";font-weight:700"
        return (
            '<div class="fml-row">'
            f'<span style="{style_label}">{_escape_prompt_html(str(label))}</span>'
            f'<span style="{style_value}">{_escape_prompt_html(str(value))}</span>'
            '</div>'
        )

    parts = [_STYLE]

    parts.append(_sec("ti-calendar", "Período"))
    parts.append(_row("Fecha Desde", d.get("fecha_desde") or "—"))
    parts.append(_row("Fecha Hasta", d.get("fecha_hasta") or "—"))

    parts.append(_sec("ti-cash", "Totales Generales"))
    parts.append(_row("Total de Ingresos", _money(d.get("total_ingresos"))))
    parts.append(_row("Cantidad de Operaciones", _entero(d.get("cantidad_operaciones"))))
    parts.append(_row("Ticket Promedio", _money(d.get("ticket_promedio"))))
    parts.append(_row("Facturación Bruta", _money(d.get("facturacion_bruta"))))
    parts.append(_row(
        "Total Descuentos",
        f"{_money(d.get('total_descuentos'))} ({_entero(d.get('cantidad_descuentos'))})",
    ))
    parts.append(_row("Facturación Neta", _money(d.get("facturacion_neta")), bold=True))
    parts.append(_row("Envíos Pagados por Comprador", _money(d.get("envios_pagados_comprador"))))

    _margen = d.get("margen") or {}
    parts.append(_sec("ti-chart-bar", "Margen Neto Estimado"))
    parts.append(_row("Ingreso Bruto por Venta", _pct(100)))
    parts.append(_row("(-) Comisiones de Venta", _pct(_margen.get("comisiones_pct"))))
    parts.append(_row("(-) Costos de Envío Neto", _pct(_margen.get("envios_pct"))))
    parts.append(_row("(-) Impuestos y Retenciones", _pct(_margen.get("impuestos_pct"))))
    parts.append(_row("(=) Margen Neto Estimado", _pct(_margen.get("margen_neto_pct")), bold=True))

    _desglose = d.get("desglose_cargos") or {}
    if _desglose:
        parts.append(_sec("ti-package", "Desglose por Tipo de Cargo"))
        for label, info in _desglose.items():
            parts.append(_row(
                _label_cargo_ml(label),
                f"{_money(info.get('total'))} ({_entero(info.get('cantidad'))} cargos)",
            ))
            if info.get("anulacion_cantidad"):
                parts.append(_row(
                    "Anulaciones",
                    f"{_money(info.get('anulacion_monto'))} ({_entero(info.get('anulacion_cantidad'))})",
                    indent=True,
                ))
                parts.append(_row(
                    "Proporción Anulaciones", _pct(info.get("proporcion_anulaciones")), indent=True,
                ))

    _por_dia = d.get("ventas_por_dia") or {}
    if _por_dia:
        parts.append(_sec("ti-calendar", "Ventas por Día de la Semana"))
        for dia in ("Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"):
            info = _por_dia.get(dia) or {}
            parts.append(_row(
                dia,
                f"{_money(info.get('monto'))} "
                f"({_entero(info.get('cantidad_ventas'))} ventas — {_pct(info.get('porcentaje'))})",
            ))

    _provs = d.get("top_provincias") or []
    if _provs:
        parts.append(_sec("ti-world", "Top 5 Provincias"))
        for i, p in enumerate(_provs, 1):
            parts.append(_row(
                f"{i}  {p.get('nombre')}",
                f"{_money(p.get('monto'))} ({_pct(p.get('porcentaje'))} — {_entero(p.get('cantidad_ventas'))} ventas)",
            ))

    _cats = d.get("top_categorias") or []
    if _cats:
        parts.append(_sec("ti-tag", "Top 10 Categorías"))
        for i, c in enumerate(_cats, 1):
            parts.append(_row(
                f"{i}  {c.get('nombre')}",
                f"{_money(c.get('monto'))} ({_pct(c.get('porcentaje'))} — {_entero(c.get('cantidad_ventas'))} ventas)",
            ))

    _prods = d.get("top_productos") or []
    if _prods:
        parts.append(_sec("ti-package", "Top 10 Productos"))
        for i, p in enumerate(_prods, 1):
            parts.append(_row(
                f"{i}  {p.get('nombre')}",
                f"{_money(p.get('monto'))} ({_entero(p.get('unidades'))} u. — {_pct(p.get('porcentaje'))})",
            ))

    _clis_f = d.get("top_clientes_facturacion") or []
    if _clis_f:
        parts.append(_sec("ti-users", "Top 10 Clientes (por facturación)"))
        for i, c in enumerate(_clis_f, 1):
            n = c.get("cantidad_ventas") or 0
            parts.append(_row(
                f"{i}  {c.get('nombre')}",
                f"{_money(c.get('monto'))} ({_pct(c.get('porcentaje'))} — {_entero(n)} venta{'s' if n != 1 else ''})",
            ))

    _clis_c = d.get("top_clientes_frecuentes") or []
    if _clis_c:
        parts.append(_sec("ti-users", "Top 10 Clientes Frecuentes (por cantidad)"))
        for i, c in enumerate(_clis_c, 1):
            n = c.get("cantidad_ventas") or 0
            parts.append(_row(
                f"{i}  {c.get('nombre')}",
                f"{_entero(n)} venta{'s' if n != 1 else ''}, {_money(c.get('monto'))}",
            ))

    _conc = d.get("concentracion") or {}
    if _conc:
        parts.append(_sec("ti-target", "Concentración"))
        parts.append(_row("Categorías únicas", _entero(_conc.get("total_categorias_unicas"))))
        parts.append(_row(
            "Concentración Top 10 categorías (facturación)",
            _pct(_conc.get("concentracion_top10_categorias_facturacion")),
        ))
        parts.append(_row(
            "Concentración Top 10 categorías (cantidad)",
            _pct(_conc.get("concentracion_top10_categorias_cantidad")),
        ))
        parts.append(_row("Clientes únicos", _entero(_conc.get("total_clientes_unicos"))))
        parts.append(_row(
            "Concentración Top 10 clientes (facturación)",
            _pct(_conc.get("concentracion_top10_clientes_facturacion")),
        ))
        parts.append(_row(
            "Concentración Top 10 clientes (cantidad)",
            _pct(_conc.get("concentracion_top10_clientes_cantidad")),
        ))

    _percs = d.get("percepciones") or {}
    if _percs:
        parts.append(_sec("ti-receipt-tax", "Percepciones Detalladas"))
        for label, info in sorted(_percs.items(), key=lambda kv: kv[1].get("monto", 0), reverse=True):
            parts.append(_row(label, _money(info.get("monto"))))
        parts.append(_row(
            "Total Percepciones",
            f"{_money(d.get('total_percepciones'))} ({_pct(d.get('total_percepciones_pct'))})",
            bold=True,
        ))

    return "".join(parts)


def _es_pagos_facturas(filename: str) -> bool:
    """Detecta el 'Reporte Pagos Facturas' de MercadoLibre por el nombre del archivo
    (p.ej. 'Reporte_Pagos_Facturas_Abr2026.xlsx'), sin importar acentos/mayúsculas."""
    norm = _strip_accents(filename.lower())
    return "pagos" in norm and "facturas" in norm


def _es_facturacion_ml(filename: str) -> bool:
    """Detecta el 'Reporte Facturación MercadoLibre' (el reporte principal, con el detalle
    completo de cargos/comisiones/bonificaciones) por el nombre del archivo
    (p.ej. 'Reporte_Facturacion_MercadoLibre_Abr2026.xlsx'), sin importar acentos/mayúsculas."""
    norm = _strip_accents(filename.lower())
    return "facturacion" in norm and "mercadolibre" in norm


REPORTE_ML_DESCRIPCIONES = {
    "nota_debito_enviosflex": (
        "Las anulaciones de bonificaciones de Mercado Envíos Flex se generan cuando se cancela "
        "una venta en la cual Mercado Libre te había bonificado parte de la tarifa. Al anularse, "
        "te emitiremos una Nota de débito para que puedas realizar tus conciliaciones y "
        "descontamos el dinero que te habíamos acreditado en Mercado Pago por esta bonificación."
    ),
    "nota_credito_enviosflex": (
        "Las bonificaciones de Mercado Envíos Flex se generan cuando tu cliente obtiene envío "
        "gratis en compras de productos nuevos que superan el importe mínimo. En ese caso, "
        "Mercado Libre te bonifica parte de la tarifa según tu reputación."
    ),
    "pagos_facturas": (
        "En este reporte no aparecen los pagos cobrados en la operación. Puedes revisar todos "
        "los cargos facturados, incluídos los cobrados en la operación, en el Reporte de "
        "Facturación de Mercado Libre y/o Mercado Pago.\n\n"
        "Las notas de crédito son devoluciones por bonificaciones de cargos o cancelación de "
        "ventas que se generaron después del cierre de una factura. Las verás aplicadas para "
        "pagar el saldo de tu período. Puedes encontrar el detalle de cada Nota de crédito en "
        "Factura y Reportes."
    ),
    "facturacion_ml": (
        "Este reporte contiene el detalle completo de la facturación de MercadoLibre del "
        "período: cargos por publicación, comisiones, bonificaciones y todos los conceptos "
        "que ML te cobra o te acredita."
    ),
}


def _tipo_reporte_ml_descripcion(filename: str) -> Optional[str]:
    """Clasifica un archivo de 'Reportes MercadoLibre' para elegir, si existe, el texto
    descriptivo de REPORTE_ML_DESCRIPCIONES a mostrar en el panel del documento original."""
    norm = _strip_accents(filename.lower())
    if "nota" in norm and "debito" in norm and "flex" in norm:
        return "nota_debito_enviosflex"
    if "nota" in norm and "credito" in norm and "flex" in norm:
        return "nota_credito_enviosflex"
    if _es_pagos_facturas(filename):
        return "pagos_facturas"
    if _es_facturacion_ml(filename):
        return "facturacion_ml"
    return None


_NOTAS_BONIF_CONFIG = {
    # Header exacto (normalizado: sin acentos, minúscula) de la columna de valor en cada
    # reporte real de Envíos Flex, confirmado contra archivos de producción.
    "credito": {
        "valor_headers": {"valor de la bonificacion"},
        "campo_total": "total_bonificado",
        "campo_cantidad": "cantidad_bonificaciones",
    },
    "debito": {
        "valor_headers": {
            "valor de la anulacion", "valor del debito", "valor debitado",
            "monto debitado", "monto debito", "importe debitado",
        },
        "campo_total": "total_debitado",
        "campo_cantidad": "cantidad_debitos",
    },
}


def _calcular_notas_bonificacion(filas: list, tipo: str) -> Optional[dict]:
    """Calcula en Python (100% preciso, sin depender de Gemini) los 4 campos de los reportes
    'notas de crédito' / 'notas de débito' de Envíos Flex: fecha_desde/fecha_hasta (rango de
    fechas de la tabla de detalle) y total/cantidad (suma en valor absoluto y conteo de filas
    con valor != 0) de la columna de valor. En estos reportes los valores pueden venir en
    NEGATIVO (son anulaciones de un cargo previo) — se toma el valor absoluto porque el total
    representa el monto acreditado/debitado, no un saldo con signo."""
    cfg = _NOTAS_BONIF_CONFIG[tipo]

    def _norm(s: str) -> str:
        return _strip_accents(s.strip().lower())

    idx_fecha = idx_valor = header_idx = None
    for i, linea in enumerate(filas):
        celdas_norm = [_norm(c) for c in linea.split("\t")]
        cand_fecha = next((j for j, c in enumerate(celdas_norm) if "fecha" in c), None)
        cand_valor = next(
            (j for j, c in enumerate(celdas_norm) if c in cfg["valor_headers"]), None
        )
        if cand_fecha is not None and cand_valor is not None:
            idx_fecha, idx_valor, header_idx = cand_fecha, cand_valor, i
            break

    if header_idx is None:
        return None

    fechas = []
    valores = []
    for linea in filas[header_idx + 1:]:
        celdas = linea.split("\t")
        if len(celdas) <= max(idx_fecha, idx_valor):
            continue
        if celdas[idx_fecha].strip():
            try:
                fechas.append(datetime.strptime(celdas[idx_fecha].strip().split(" ")[0], "%Y-%m-%d"))
            except ValueError:
                pass
        try:
            v = float(celdas[idx_valor])
        except ValueError:
            v = 0.0
        if v != 0:
            valores.append(abs(v))

    return {
        "fecha_desde": fechas and min(fechas).strftime("%d/%m/%Y") or None,
        "fecha_hasta": fechas and max(fechas).strftime("%d/%m/%Y") or None,
        cfg["campo_total"]: round(sum(valores), 2),
        cfg["campo_cantidad"]: len(valores),
    }


def _extraer_fecha_actualizacion(filas: list) -> Optional[str]:
    """Busca la fila 'Fecha de actualización' de la cabecera del Excel y devuelve su
    valor formateado DD/MM/AAAA."""
    for linea in filas:
        celdas = linea.split("\t")
        if not celdas or _strip_accents(celdas[0].strip().lower()) != "fecha de actualizacion":
            continue
        for c in celdas[1:]:
            if not c.strip():
                continue
            try:
                return datetime.strptime(c.strip().split(" ")[0], "%Y-%m-%d").strftime("%d/%m/%Y")
            except ValueError:
                continue
    return None


def _calcular_pagos_facturas(filas: list) -> Optional[dict]:
    """Calcula en Python (100% preciso, sin depender de Gemini) el detalle completo del
    reporte 'Pagos Facturas': fecha_actualizacion (de la cabecera del Excel) y la lista
    completa de filas de la tabla 'Pagos y notas de crédito' (fecha_pago, estado,
    nota_credito, importe_total). Las filas con estado 'Facturado' en realidad son notas de
    crédito (no pagos) — se normalizan a estado='Nota de crédito' + su número, tomado de la
    columna 'Número de nota de crédito'. El archivo real trae 2 hojas; la segunda
    ('Detalle de Pagos del mes') tiene menos columnas y queda excluida naturalmente por el
    chequeo de ancho de fila."""
    def _norm(s: str) -> str:
        return _strip_accents(s.strip().lower())

    idx_fecha = idx_estado = idx_nc = idx_importe = header_idx = None
    for i, linea in enumerate(filas):
        celdas_norm = [_norm(c) for c in linea.split("\t")]
        cand_fecha = next((j for j, c in enumerate(celdas_norm) if "fecha de pago" in c), None)
        cand_estado = next((j for j, c in enumerate(celdas_norm) if c == "estado"), None)
        cand_nc = next(
            (j for j, c in enumerate(celdas_norm)
             if "numero de nota de credito" in c or "nro nc" in c or "numero de nc" in c),
            None,
        )
        cand_importe = next((j for j, c in enumerate(celdas_norm) if "importe total" in c), None)
        if cand_fecha is not None and cand_estado is not None and cand_importe is not None:
            idx_fecha, idx_estado, idx_nc, idx_importe, header_idx = (
                cand_fecha, cand_estado, cand_nc, cand_importe, i
            )
            break

    if header_idx is None:
        return None

    pagos = []
    for linea in filas[header_idx + 1:]:
        celdas = linea.split("\t")
        if len(celdas) <= max(idx_fecha, idx_estado, idx_importe):
            continue
        try:
            importe = float(celdas[idx_importe])
        except ValueError:
            continue

        fecha_pago = None
        if celdas[idx_fecha].strip():
            try:
                fecha_pago = datetime.strptime(
                    celdas[idx_fecha].strip().split(" ")[0], "%Y-%m-%d"
                ).strftime("%d/%m/%Y")
            except ValueError:
                pass

        estado_raw = celdas[idx_estado].strip() or None
        estado_final = estado_raw
        nota_credito = None
        if estado_raw == "Facturado":
            estado_final = "Nota de crédito"
            if idx_nc is not None and idx_nc < len(celdas):
                nota_credito = celdas[idx_nc].strip() or None

        pagos.append({
            "fecha_pago": fecha_pago,
            "estado": estado_final,
            "nota_credito": nota_credito,
            "importe_total": round(importe, 2),
        })

    if not pagos:
        return None

    return {
        "fecha_actualizacion": _extraer_fecha_actualizacion(filas),
        "pagos": pagos,
    }


# ---------------------------------------------------------------------------
# Análisis consolidado del período (cruza las 6 secciones de Gastos)
# ---------------------------------------------------------------------------


def _ar_money(v) -> str:
    try:
        n = float(v)
    except (TypeError, ValueError):
        return "—"
    s = f"{round(abs(n)):,}".replace(",", ".")
    return f"$ {'-' if n < 0 else ''}{s}"


def _ar_num(v) -> str:
    try:
        n = int(round(float(v)))
    except (TypeError, ValueError):
        return "—"
    return f"{n:,}".replace(",", ".")


def _ed(archivo: Optional[Dict[str, Any]]) -> dict:
    """Parsea extracted_data de un registro de gastos_archivos; {} si no hay o es inválido."""
    if not archivo:
        return {}
    try:
        return json.loads(archivo.get("extracted_data") or "{}") or {}
    except Exception:
        return {}


def _clasificar_reportes_ml(archivos: list) -> Dict[str, Optional[dict]]:
    """Separa los archivos de 'Reportes MercadoLibre' del período por subtipo, según el nombre."""
    out: Dict[str, Optional[dict]] = {
        "facturacion_ml": None, "pagos_facturas": None,
        "nc_ml": None, "nc_flex": None, "nd_flex": None,
    }
    for f in archivos:
        fn = f.get("filename", "")
        if _es_facturacion_ml(fn):
            out["facturacion_ml"] = f
        elif _es_pagos_facturas(fn):
            out["pagos_facturas"] = f
        elif _es_notas_credito_ml(fn):
            out["nc_ml"] = f
        elif _es_notas_credito(fn):
            out["nc_flex"] = f
        elif _es_notas_debito(fn):
            out["nd_flex"] = f
    return out


def _buscar_cargo_neto(desglose: dict, *nombres_norm: str) -> float:
    """Busca en desglose_cargos (de analizar_facturacion_ml) el/los label(s) normalizados
    dados y devuelve total + anulacion_monto (neto de anulaciones)."""
    total = 0.0
    for label, info in (desglose or {}).items():
        if _strip_accents(label.strip().lower()) in nombres_norm:
            total += (info.get("total") or 0.0) + (info.get("anulacion_monto") or 0.0)
    return round(total, 2)


# Mapeo de patterns (sin acentos, lowercase) a jurisdicción canónica.
# "CABA" agrupa tanto el régimen general como "Comercio Electrónico" — Diego
# los quiere sumados en una sola fila.
_JURISDICCION_PERCEPCION_MAP = [
    ("caba", "CABA"),
    ("ciudad autonoma", "CABA"),
    ("capital federal", "CABA"),
    ("buenos aires", "Buenos Aires"),
    ("bs as", "Buenos Aires"),
    ("catamarca", "Catamarca"),
    ("chaco", "Chaco"),
    ("chubut", "Chubut"),
    ("cordoba", "Córdoba"),
    ("corrientes", "Corrientes"),
    ("entre rios", "Entre Ríos"),
    ("formosa", "Formosa"),
    ("jujuy", "Jujuy"),
    ("la pampa", "La Pampa"),
    ("la rioja", "La Rioja"),
    ("mendoza", "Mendoza"),
    ("misiones", "Misiones"),
    ("neuquen", "Neuquén"),
    ("rio negro", "Río Negro"),
    ("salta", "Salta"),
    ("san juan", "San Juan"),
    ("san luis", "San Luis"),
    ("santa cruz", "Santa Cruz"),
    ("santa fe", "Santa Fe"),
    ("sgo del estero", "Santiago del Estero"),
    ("santiago del estero", "Santiago del Estero"),
    ("tierra del fuego", "Tierra del Fuego"),
    ("tucuman", "Tucumán"),
]


def normalizar_jurisdiccion_percepcion(concepto: str) -> Optional[str]:
    """Devuelve el nombre canónico de la jurisdicción/provincia a partir del concepto
    de una línea de percepción de IIBB (Facturas ML o reportes de Percepciones).
    Usada en ambos lados para que las claves coincidan al cruzar. None si no matchea."""
    if not concepto:
        return None
    s = _strip_accents(concepto.lower())
    for pattern, jurisdiccion in _JURISDICCION_PERCEPCION_MAP:
        if pattern in s:
            return jurisdiccion
    return None


def _clasificar_impuesto(nombre: str) -> str:
    """Clasifica un nombre de impuesto (Retenciones/Pagos ARCA) en una categoría común
    para poder cruzarlo contra percepciones y pagos a ARCA."""
    n = _strip_accents((nombre or "").strip().lower())
    if "sirtac" in n:
        return "SIRTAC"
    if "credito" in n and "debito" in n:
        return "Impuesto Créditos y Débitos"
    if "ganancias" in n:
        return "Ganancias"
    if "iibb" in n or "ingresos brutos" in n or "convenio multilateral" in n or "sifere" in n:
        return "IIBB"
    if "iva" in n:
        return "IVA"
    return "Otros"


def _neto_gravado(fd: dict) -> float:
    v = fd.get("neto_gravado")
    if v is None:
        v = fd.get("subtotal")
    return v or 0.0


# ---------------------------------------------------------------------------
# Sección 7 — Cruce Ventas nuestras (BD) vs Reporte de Facturación ML
# ---------------------------------------------------------------------------

TOLERANCIA_CRUCE_VENTAS_PCT = 0.005  # 0.5%


def _rango_fechas_periodo(periodo: str) -> tuple:
    """(inicio, fin) del mes 'YYYY-MM' como datetime, fin al último segundo del mes."""
    anio, mes = int(periodo[:4]), int(periodo[5:7])
    inicio = datetime(anio, mes, 1)
    fin = (datetime(anio + 1, 1, 1) if mes == 12 else datetime(anio, mes + 1, 1)) - timedelta(seconds=1)
    return inicio, fin


def _parsear_filas_facturacion_ml(path: Path) -> Dict[str, dict]:
    """Re-lee el Excel de 'Reporte de Facturación ML' directamente del disco (misma
    estructura que analizar_facturacion_ml: hoja REPORT, headers filas 7-8, detalle
    desde fila 9) para conservar el detalle por Número de venta — analizar_facturacion_ml
    lo descarta luego de calcular los agregados, así que para el cruce de la Sección 7
    hace falta volver a leer el archivo original."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_name = next(
        (s for s in wb.sheetnames if _strip_accents(s.strip().lower()) == "report"), None
    )
    if sheet_name is None:
        wb.close()
        return {}
    ws = wb[sheet_name]

    def _col(row, idx1: int):
        j = idx1 - 1
        return row[j] if j < len(row) else None

    def _num(v) -> float:
        if v is None or v == "":
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    def _fecha(v) -> Optional[str]:
        if v is None or v == "":
            return None
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        try:
            return datetime.strptime(str(v).strip().split(" ")[0], "%Y-%m-%d").strftime("%Y-%m-%d")
        except ValueError:
            return None

    filas = [
        row for row in ws.iter_rows(min_row=9, values_only=True)
        if not all(c is None for c in row)
    ]
    wb.close()

    ventas: Dict[str, dict] = {}
    for row in filas:
        nro = _col(row, 13)
        if nro is None or str(nro).strip() == "":
            continue
        nro = str(nro).strip()
        v = ventas.setdefault(nro, {
            "fecha_venta": None, "cliente": None, "total_venta": None,
            "provincia": None, "categoria": None,
            "comision_ml": 0.0, "fue_anulada": False,
        })
        _fv = _fecha(_col(row, 15))
        if _fv and not v["fecha_venta"]:
            v["fecha_venta"] = _fv
        for campo, idx in (("cliente", 18), ("provincia", 19), ("categoria", 31)):
            _val = _col(row, idx)
            if _val and not v[campo]:
                v[campo] = str(_val).strip()
        _tv = _num(_col(row, 23))
        if _tv and not v["total_venta"]:
            v["total_venta"] = _tv

        detalle = _strip_accents(str(_col(row, 4) or "").strip().lower())
        valor_cargo = _num(_col(row, 8))
        if detalle == "cargo por vender":
            v["comision_ml"] += valor_cargo
        elif detalle == "anulacion del cargo por vender":
            v["fue_anulada"] = True

    for v in ventas.values():
        v["comision_ml"] = round(v["comision_ml"], 2)
        v["total_venta"] = round(v["total_venta"] or 0.0, 2)
    return ventas


def _fetch_meli_fee_mp(access_token: str, payment_id: str) -> Optional[float]:
    """Comisión real cobrada por ML para un pago, vía la misma API que usa
    tabs/ventas.py en 'Completar datos' (MercadoPago Payments, no un endpoint de ML)."""
    try:
        r = requests.get(
            f"https://api.mercadopago.com/v1/payments/{payment_id}",
            headers={"Authorization": f"Bearer {access_token}"}, timeout=15,
        )
        if r.status_code != 200:
            return None
        charges = r.json().get("charges_details") or []
        return sum(
            float((c.get("amounts") or {}).get("original", 0))
            for c in charges if c.get("name") == "meli_percentage_fee"
        )
    except Exception:
        return None


def _obtener_ventas_bd_periodo(user_id: int, inicio: datetime, fin: datetime) -> tuple:
    """Trae del lado nuestro (API de MercadoLibre, no ml_orders_cache — ese cache solo
    crece hacia adelante desde que se abrió el Panel por primera vez y no garantiza
    cobertura de un mes cerrado ya pasado) las órdenes entre [inicio, fin] (el período de
    corte de ML — ver _calcular_periodo_ml — no el mes calendario), indexadas por
    order_id. comision_bd es la comisión REAL cobrada por ML: primero se busca en
    ventas_datos (ya cacheada si el usuario usó "Completar datos" en Ventas), y para
    las que falten se pide en vivo a la API de pagos de MercadoPago
    (charges_details -> meli_percentage_fee), igual mecanismo que tabs/ventas.py. Solo
    si no hay payment_id o falla la consulta se cae a una estimación por tasa plana
    (ml_comision del cotizador × total). Retorna (ventas_bd, seller_id)."""
    access_token = get_ml_access_token(user_id)
    if not access_token:
        return {}, None
    profile = ml_get_user_profile(access_token) or {}
    seller_id = profile.get("id") or ml_get_user_id(access_token)
    if not seller_id:
        return {}, None

    date_from = inicio.strftime("%Y-%m-%dT00:00:00.000-03:00")
    date_to = fin.strftime("%Y-%m-%dT23:59:59.999-03:00")
    orders_data = ml_get_orders(
        access_token, str(seller_id), limit=5000, offset=0,
        date_from=date_from, date_to=date_to,
    )
    raw_orders = orders_data.get("results") or orders_data.get("orders") or orders_data.get("elements") or []

    ml_com = 0.15
    try:
        ml_com = float(get_cotizador_param("ml_comision", user_id) or 0.15)
    except (TypeError, ValueError):
        pass

    # payment_ids por orden: TODOS los pagos de la orden, sin filtrar por status. Un
    # comprador puede dividir el pago en más de uno (p.ej. cuenta + tarjeta) y en
    # órdenes canceladas/reembolsadas el cargo meli_percentage_fee puede quedar en un
    # pago "refunded"/"rejected" en vez de uno "approved" — para reconciliar el total
    # cobrado hace falta sumar el cargo de todos los pagos de la orden, no solo el
    # "principal" (a diferencia de tabs/ventas.py, que sí necesita uno solo para
    # calcular la ganancia por venta).
    pids_por_orden: Dict[str, List[str]] = {}
    for o in raw_orders:
        if not isinstance(o, dict) or not o.get("id"):
            continue
        pays = o.get("payments") or []
        pids = [str(p["id"]) for p in pays if p.get("id")]
        if pids:
            pids_por_orden[str(o["id"])] = pids

    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute("SELECT payment_id, meli_fee FROM ventas_datos WHERE user_id=?", (user_id,))
        # meli_fee=0.0 cacheado se descarta: un cargo real de ML nunca es $0 exacto —
        # 0.0 en ventas_datos indica que "Completar datos" no encontró el cargo en su
        # momento (p.ej. en una orden que luego se canceló/reembolsó), no que sea gratis.
        # Se prefiere re-consultar en vivo antes que arrastrar ese valor stale.
        fee_cache: Dict[str, float] = {r[0]: r[1] for r in cur.fetchall() if r[1]}
    finally:
        conn.close()

    todos_pids = sorted({pid for pids in pids_por_orden.values() for pid in pids})
    faltantes = [pid for pid in todos_pids if pid not in fee_cache]
    if faltantes:
        with ThreadPoolExecutor(max_workers=min(16, len(faltantes))) as ex:
            for pid, fee in zip(faltantes, ex.map(lambda p: _fetch_meli_fee_mp(access_token, p), faltantes)):
                if fee is not None:
                    fee_cache[pid] = fee

    ventas_bd: Dict[str, dict] = {}
    for o in raw_orders:
        if not isinstance(o, dict) or not o.get("id"):
            continue
        order_id = str(o["id"])
        total_bd = float(o.get("total_amount") or o.get("paid_amount") or 0.0)
        status_raw = str(o.get("status") or "").strip().lower()
        buyer = o.get("buyer") or {}
        cliente = buyer.get("nickname") or f'{buyer.get("first_name", "")} {buyer.get("last_name", "")}'.strip() or "—"
        fecha = str(o.get("date_created") or "")[:10]

        pids = pids_por_orden.get(order_id) or []
        fees = [fee_cache[p] for p in pids if p in fee_cache]
        fee_real = sum(fees) if pids and len(fees) == len(pids) else None
        comision_bd = round(fee_real, 2) if fee_real is not None else round(total_bd * ml_com, 2)

        ventas_bd[order_id] = {
            "fecha": fecha,
            "cliente": cliente,
            "total_bd": round(total_bd, 2),
            "comision_bd": comision_bd,
            "status_raw": status_raw,
        }
    return ventas_bd, str(seller_id)


def _cruzar_ventas_reporte_vs_bd(
    filas_reporte: Dict[str, dict], ventas_bd: Dict[str, dict],
    tolerancia_pct: float = TOLERANCIA_CRUCE_VENTAS_PCT,
) -> dict:
    """Cruza por Número de venta / order_id. Recorre cada diccionario una sola vez
    (ya indexados por clave) para cumplir el requisito de performance (<2s ante ~1200
    ventas por lado)."""
    cruzadas_ok: List[dict] = []
    con_diferencias: List[dict] = []
    anuladas: List[dict] = []
    solo_reporte: List[dict] = []

    for nro, fr in filas_reporte.items():
        bd = ventas_bd.get(nro)
        if fr.get("fue_anulada"):
            marcada_bd = bd is not None and bd["status_raw"] in ("cancelled", "canceled")
            anuladas.append({
                "nro_venta": nro, "fecha": fr.get("fecha_venta") or "", "cliente": fr.get("cliente") or "",
                "total_venta": fr.get("total_venta") or 0.0,
                "status_bd": (bd or {}).get("status_raw") or "(no encontrada en BD)",
                "observacion": "OK — marcada como cancelada en BD" if marcada_bd else "⚠ ML la anuló pero no está cancelada en nuestra BD",
                "_marcada_ok": marcada_bd,
            })
            continue
        if bd is None:
            solo_reporte.append({
                "nro_venta": nro, "fecha": fr.get("fecha_venta") or "", "cliente": fr.get("cliente") or "",
                "total_venta": fr.get("total_venta") or 0.0, "comision_ml": fr.get("comision_ml") or 0.0,
                "provincia": fr.get("provincia") or "", "categoria": fr.get("categoria") or "",
            })
            continue

        total_ml = fr.get("total_venta") or 0.0
        comision_ml = fr.get("comision_ml") or 0.0
        diff_total = abs(total_ml - bd["total_bd"]) / total_ml if total_ml else 0.0
        diff_comision = abs(comision_ml - bd["comision_bd"]) / comision_ml if comision_ml else 0.0

        fila = {
            "nro_venta": nro, "fecha": fr.get("fecha_venta") or "", "cliente": fr.get("cliente") or "",
            "total_reporte": total_ml, "total_bd": bd["total_bd"], "diff_total_pct": round(diff_total * 100, 2),
            "comision_reporte": comision_ml, "comision_bd": bd["comision_bd"], "diff_comision_pct": round(diff_comision * 100, 2),
            "ganancia_calculada": None,  # No se persiste en ningún lado del proyecto para períodos pasados; requeriría
                                          # rearmar el costeo por SKU igual que tabs/ventas.py, fuera del alcance del cruce.
        }
        if diff_total < tolerancia_pct and diff_comision < tolerancia_pct:
            cruzadas_ok.append(fila)
        else:
            tipos = []
            if diff_total >= tolerancia_pct:
                tipos.append("total")
            if diff_comision >= tolerancia_pct:
                tipos.append("comision")
            fila["tipo_diff"] = "+".join(tipos)
            con_diferencias.append(fila)

    solo_bd = [
        {
            "nro_venta": nro, "fecha": bd["fecha"], "cliente": bd["cliente"],
            "total_venta": bd["total_bd"], "comision_calculada": bd["comision_bd"],
            "ganancia_calculada": None, "status": bd["status_raw"],
        }
        for nro, bd in ventas_bd.items() if nro not in filas_reporte
    ]

    total_reporte = len(filas_reporte)
    total_bd = len(ventas_bd)
    salud = (len(cruzadas_ok) / total_reporte * 100) if total_reporte else 0.0

    return {
        "cruzadas_ok": cruzadas_ok, "solo_reporte": solo_reporte, "solo_bd": solo_bd,
        "con_diferencias": con_diferencias, "anuladas": anuladas,
        "total_reporte": total_reporte, "total_bd": total_bd,
        "salud_pct": round(salud, 1),
        "tolerancia_pct": tolerancia_pct * 100,
    }


def _calcular_periodo_ml(fact: dict, inicio_fiscal: datetime, fin_fiscal: datetime) -> tuple:
    """Rango del 'período de corte' de ML (ej. mayo = 17/04-16/05 aprox.), tomado del
    MIN/MAX de fecha de venta del Reporte de Facturación ML del período (fact["fecha_desde"]/
    ["fecha_hasta"], ya calculados por analizar_facturacion_ml sobre TODAS las ventas del
    archivo). Si no hay reporte procesado, cae al mes calendario (periodo_fiscal) y lo marca
    con fallback=True para que el modal avise que está usando el mes calendario en su lugar.

    Devuelve (desde: datetime, hasta: datetime, fallback: bool)."""
    try:
        desde = datetime.strptime(fact.get("fecha_desde") or "", "%d/%m/%Y")
        hasta = datetime.strptime(fact.get("fecha_hasta") or "", "%d/%m/%Y")
        return desde, hasta, False
    except ValueError:
        return inicio_fiscal, fin_fiscal, True


def _periodos_del_resultado(resultado: dict) -> tuple:
    """periodo_fiscal/periodo_ml (dicts con 'desde'/'hasta' en dd/mm/YYYY) ya calculados y
    guardados en el resultado. Fallback para análisis guardados antes de que existieran estas
    claves: reconstruye el mes calendario como ambos períodos, marcando periodo_ml como
    fallback."""
    periodo_fiscal = resultado.get("periodo_fiscal")
    periodo_ml = resultado.get("periodo_ml")
    if periodo_fiscal and periodo_ml:
        return periodo_fiscal, periodo_ml
    try:
        inicio, fin = _rango_fechas_periodo(resultado.get("periodo", ""))
        _desde, _hasta = inicio.strftime("%d/%m/%Y"), fin.strftime("%d/%m/%Y")
    except ValueError:
        _desde, _hasta = "", ""
    periodo_fiscal = periodo_fiscal or {"desde": _desde, "hasta": _hasta}
    periodo_ml = periodo_ml or {"desde": _desde, "hasta": _hasta, "fallback": True}
    return periodo_fiscal, periodo_ml


def _render_periodo_badge(tipo: str, periodo_fiscal: dict, periodo_ml: dict) -> str:
    """Mini-badge '📅 Fiscal/Período ML/Mixto' para el header de una sección, con tooltip
    mostrando las fechas exactas de ambos períodos."""
    label = {"fiscal": "Fiscal", "ml": "Período ML", "mixto": "Mixto"}.get(tipo, tipo)
    tooltip = (
        f"Período fiscal: {periodo_fiscal.get('desde', '')} - {periodo_fiscal.get('hasta', '')}\n"
        f"Período ML: {periodo_ml.get('desde', '')} - {periodo_ml.get('hasta', '')}"
    ).replace('"', "&quot;")
    return (
        f'<span title="{tooltip}" style="display:inline-flex;align-items:center;'
        "background:var(--color-background-secondary);color:var(--color-text-secondary);"
        'font-size:10px;font-weight:500;padding:2px 6px;border-radius:3px;cursor:help">'
        f"📅 {label}</span>"
    )


def _calcular_seccion_cruce_ventas(
    user_id: int, archivo_facturacion_ml: Optional[dict],
    periodo_ml_desde: datetime, periodo_ml_hasta: datetime,
) -> dict:
    """Arma la Sección 7 completa: si no hay Reporte de Facturación ML procesado en el
    período, devuelve _disponible=False para que el render muestre el aviso pedido.

    Filtra las ventas de nuestra BD por periodo_ml (el período de corte de ML, tomado del
    propio Reporte de Facturación), NO por el mes calendario — las ventas de nuestra BD y las
    del reporte deben cubrir la misma ventana de fechas para que el cruce por Número de venta
    tenga sentido; usar el mes calendario dejaba ~600 ventas de fin de mes "solo en BD" que en
    realidad pertenecían al período de corte de ML del mes siguiente."""
    if not archivo_facturacion_ml or not archivo_facturacion_ml.get("filepath"):
        return {"_disponible": False}
    filepath = Path(archivo_facturacion_ml["filepath"])
    if not filepath.exists():
        return {"_disponible": False}

    filas_reporte = _parsear_filas_facturacion_ml(filepath)
    if not filas_reporte:
        return {"_disponible": False}

    ventas_bd, seller_id = _obtener_ventas_bd_periodo(user_id, periodo_ml_desde, periodo_ml_hasta)
    cruce = _cruzar_ventas_reporte_vs_bd(filas_reporte, ventas_bd)
    cruce["_disponible"] = True
    cruce["seller_id"] = seller_id or ""
    return cruce


_HOJAS_CRUCE_VENTAS = {
    "ok": (
        "Cruzadas OK",
        ["Nro Venta", "Fecha", "Cliente", "Total Reporte", "Total BD", "Diff Total %",
         "Comisión Reporte", "Comisión BD", "Diff Comisión %", "Ganancia Calculada"],
        lambda f: [
            f["nro_venta"], f["fecha"], f["cliente"], f["total_reporte"], f["total_bd"],
            f["diff_total_pct"], f["comision_reporte"], f["comision_bd"], f["diff_comision_pct"],
            f["ganancia_calculada"] if f["ganancia_calculada"] is not None else "N/D",
        ],
        {4, 5, 7, 8}, {6, 9},
    ),
    "solo_reporte": (
        "Solo en Reporte ML",
        ["Nro Venta", "Fecha", "Cliente", "Total Venta", "Comisión ML", "Provincia", "Categoría"],
        lambda f: [f["nro_venta"], f["fecha"], f["cliente"], f["total_venta"], f["comision_ml"], f["provincia"], f["categoria"]],
        {4, 5}, set(),
    ),
    "solo_bd": (
        "Solo en nuestra BD",
        ["Nro Venta", "Fecha", "Cliente", "Total Venta", "Comisión Calculada", "Ganancia Calculada", "Status"],
        lambda f: [
            f["nro_venta"], f["fecha"], f["cliente"], f["total_venta"], f["comision_calculada"],
            f["ganancia_calculada"] if f["ganancia_calculada"] is not None else "N/D", f["status"],
        ],
        {4, 5}, set(),
    ),
    "diff": (
        "Con diferencias",
        ["Nro Venta", "Fecha", "Cliente", "Total Reporte", "Total BD", "Diff Total %",
         "Comisión Reporte", "Comisión BD", "Diff Comisión %", "Tipo Diff"],
        lambda f: [
            f["nro_venta"], f["fecha"], f["cliente"], f["total_reporte"], f["total_bd"],
            f["diff_total_pct"], f["comision_reporte"], f["comision_bd"], f["diff_comision_pct"], f["tipo_diff"],
        ],
        {4, 5, 7, 8}, {6, 9},
    ),
    "anuladas": (
        "Anuladas",
        ["Nro Venta", "Fecha", "Cliente", "Total Venta", "Status BD", "Observación"],
        lambda f: [f["nro_venta"], f["fecha"], f["cliente"], f["total_venta"], f["status_bd"], f["observacion"]],
        {4}, set(),
    ),
}


def _generar_excel_cruce_ventas(resultado: dict, solo_tipo: Optional[str] = None) -> tuple:
    """Genera el Excel de la Sección 7 (openpyxl). Sin solo_tipo: Resumen + las 5 hojas
    de detalle. Con solo_tipo ('ok'/'solo_reporte'/'solo_bd'/'diff'/'anuladas'): Resumen +
    únicamente esa hoja. Devuelve (path, filename)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    cr = resultado.get("cruce_ventas") or {}
    periodo = resultado.get("periodo", "")
    seller_id = cr.get("seller_id") or "sinid"

    header_fill = PatternFill(start_color="2A7AC7", end_color="2A7AC7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill(start_color="F5F8FB", end_color="F5F8FB", fill_type="solid")
    thin = Side(border_style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    right_align = Alignment(horizontal="right")

    def _escribir_hoja(ws, headers: list, filas: list, money_cols: set, num_cols: set) -> None:
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = header_fill
            c.font = header_font
            c.border = border
            c.alignment = Alignment(horizontal="center", vertical="center")
        for r_idx, fila in enumerate(filas, 2):
            for col, val in enumerate(fila, 1):
                c = ws.cell(row=r_idx, column=col, value=val)
                c.border = border
                if r_idx % 2 == 0:
                    c.fill = alt_fill
                if col in money_cols:
                    c.number_format = "$ #,##0"
                    c.alignment = right_align
                elif col in num_cols:
                    c.number_format = "0"
                    c.alignment = right_align
        for col_idx, h in enumerate(headers, 1):
            max_len = max([len(str(h))] + [len(str(f[col_idx - 1])) for f in filas] or [10])
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 10), 40)
        ws.freeze_panes = "A2"

    wb = Workbook()
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    salud = cr.get("salud_pct", 0.0)
    estado = "Excelente" if salud >= 95 else ("Con observaciones" if salud >= 85 else "Requiere atención")
    resumen_filas = [
        ("Período analizado", periodo),
        ("Salud del período", f"{salud:.1f}%".replace(".", ",")),
        ("Estado", estado),
        ("Cruzadas OK", len(cr.get("cruzadas_ok") or [])),
        ("Solo en Reporte ML", len(cr.get("solo_reporte") or [])),
        ("Solo en nuestra BD", len(cr.get("solo_bd") or [])),
        ("Con diferencias en montos", len(cr.get("con_diferencias") or [])),
        ("Anuladas", len(cr.get("anuladas") or [])),
        ("Total en Reporte ML", cr.get("total_reporte", 0)),
        ("Total en nuestra BD", cr.get("total_bd", 0)),
        ("Fecha del análisis", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Usuario", str(resultado.get("user_id", ""))),
    ]
    for r_idx, (label, val) in enumerate(resumen_filas, 1):
        ws_resumen.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        ws_resumen.cell(row=r_idx, column=2, value=val)
    ws_resumen.column_dimensions["A"].width = 28
    ws_resumen.column_dimensions["B"].width = 24

    tipos = [solo_tipo] if solo_tipo else ["ok", "solo_reporte", "solo_bd", "diff", "anuladas"]
    claves_datos = {"ok": "cruzadas_ok", "solo_reporte": "solo_reporte", "solo_bd": "solo_bd",
                     "diff": "con_diferencias", "anuladas": "anuladas"}
    for tipo in tipos:
        filas_datos = cr.get(claves_datos[tipo]) or []
        if not solo_tipo and tipo == "anuladas" and not filas_datos:
            continue
        titulo, headers, mapper, money_cols, num_cols = _HOJAS_CRUCE_VENTAS[tipo]
        ws = wb.create_sheet(titulo[:31])
        _escribir_hoja(ws, headers, [mapper(f) for f in filas_datos], money_cols, num_cols)

    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    wb.save(path)
    sufijo = f"_{solo_tipo}" if solo_tipo else ""
    nombre = f"cruce_ventas{sufijo}_{seller_id}_{periodo}.xlsx"
    return path, nombre


def analizar_periodo_consolidado(user_id: int, periodo: str) -> dict:
    """Cruza los datos extraídos de las 6 secciones de Gastos (Facturas ML, Retenciones,
    Percepciones, Pagos ARCA, Reportes ML, Análisis ML) del período dado y arma el reporte
    consolidado que muestra el modal de 'Procesar Análisis Final'.

    Fase 1: estructura y cálculos base sobre los datos ya extraídos por archivo — los
    cruces y validaciones se van a ir refinando en próximas iteraciones."""
    archivos: Dict[str, list] = {
        sk: [f for f in get_gastos_archivos(user_id, periodo, sk) if f.get("extraction_status") == "procesado"]
        for sk, _, _, _, _ in _SECCIONES
    }
    faltantes = [lbl for sk, lbl, _, _, _ in _SECCIONES if not archivos.get(sk)]

    # --- Reportes ML clasificados por subtipo ---
    reportes = _clasificar_reportes_ml(archivos.get("reportes_ml", []))
    fact       = _ed(reportes["facturacion_ml"])
    nc_ml      = _ed(reportes["nc_ml"])
    nc_flex    = _ed(reportes["nc_flex"])
    nd_flex    = _ed(reportes["nd_flex"])

    # --- Facturas ML ---
    facturas_raw = archivos.get("facturas_ml", [])
    facturas_data = [_ed(f) for f in facturas_raw]
    print(f"[DBG-CONSOLIDADO] Facturas ML encontradas: {len(facturas_data)}", flush=True)

    def _lineas_intermedias(fd: dict) -> list:
        return fd.get("lineas_intermedias") or []

    total_facturas_ml = round(sum(fd.get("total") or 0.0 for fd in facturas_data), 2)
    cantidad_facturas = len(facturas_data)
    neto_gravado_total = round(sum(_neto_gravado(fd) for fd in facturas_data), 2)
    iva_21_total = round(sum(
        l.get("monto") or 0.0
        for fd in facturas_data for l in _lineas_intermedias(fd)
        if "iva" in _concepto_norm(l.get("concepto", "")) and "21" in str(l.get("concepto", ""))
    ), 2)
    percepciones_ml_total = round(sum(
        l.get("monto") or 0.0
        for fd in facturas_data for l in _lineas_intermedias(fd)
        if "percep" in _strip_accents(str(l.get("concepto", "")).lower())
    ), 2)

    # Percepciones por provincia según las Facturas ML — excluye Notas de Crédito/Débito:
    # sus percepciones revierten lo cobrado en facturas previas y rompen el cruce vs Reportes.
    def _es_factura_para_percepciones(tipo_documento: str) -> bool:
        t = _strip_accents(str(tipo_documento or "").lower())
        if "nota de" in t or "credito" in t or "debito" in t:
            return False
        return "factura" in t

    perc_facturas: Dict[str, float] = defaultdict(float)
    perc_facturas_detalle: Dict[str, list] = defaultdict(list)
    for f_raw, fd in zip(facturas_raw, facturas_data):
        tipo_doc = fd.get("tipo_documento", "")
        if not _es_factura_para_percepciones(tipo_doc):
            print(f"[DBG-CONSOL] Excluyendo archivo por tipo='{tipo_doc}': filename={f_raw.get('filename')}", flush=True)
            continue
        print(f"[DBG-CONSOL] Incluyendo archivo por tipo='{tipo_doc}': filename={f_raw.get('filename')}", flush=True)
        for l in _lineas_intermedias(fd):
            concepto = str(l.get("concepto", ""))
            if "percep" not in _strip_accents(concepto.lower()):
                continue
            monto = l.get("monto") or 0.0
            jurisdiccion = normalizar_jurisdiccion_percepcion(concepto) or "Sin identificar"
            perc_facturas[jurisdiccion] += monto
            perc_facturas_detalle[jurisdiccion].append(f"{concepto}: {_ar_money(round(monto, 2))}")
    perc_facturas = {k: round(v, 2) for k, v in perc_facturas.items()}

    _dbg_suma_lineas, _dbg_percs_por_factura = {}, {}
    for _f_raw, _fd in zip(facturas_raw, facturas_data):
        _fid = _f_raw.get("filename") or _f_raw.get("id")
        _dbg_suma_lineas[_fid] = round(sum(l.get("monto") or 0.0 for l in _lineas_intermedias(_fd)), 2)
        _dbg_percs_por_factura[_fid] = [
            {"concepto": l.get("concepto"), "monto": l.get("monto")}
            for l in _lineas_intermedias(_fd)
            if "percep" in _strip_accents(str(l.get("concepto", "")).lower())
        ]
    print(f"[DBG-CONSOLIDADO] Suma lineas_intermedias por factura: {_dbg_suma_lineas}", flush=True)
    print(f"[DBG-CONSOLIDADO] Percepciones extraídas por factura: {_dbg_percs_por_factura}", flush=True)

    # --- SECCIÓN 1 — Ventas ---
    seccion_ventas = {
        "total_ingresos_brutos": fact.get("total_ingresos"),
        "cantidad_operaciones": fact.get("cantidad_operaciones"),
        "ticket_promedio": fact.get("ticket_promedio"),
        "descuentos_aplicados": {
            "monto": fact.get("total_descuentos"),
            "cantidad": fact.get("cantidad_descuentos"),
        },
        "envios_pagados_por_comprador": fact.get("envios_pagados_comprador"),
        "notas_credito_ml": nc_ml.get("total"),
        "notas_credito_flex": nc_flex.get("total_bonificado"),
        "notas_debito_flex": nd_flex.get("total_debitado"),
        "_incompleto": reportes["facturacion_ml"] is None,
    }

    # --- SECCIÓN 2 — Costos de MercadoLibre ---
    desglose = fact.get("desglose_cargos", {})
    comisiones_venta      = _buscar_cargo_neto(desglose, "cargo por vender")
    costos_envio_ml_neto  = _buscar_cargo_neto(desglose, "cargo por envios de mercado libre")
    cuotas                = _buscar_cargo_neto(desglose, "costo por ofrecer cuotas")
    total_valor_del_cargo = fact.get("facturacion_neta")
    facturacion_bruta     = fact.get("facturacion_bruta")
    diff_facturado_vs_reporte = (
        round(total_facturas_ml - total_valor_del_cargo, 2)
        if total_valor_del_cargo is not None else None
    )
    seccion_costos_ml = {
        "comisiones_venta": comisiones_venta,
        "costos_envio_ml_neto": costos_envio_ml_neto,
        "cuotas": cuotas,
        "total_valor_del_cargo": total_valor_del_cargo,
        "facturacion_bruta": facturacion_bruta,
        "total_facturas_ml": total_facturas_ml,
        "cantidad_facturas": cantidad_facturas,
        "neto_gravado_total": neto_gravado_total,
        "iva_21_total": iva_21_total,
        "percepciones_ml_total": percepciones_ml_total,
        "diff_facturado_vs_reporte": diff_facturado_vs_reporte,
        "diff_alerta": diff_facturado_vs_reporte is not None and abs(diff_facturado_vs_reporte) > 1000,
        "_incompleto": not facturas_data or reportes["facturacion_ml"] is None,
    }

    # --- Percepciones (reportes de Percepciones), por provincia ---
    perc_reportes: Dict[str, float] = defaultdict(float)
    perc_reportes_detalle: Dict[str, list] = defaultdict(list)
    for f in archivos.get("percepciones", []):
        d = _ed(f)
        etiqueta = _extraer_impuesto_percepciones(f.get("filename", "")) or d.get("impuesto") or ""
        monto = d.get("monto_percibido") or 0.0
        jurisdiccion = normalizar_jurisdiccion_percepcion(etiqueta) or "Sin identificar"
        perc_reportes[jurisdiccion] += monto
        perc_reportes_detalle[jurisdiccion].append(f"{etiqueta or f.get('filename', '')}: {_ar_money(round(monto, 2))}")
    perc_reportes = {k: round(v, 2) for k, v in perc_reportes.items()}

    print(f"[DBG-CONSOLIDADO] percepciones_por_jurisdiccion_facturas = {perc_facturas}", flush=True)
    print(f"[DBG-CONSOLIDADO] percepciones_por_jurisdiccion_reportes = {perc_reportes}", flush=True)

    _todas_juris = set(perc_facturas) | set(perc_reportes)
    _provs_ordenadas = sorted(_todas_juris - {"Sin identificar"}) + (
        ["Sin identificar"] if "Sin identificar" in _todas_juris else []
    )
    cruce_percepciones = []
    for prov in _provs_ordenadas:
        a, b = perc_facturas.get(prov, 0.0), perc_reportes.get(prov, 0.0)
        ok = abs(a - b) < 1
        cruce_percepciones.append({
            "provincia": prov, "facturas_ml": a, "reportes": b,
            "diff": round(a - b, 2), "ok": ok,
            "detalle_facturas": perc_facturas_detalle.get(prov, []) if not ok else [],
            "detalle_reportes": perc_reportes_detalle.get(prov, []) if not ok else [],
        })

    # --- Retenciones sufridas ---
    retenciones_detalle = []
    total_retenciones_neto = 0.0
    for f in archivos.get("retenciones", []):
        d = _ed(f)
        neto = (d.get("importe_retenido") or 0.0) - (d.get("importe_devuelto") or 0.0)
        retenciones_detalle.append({
            "impuesto": d.get("impuesto") or f.get("filename"),
            "base_imponible": d.get("base_imponible"),
            "importe_retenido": d.get("importe_retenido"),
            "importe_devuelto": d.get("importe_devuelto"),
            "neto": round(neto, 2),
        })
        total_retenciones_neto += neto
    total_retenciones_neto = round(total_retenciones_neto, 2)

    # --- Créditos fiscales (percepciones + retenciones) por categoría de impuesto ---
    creditos_fiscales: Dict[str, float] = defaultdict(float)
    for monto in perc_reportes.values():
        creditos_fiscales["IIBB"] += monto
    for r in retenciones_detalle:
        creditos_fiscales[_clasificar_impuesto(r["impuesto"])] += r["neto"]
    creditos_fiscales = {k: round(v, 2) for k, v in creditos_fiscales.items()}

    # --- Pagos a ARCA por tipo ---
    pagos_arca_por_tipo: Dict[str, float] = defaultdict(float)
    for f in archivos.get("pagos_arca", []):
        d = _ed(f)
        tipo = d.get("tipo") or "Otro"
        pagos_arca_por_tipo[tipo] += d.get("importe_total_a_pagar") or 0.0
    pagos_arca_por_tipo = {k: round(v, 2) for k, v in pagos_arca_por_tipo.items()}

    # --- IVA: Débito Fiscal vs Crédito Fiscal vs Pago a ARCA ---
    ventas_gravadas = fact.get("total_ingresos") or 0.0
    iva_debito_fiscal = round(ventas_gravadas * 0.105 / 1.105, 2)
    iva_credito_fiscal_facturas = round(sum(
        l.get("monto") or 0.0
        for fd in facturas_data for l in _lineas_intermedias(fd)
        if "iva" in _concepto_norm(l.get("concepto", ""))
        and "percep" not in _concepto_norm(l.get("concepto", ""))
    ), 2)
    retenciones_iva = round(sum(
        r["neto"] for r in retenciones_detalle
        if "iva" in _concepto_norm(r["impuesto"])
    ), 2)
    iva_a_pagar_estimado = round(iva_debito_fiscal - iva_credito_fiscal_facturas - retenciones_iva, 2)
    iva_pagado_arca = pagos_arca_por_tipo.get("IVA", 0.0)
    diff_iva = round(iva_a_pagar_estimado - iva_pagado_arca, 2)
    _umbral_iva = max(abs(iva_pagado_arca), abs(iva_a_pagar_estimado)) * 0.05
    iva_analisis = {
        "ventas_gravadas": ventas_gravadas,
        "iva_debito_fiscal": iva_debito_fiscal,
        "iva_credito_fiscal_facturas": iva_credito_fiscal_facturas,
        "retenciones_iva": retenciones_iva,
        "iva_a_pagar_estimado": iva_a_pagar_estimado,
        "iva_pagado_arca": iva_pagado_arca,
        "diff": diff_iva,
        "ok": abs(diff_iva) <= _umbral_iva,
    }

    # --- Cruce Impuestos vs Pagos ARCA, por categoría ---
    categorias = sorted(set(creditos_fiscales) | {_clasificar_impuesto(t) for t in pagos_arca_por_tipo})
    cruce_impuestos_pagos = []
    for cat in categorias:
        credito = creditos_fiscales.get(cat, 0.0)
        pagado = round(sum(v for t, v in pagos_arca_por_tipo.items() if _clasificar_impuesto(t) == cat), 2)
        neto = round(credito - pagado, 2)
        cruce_impuestos_pagos.append({
            "concepto": cat, "total_credito": credito, "pagado_arca": pagado, "neto": neto,
            "saldo": "A favor" if neto > 0 else ("A pagar" if neto < 0 else "Sin saldo"),
        })

    seccion_impuestos = {
        "percepciones_facturas_ml": perc_facturas,
        "percepciones_reportes": perc_reportes,
        "cruce_percepciones": cruce_percepciones,
        "iva_analisis": iva_analisis,
        "retenciones_detalle": retenciones_detalle,
        "total_retenciones_neto": total_retenciones_neto,
        "creditos_fiscales": creditos_fiscales,
        "pagos_arca_por_tipo": pagos_arca_por_tipo,
        "cruce_impuestos_pagos": cruce_impuestos_pagos,
        "_incompleto": (
            not archivos.get("retenciones") or not archivos.get("percepciones")
            or not archivos.get("pagos_arca")
        ),
    }

    # --- SECCIÓN 3.5 — Facturación por Provincia (Reporte Facturación ML) ---
    facturacion_provincia_disponible = reportes["facturacion_ml"] is not None
    facturacion_provincia_filas = []
    if facturacion_provincia_disponible:
        _prov_agg_consol: Dict[str, dict] = defaultdict(lambda: {"monto": 0.0, "cantidad": 0})
        for p in fact.get("provincias_completas") or []:
            jurisdiccion = normalizar_jurisdiccion_percepcion(p.get("nombre", "")) or p.get("nombre") or "Sin identificar"
            _prov_agg_consol[jurisdiccion]["monto"] += p.get("monto") or 0.0
            _prov_agg_consol[jurisdiccion]["cantidad"] += p.get("cantidad_ventas") or 0
        _total_fact_prov = sum(x["monto"] for x in _prov_agg_consol.values())
        for nombre, info in sorted(_prov_agg_consol.items(), key=lambda kv: kv[1]["monto"], reverse=True):
            monto = round(info["monto"], 2)
            cantidad = info["cantidad"]
            facturacion_provincia_filas.append({
                "provincia": nombre,
                "monto": monto,
                "porcentaje": round(monto / _total_fact_prov * 100, 2) if _total_fact_prov else 0.0,
                "cantidad_ventas": cantidad,
                "ticket_promedio": round(monto / cantidad, 2) if cantidad else 0.0,
            })
    seccion_facturacion_provincia = {
        "disponible": facturacion_provincia_disponible,
        "filas": facturacion_provincia_filas,
        "total_monto": round(sum(f["monto"] for f in facturacion_provincia_filas), 2),
        "total_ventas": sum(f["cantidad_ventas"] for f in facturacion_provincia_filas),
    }

    # --- SECCIÓN 4 — Flujo financiero neto ---
    ingresos_brutos = fact.get("total_ingresos") or 0.0
    facturacion_neta = fact.get("facturacion_neta") or 0.0
    otros_costos_ml = round(facturacion_neta - comisiones_venta - costos_envio_ml_neto - cuotas, 2)
    total_percepciones_ml = fact.get("total_percepciones") or 0.0
    nd_flex_total = nd_flex.get("total_debitado") or 0.0
    nc_ml_total = nc_ml.get("total") or 0.0
    envios_comprador = fact.get("envios_pagados_comprador") or 0.0

    _fuentes_percepciones_flujo = ["fact", "perc"] if (perc_facturas and perc_reportes) else ["repo"]
    lineas_flujo = [
        {"concepto": "Ingresos brutos por ventas", "monto": round(ingresos_brutos, 2), "fuentes": ["repo"]},
        {"concepto": "Comisiones ML netas", "monto": round(-comisiones_venta, 2), "fuentes": ["repo"]},
        {"concepto": "Costos de envío ML netos", "monto": round(-costos_envio_ml_neto, 2), "fuentes": ["repo"]},
        {"concepto": "Cuotas", "monto": round(-cuotas, 2), "fuentes": ["repo"]},
        {"concepto": "Otros costos ML", "monto": round(-otros_costos_ml, 2), "fuentes": ["calc"]},
        {"concepto": "Percepciones (van a AFIP)", "monto": round(-total_percepciones_ml, 2), "fuentes": _fuentes_percepciones_flujo},
        {"concepto": "Retenciones sufridas", "monto": round(-total_retenciones_neto, 2), "fuentes": ["reten"]},
        {"concepto": "Notas de débito EnvíosFlex", "monto": round(-nd_flex_total, 2), "fuentes": ["repo"]},
        {"concepto": "Notas de crédito ML", "monto": round(nc_ml_total, 2), "fuentes": ["repo"]},
        {"concepto": "Envíos pagados por comprador", "monto": round(envios_comprador, 2), "fuentes": ["repo"]},
    ]
    cobrado_neto = round(sum(l["monto"] for l in lineas_flujo), 2)
    seccion_flujo = {"lineas": lineas_flujo, "cobrado_neto": cobrado_neto}

    # --- SECCIÓN 5 — Validaciones y alertas ---
    validaciones = []

    if total_valor_del_cargo is None or not facturas_data:
        validaciones.append({
            "check": "Facturas ML suman coherente con Facturación neta del reporte",
            "status": "warn", "detalle": "Datos incompletos para comparar.",
            "fuentes": ["fact", "repo"],
        })
    elif abs(diff_facturado_vs_reporte) <= 1000:
        validaciones.append({
            "check": "Facturas ML suman coherente con Facturación neta del reporte",
            "status": "ok",
            "detalle": f"Facturas ML suman {_ar_money(total_facturas_ml)} — coincide con Facturación neta.",
            "fuentes": ["fact", "repo"],
        })
    else:
        validaciones.append({
            "check": "Facturas ML suman coherente con Facturación neta del reporte",
            "status": "error",
            "detalle": (
                f"Facturas ML ({_ar_money(total_facturas_ml)}) vs Facturación neta "
                f"({_ar_money(total_valor_del_cargo)}) — diff {_ar_money(diff_facturado_vs_reporte)}"
            ),
            "fuentes": ["fact", "repo"],
        })

    if not cruce_percepciones:
        validaciones.append({
            "check": "Percepciones en Facturas ML coinciden con reportes de Percepciones",
            "status": "warn", "detalle": "Sin datos de percepciones para cruzar.",
            "fuentes": ["fact", "perc"],
        })
    else:
        malas = [x for x in cruce_percepciones if not x["ok"]]
        if malas:
            detalle = "; ".join(
                f"{x['provincia']}: Facturas {_ar_money(x['facturas_ml'])} vs Reporte "
                f"{_ar_money(x['reportes'])} (diff {_ar_money(x['diff'])})"
                for x in malas
            )
            validaciones.append({
                "check": "Percepciones en Facturas ML coinciden con reportes de Percepciones",
                "status": "error", "detalle": detalle,
                "fuentes": ["fact", "perc"],
            })
        else:
            validaciones.append({
                "check": "Percepciones en Facturas ML coinciden con reportes de Percepciones",
                "status": "ok", "detalle": "Todas las jurisdicciones coinciden.",
                "fuentes": ["fact", "perc"],
            })

    if faltantes:
        validaciones.append({
            "check": "Todos los archivos de las secciones están procesados y aprobados",
            "status": "warn", "detalle": f"Falta procesar: {', '.join(faltantes)}.",
            "fuentes": ["calc"],
        })
    else:
        validaciones.append({
            "check": "Todos los archivos de las secciones están procesados y aprobados",
            "status": "ok", "detalle": "Las 6 secciones tienen archivos procesados.",
            "fuentes": ["calc"],
        })

    if not cruce_impuestos_pagos:
        validaciones.append({
            "check": "Pagos a ARCA cubren los impuestos declarados",
            "status": "warn", "detalle": "Sin datos suficientes para el cruce.",
            "fuentes": ["perc", "reten", "arca"],
        })
    else:
        saldos_a_pagar = [x for x in cruce_impuestos_pagos if x["neto"] < 0]
        if saldos_a_pagar:
            detalle = "; ".join(
                f"{x['concepto']}: falta pagar {_ar_money(-x['neto'])}" for x in saldos_a_pagar
            )
            validaciones.append({
                "check": "Pagos a ARCA cubren los impuestos declarados",
                "status": "error", "detalle": detalle,
                "fuentes": ["perc", "reten", "arca"],
            })
        else:
            validaciones.append({
                "check": "Pagos a ARCA cubren los impuestos declarados",
                "status": "ok", "detalle": "Los pagos a ARCA cubren los impuestos declarados.",
                "fuentes": ["perc", "reten", "arca"],
            })

    if ingresos_brutos:
        pct_retenciones = abs(total_retenciones_neto) / ingresos_brutos * 100
        validaciones.append({
            "check": "No hay retenciones inusuales (> 15% de la facturación)",
            "status": "warn" if pct_retenciones > 15 else "ok",
            "detalle": f"Retenciones representan {_ar_pct_simple(pct_retenciones)} de la facturación.",
            "fuentes": ["reten"],
        })
    else:
        validaciones.append({
            "check": "No hay retenciones inusuales (> 15% de la facturación)",
            "status": "warn", "detalle": "Sin facturación para calcular el ratio.",
            "fuentes": ["reten"],
        })

    cargo_venta = next(
        (info for label, info in desglose.items() if _strip_accents(label.lower()) == "cargo por vender"),
        None,
    )
    if cargo_venta is None:
        validaciones.append({
            "check": "Ratio de anulaciones aceptable (< 10%)",
            "status": "warn", "detalle": "Sin datos de cargos por venta.",
            "fuentes": ["repo"],
        })
    else:
        prop = cargo_venta.get("proporcion_anulaciones") or 0.0
        validaciones.append({
            "check": "Ratio de anulaciones aceptable (< 10%)",
            "status": "ok" if prop < 10 else "warn",
            "detalle": f"Anulaciones: {_ar_pct_simple(prop)} de las ventas.",
            "fuentes": ["repo"],
        })

    # --- SECCIÓN 6 — Panorama impositivo ---
    panorama = []
    for cat in ("IIBB", "IVA", "Ganancias", "Impuesto Créditos y Débitos"):
        percepciones_cat = round(sum(perc_reportes.values()), 2) if cat == "IIBB" else 0.0
        retenciones_cat = round(sum(
            r["neto"] for r in retenciones_detalle if _clasificar_impuesto(r["impuesto"]) == cat
        ), 2)
        pagado_cat = round(sum(
            v for t, v in pagos_arca_por_tipo.items() if _clasificar_impuesto(t) == cat
        ), 2)
        saldo = round(percepciones_cat + retenciones_cat - pagado_cat, 2)
        if saldo > 0:
            recomendacion = "Podés compensar contra el próximo período"
        elif saldo < 0:
            recomendacion = f"Tenés que pagar {_ar_money(abs(saldo))} en la próxima DDJJ"
        else:
            recomendacion = "Sin saldo pendiente"
        panorama.append({
            "impuesto": cat,
            "total_percepciones": percepciones_cat,
            "total_retenciones": retenciones_cat,
            "total_pagado_arca": pagado_cat,
            "saldo": saldo,
            "recomendacion": recomendacion,
        })

    # --- Doble período: fiscal (mes calendario, para Impuestos/ARCA) vs ML (período de
    # corte tomado del propio Reporte de Facturación, para Ventas/Facturación/Cruce) ---
    inicio_fiscal, fin_fiscal = _rango_fechas_periodo(periodo)
    ml_desde_dt, ml_hasta_dt, ml_fallback = _calcular_periodo_ml(fact, inicio_fiscal, fin_fiscal)
    periodo_fiscal = {"desde": inicio_fiscal.strftime("%d/%m/%Y"), "hasta": fin_fiscal.strftime("%d/%m/%Y")}
    periodo_ml = {
        "desde": ml_desde_dt.strftime("%d/%m/%Y"), "hasta": ml_hasta_dt.strftime("%d/%m/%Y"),
        "fallback": ml_fallback,
    }

    seccion_cruce_ventas = _calcular_seccion_cruce_ventas(
        user_id, reportes["facturacion_ml"], ml_desde_dt, ml_hasta_dt
    )

    return {
        "user_id": user_id,
        "periodo": periodo,
        "periodo_fiscal": periodo_fiscal,
        "periodo_ml": periodo_ml,
        "faltantes": faltantes,
        "ventas": seccion_ventas,
        "costos_ml": seccion_costos_ml,
        "impuestos": seccion_impuestos,
        "facturacion_provincia": seccion_facturacion_provincia,
        "flujo_financiero": seccion_flujo,
        "validaciones": validaciones,
        "panorama_impositivo": panorama,
        "cruce_ventas": seccion_cruce_ventas,
    }


def _ar_pct_simple(n: float) -> str:
    return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + " %"


def _render_consolidado_html(resultado: dict) -> str:
    """Arma el HTML de las 6 secciones del análisis consolidado para el modal."""
    v         = resultado["ventas"]
    c         = resultado["costos_ml"]
    imp       = resultado["impuestos"]
    fp        = resultado["facturacion_provincia"]
    flujo     = resultado["flujo_financiero"]
    validaciones = resultado["validaciones"]
    panorama  = resultado["panorama_impositivo"]
    faltantes = resultado["faltantes"]
    periodo_fiscal, periodo_ml = _periodos_del_resultado(resultado)

    _SEP = (
        f"font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.06em;"
        f"color:{_HDR_COLOR};background:{_HDR_BG};padding:10px 14px;"
        f"border-bottom:2px solid {_HDR_BORDER};display:flex;align-items:center;gap:8px"
    )
    _WRAP = "border-bottom:1px solid #e5e5e5;padding-bottom:14px;margin-bottom:14px"

    def _sec(icon: str, titulo: str, periodo_tipo: Optional[str] = None) -> str:
        badge = _render_periodo_badge(periodo_tipo, periodo_fiscal, periodo_ml) if periodo_tipo else ""
        return f'<div style="{_SEP}"><i class="ti {icon}"></i><span>{titulo}</span>{badge}</div>'

    def _row(label, value, bold: bool = False, fuentes: Optional[list] = None) -> str:
        w = ";font-weight:700" if bold else ""
        badges = render_fuente_badges(fuentes)
        return (
            '<div style="display:flex;justify-content:space-between;align-items:center;padding:3px 14px;'
            f'border-bottom:1px solid #f5f5f5;font-size:12px{w}">'
            f'<span style="color:#555">{label}</span>'
            f'<span style="display:flex;align-items:center;gap:6px">'
            f'<span style="color:{_BLUE};font-variant-numeric:tabular-nums">{value}</span>'
            f'{badges}</span></div>'
        )

    def _incompleto_html() -> str:
        return (
            '<div style="padding:6px 14px;font-size:11px;color:#A32D2D;font-style:italic">'
            "Datos incompletos</div>"
        )

    def _tabla(headers: list, filas: list) -> str:
        ths = "".join(
            f'<th style="border:1px solid #d0d0d0;padding:3px 8px;background:#eef3f8;'
            f'font-size:10px;font-weight:600;color:{_HDR_COLOR};text-align:left">{h}</th>'
            for h in headers
        )
        body = ""
        for fila in filas:
            tds = "".join(
                f'<td style="border:1px solid #e8e8e8;padding:3px 8px;font-size:11px;color:#333">{c}</td>'
                for c in fila
            )
            body += f"<tr>{tds}</tr>"
        return (
            '<table style="border-collapse:collapse;width:100%;margin-top:4px">'
            f"<thead><tr>{ths}</tr></thead><tbody>{body}</tbody></table>"
        )

    def _leyenda_fuentes_html() -> str:
        badges = "".join(render_fuente_badge(k, with_label=True) for k in FUENTES_CONSOLIDADO)
        return (
            '<div style="position:sticky;top:0;z-index:100;'
            "background:#FFFFFF !important;background-color:#FFFFFF !important;"
            'border-bottom:1px solid #E5E7EB;'
            'padding:12px 14px;margin:-1px 0 12px 0;'
            'box-shadow:0 2px 6px rgba(0,0,0,0.06)">'
            '<div style="text-transform:uppercase;font-size:10px;letter-spacing:0.05em;'
            f'color:{_GRAY};font-weight:600;margin-bottom:8px">'
            "Íconos de fuente de datos</div>"
            f'<div style="display:flex;flex-wrap:wrap;gap:8px">{badges}</div>'
            "</div>"
        )

    partes = [_leyenda_fuentes_html()]

    if faltantes:
        partes.append(
            '<div style="background:#FBF1DC;color:#7A5A0E;border:1px solid #E2A93B;'
            'border-radius:4px;padding:8px 14px;margin-bottom:14px;font-size:12px">'
            f'⚠ Falta procesar: {", ".join(faltantes)} (0 archivos)</div>'
        )

    # 1 — Ventas
    s = [f'<div style="{_WRAP}">', _sec("ti-shopping-cart", "1 · Ventas", "ml")]
    if v["_incompleto"]:
        s.append(_incompleto_html())
    s.append(_row("Ingresos brutos", _ar_money(v["total_ingresos_brutos"]), fuentes=["repo"]))
    s.append(_row("Cantidad de operaciones", _ar_num(v["cantidad_operaciones"]), fuentes=["repo"]))
    s.append(_row("Ticket promedio", _ar_money(v["ticket_promedio"]), fuentes=["calc"]))
    desc = v["descuentos_aplicados"]
    s.append(_row(
        "Descuentos aplicados",
        f'{_ar_money(desc["monto"])} ({_ar_num(desc["cantidad"] or 0)} descuentos)',
        fuentes=["repo"],
    ))
    s.append(_row("Envíos pagados por comprador", _ar_money(v["envios_pagados_por_comprador"]), fuentes=["repo"]))
    s.append(_row("Notas de crédito ML", _ar_money(v["notas_credito_ml"]), fuentes=["repo"]))
    s.append(_row("Notas de crédito EnvíosFlex", _ar_money(v["notas_credito_flex"]), fuentes=["repo"]))
    s.append(_row("Notas de débito EnvíosFlex", _ar_money(v["notas_debito_flex"]), fuentes=["repo"]))
    s.append("</div>")
    partes.append("".join(s))

    # 2 — Costos de MercadoLibre
    s = [f'<div style="{_WRAP}">', _sec("ti-receipt-tax", "2 · Costos de MercadoLibre", "fiscal")]
    if c["_incompleto"]:
        s.append(_incompleto_html())
    s.append(_row("Comisiones de venta (neto anulaciones)", _ar_money(c["comisiones_venta"]), fuentes=["repo"]))
    s.append(_row("Costos de envío ML (neto)", _ar_money(c["costos_envio_ml_neto"]), fuentes=["repo"]))
    s.append(_row("Cuotas", _ar_money(c["cuotas"]), fuentes=["repo"]))
    s.append(_row("Total valor del cargo (facturación neta ML)", _ar_money(c["total_valor_del_cargo"]), fuentes=["repo"]))
    s.append(_row("Facturación bruta", _ar_money(c["facturacion_bruta"]), fuentes=["repo"]))
    s.append(_row("Total Facturas ML", _ar_money(c["total_facturas_ml"]), fuentes=["fact"]))
    s.append(_row("Cantidad de facturas", _ar_num(c["cantidad_facturas"]), fuentes=["fact"]))
    s.append(_row("Neto gravado total", _ar_money(c["neto_gravado_total"]), fuentes=["fact"]))
    s.append(_row("IVA 21% total", _ar_money(c["iva_21_total"]), fuentes=["fact"]))
    s.append(_row("Percepciones (Facturas ML) total", _ar_money(c["percepciones_ml_total"]), fuentes=["fact"]))
    diff_color = _RED if c["diff_alerta"] else _GREEN
    s.append(
        '<div style="display:flex;justify-content:space-between;align-items:center;padding:5px 14px;'
        f'font-size:12px;font-weight:700"><span style="color:#555">Diff. Facturado vs Reporte</span>'
        f'<span style="display:flex;align-items:center;gap:6px">'
        f'<span style="color:{diff_color}">{_ar_money(c["diff_facturado_vs_reporte"])}</span>'
        f'{render_fuente_badges(["calc"])}</span></div>'
    )
    s.append("</div>")
    partes.append("".join(s))

    # 3 — Impuestos y Retenciones
    s = [f'<div style="{_WRAP}">', _sec("ti-building-bank", "3 · Impuestos y Retenciones", "fiscal")]
    if imp["_incompleto"]:
        s.append(_incompleto_html())
    s.append(
        '<div style="padding:6px 14px 0;font-size:11px;font-weight:700;color:#555">'
        "Percepciones por provincia — Facturas ML vs Reportes</div>"
    )
    if imp["cruce_percepciones"]:
        filas = []
        for r in imp["cruce_percepciones"]:
            if r["ok"]:
                simbolo = (
                    f'<div style="text-align:center"><i class="ti ti-check" '
                    f'style="color:{_GREEN};font-size:16px"></i></div>'
                )
            else:
                _tt = "Facturas ML:\n" + ("\n".join(r["detalle_facturas"]) or "(sin líneas)")
                _tt += "\n\nReportes Percepciones:\n" + ("\n".join(r["detalle_reportes"]) or "(sin líneas)")
                _tt_esc = _tt.replace('"', "&quot;").replace("<", "&lt;")
                simbolo = (
                    f'<div style="text-align:center" title="{_tt_esc}">'
                    f'<i class="ti ti-x" style="color:{_RED};font-size:16px;cursor:help"></i></div>'
                )
            filas.append([
                r["provincia"], _ar_money(r["facturas_ml"]), _ar_money(r["reportes"]),
                _ar_money(r["diff"]), simbolo, render_fuente_badges(["fact", "perc"]),
            ])
        s.append(f'<div style="padding:0 14px">{_tabla(["Provincia", "Facturas ML", "Reportes Perc.", "Diff", "OK", "Fuente"], filas)}</div>')
    else:
        s.append('<div style="padding:4px 14px;font-size:11px;color:#9e9e9e">Sin datos</div>')

    s.append(_sec("ti-percentage", "IVA — Débito Fiscal vs Crédito Fiscal vs Pago a ARCA"))
    iva = imp["iva_analisis"]
    s.append(_row("Ventas gravadas (aprox.)", _ar_money(iva["ventas_gravadas"]), fuentes=["repo"]))
    s.append(_row("IVA Débito Fiscal (10,5% estimado)", _ar_money(iva["iva_debito_fiscal"]), fuentes=["calc"]))
    s.append(_row("IVA Crédito Fiscal (Facturas ML)", _ar_money(iva["iva_credito_fiscal_facturas"]), fuentes=["fact"]))
    s.append(_row("Retenciones IVA sufridas", _ar_money(iva["retenciones_iva"]), fuentes=["reten"]))
    s.append(_row("IVA a Pagar Estimado", _ar_money(iva["iva_a_pagar_estimado"]), bold=True, fuentes=["calc"]))
    s.append(_row("IVA Pagado a ARCA", _ar_money(iva["iva_pagado_arca"]), fuentes=["arca"]))
    _iva_diff_color = _GREEN if iva["ok"] else _RED
    s.append(
        '<div style="display:flex;justify-content:space-between;align-items:center;padding:5px 14px;'
        f'font-size:12px;font-weight:700"><span style="color:#555">Diferencia</span>'
        f'<span style="display:flex;align-items:center;gap:6px">'
        f'<span style="color:{_iva_diff_color}">{_ar_money(iva["diff"])}</span>'
        f'{render_fuente_badges(["calc"])}</span></div>'
    )
    s.append(
        '<div style="padding:4px 14px 8px;font-size:10px;color:#9e9e9e;font-style:italic">'
        "El IVA débito fiscal se calcula al 10,5% asumiendo que todas las ventas están gravadas a esa "
        "alícuota. Puede diferir del real si hay ventas al 21% u otras alícuotas.</div>"
    )

    s.append('<div style="padding:10px 14px 0;font-size:11px;font-weight:700;color:#555">Retenciones sufridas</div>')
    if imp["retenciones_detalle"]:
        filas = [
            [r["impuesto"], _ar_money(r["base_imponible"]), _ar_money(r["importe_retenido"]),
             _ar_money(r["importe_devuelto"]), _ar_money(r["neto"]), render_fuente_badges(["reten"])]
            for r in imp["retenciones_detalle"]
        ]
        s.append(f'<div style="padding:0 14px">{_tabla(["Impuesto", "Base Imponible", "Retenido", "Devuelto", "Neto", "Fuente"], filas)}</div>')
    else:
        s.append('<div style="padding:4px 14px;font-size:11px;color:#9e9e9e">Sin datos</div>')

    s.append('<div style="padding:10px 14px 0;font-size:11px;font-weight:700;color:#555">Cruce Impuestos vs Pagos a ARCA</div>')
    if imp["cruce_impuestos_pagos"]:
        filas = [
            [r["concepto"], _ar_money(r["total_credito"]), _ar_money(r["pagado_arca"]),
             _ar_money(r["neto"]), r["saldo"], render_fuente_badges(["perc", "reten", "arca"])]
            for r in imp["cruce_impuestos_pagos"]
        ]
        s.append(f'<div style="padding:0 14px">{_tabla(["Concepto", "Total Crédito", "Pagado ARCA", "Neto", "Saldo", "Fuente"], filas)}</div>')
    else:
        s.append('<div style="padding:4px 14px;font-size:11px;color:#9e9e9e">Sin datos</div>')
    s.append("</div>")
    partes.append("".join(s))

    # 3.5 — Facturación por Provincia
    s = [f'<div style="{_WRAP}">', _sec("ti-map-pin", "3.5 · Facturación por Provincia", "ml")]
    if not fp["disponible"]:
        s.append(
            '<div style="padding:6px 14px;font-size:11px;color:#A32D2D;font-style:italic">'
            "⚠ No hay Reporte de Facturación MercadoLibre procesado en este período. "
            "Subir el archivo en la tarjeta Reportes MercadoLibre para ver el detalle.</div>"
        )
    elif fp["filas"]:
        filas = [
            [
                r["provincia"], _ar_money(r["monto"]), _ar_pct_simple(r["porcentaje"]),
                _ar_num(r["cantidad_ventas"]), _ar_money(r["ticket_promedio"]), render_fuente_badges(["repo"]),
            ]
            for r in fp["filas"]
        ]
        filas.append([
            "<b>TOTAL</b>", f'<b>{_ar_money(fp["total_monto"])}</b>', "<b>100,00 %</b>",
            f'<b>{_ar_num(fp["total_ventas"])}</b>', "", render_fuente_badges(["repo"]),
        ])
        s.append(
            f'<div style="padding:0 14px">'
            f'{_tabla(["Provincia", "Facturación", "%", "Ventas", "Ticket Prom.", "Fuente"], filas)}</div>'
        )
    else:
        s.append('<div style="padding:4px 14px;font-size:11px;color:#9e9e9e">Sin datos</div>')
    s.append("</div>")
    partes.append("".join(s))

    # 4 — Flujo Financiero Neto
    s = [f'<div style="{_WRAP}">', _sec("ti-cash", "4 · Flujo Financiero Neto", "mixto")]
    for l in flujo["lineas"]:
        s.append(_row(l["concepto"], _ar_money(l["monto"]), fuentes=l.get("fuentes")))
    cobrado_color = _GREEN if flujo["cobrado_neto"] >= 0 else _RED
    s.append(
        '<div style="display:flex;justify-content:space-between;align-items:center;padding:6px 14px;'
        f'font-size:13px;font-weight:700;border-top:2px solid {_HDR_BORDER};margin-top:4px">'
        f'<span style="color:#333">Cobrado neto de ML</span>'
        f'<span style="display:flex;align-items:center;gap:6px">'
        f'<span style="color:{cobrado_color}">{_ar_money(flujo["cobrado_neto"])}</span>'
        f'{render_fuente_badges(["calc"])}</span></div>'
    )
    s.append("</div>")
    partes.append("".join(s))

    # 5 — Validaciones y Alertas
    s = [f'<div style="{_WRAP}">', _sec("ti-shield-check", "5 · Validaciones y Alertas")]
    _ICONS = {"ok": ("✓", _GREEN), "warn": ("⚠", _YELLOW), "error": ("✗", _RED)}
    for chk in validaciones:
        icono, color = _ICONS.get(chk["status"], ("•", "#9e9e9e"))
        s.append(
            '<div style="display:flex;gap:8px;padding:5px 14px;font-size:12px;align-items:flex-start">'
            f'<span style="color:{color};font-weight:700">{icono}</span>'
            f'<span style="flex:1"><b>{chk["check"]}</b> — {chk["detalle"]}</span>'
            f'<span style="display:flex;gap:3px">{render_fuente_badges(chk.get("fuentes"))}</span></div>'
        )
    s.append("</div>")
    partes.append("".join(s))

    # 6 — Panorama Impositivo
    s = ['<div>', _sec("ti-scale", "6 · Panorama Impositivo")]
    filas = [
        [p["impuesto"], _ar_money(p["total_percepciones"]), _ar_money(p["total_retenciones"]),
         _ar_money(p["total_pagado_arca"]), _ar_money(p["saldo"]), p["recomendacion"],
         render_fuente_badges(["perc", "reten", "arca"])]
        for p in panorama
    ]
    s.append(
        f'<div style="padding:6px 14px">'
        f'{_tabla(["Impuesto", "Percepciones", "Retenciones", "Pagado ARCA", "Saldo", "Recomendación", "Fuente"], filas)}'
        f'</div>'
    )
    s.append("</div>")
    partes.append("".join(s))

    return "".join(partes)


_MESES_ES = [
    "", "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre",
]


def _render_seccion_cruce_ventas(resultado: dict) -> None:
    """Sección 7 — se renderiza con componentes NiceGUI nativos (no HTML estático como
    las secciones 1-6) porque los íconos de descarga por card necesitan un callback
    real de Python (ui.download), algo que un string de HTML inyectado no puede disparar."""
    cr = resultado.get("cruce_ventas") or {"_disponible": False}
    periodo_fiscal, periodo_ml = _periodos_del_resultado(resultado)

    _sep_style = (
        "font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.06em;"
        f"color:{_HDR_COLOR};background:{_HDR_BG};padding:10px 14px;"
        f"border-bottom:2px solid {_HDR_BORDER};display:flex;align-items:center;gap:8px"
    )
    ui.html(
        '<style>.cruce-card{background:var(--color-background-secondary);'
        'border:0.5px solid var(--color-border-tertiary);border-radius:6px;'
        'padding:12px 14px;display:flex;align-items:center;gap:12px}'
        '.cruce-card.is-zero{opacity:0.55}</style>'
        f'<div style="{_sep_style}"><i class="ti ti-arrows-exchange"></i>'
        f'<span>7 · Cruce Ventas Nuestras vs Reporte ML</span>'
        f'{_render_periodo_badge("ml", periodo_fiscal, periodo_ml)}</div>'
    )

    if not cr.get("_disponible"):
        ui.html(
            '<div style="padding:6px 14px;font-size:11px;color:#A32D2D;font-style:italic">'
            "⚠ No hay Reporte de Facturación MercadoLibre procesado en este período. "
            "Subir el archivo en la tarjeta Reportes MercadoLibre para ver el cruce.</div>"
        )
        return

    async def _descargar(tipo: Optional[str]) -> None:
        path, nombre = await run.io_bound(_generar_excel_cruce_ventas, resultado, tipo)
        ui.download(path, nombre)
        ui.notify(f"Exportado: {nombre}", color="positive")

        def _cleanup() -> None:
            try:
                if path and os.path.exists(path):
                    os.unlink(path)
            except Exception:
                pass
        ui.timer(5.0, _cleanup, once=True)

    cruzadas_ok = cr.get("cruzadas_ok") or []
    solo_reporte = cr.get("solo_reporte") or []
    solo_bd = cr.get("solo_bd") or []
    con_diferencias = cr.get("con_diferencias") or []
    anuladas = cr.get("anuladas") or []
    total_reporte = cr.get("total_reporte", 0)

    # --- Hero de salud del período ---
    salud = cr.get("salud_pct", 0.0)
    if salud >= 95:
        hero_bg1, hero_bg2, hero_border = "#EAF3DE", "#F7F9F2", "#7FCFA0"
        hero_icon_color, hero_value_color, hero_title_color = "#3B6D11", "#27500A", "#5A7A3F"
        hero_icon = "ti-mood-happy"
    elif salud >= 85:
        hero_bg1, hero_bg2, hero_border = "#FBF1DC", "#FCFAF4", "#E9C77B"
        hero_icon_color, hero_value_color, hero_title_color = "#7A5A0E", "#6B4A08", "#8A6B2E"
        hero_icon = "ti-mood-neutral"
    else:
        hero_bg1, hero_bg2, hero_border = "#F8DFDD", "#FBF4F3", "#E39B94"
        hero_icon_color, hero_value_color, hero_title_color = "#A32D2D", "#7A1F1F", "#A85C56"
        hero_icon = "ti-mood-sad"

    salud_str = f"{salud:.1f}".replace(".", ",") + "%"
    ui.html(
        f'<div style="padding:16px 20px;background:linear-gradient(135deg,{hero_bg1},{hero_bg2});'
        f'border:0.5px solid {hero_border};border-radius:8px;margin:12px 14px;'
        'display:flex;align-items:center;gap:16px">'
        f'<i class="ti {hero_icon}" style="font-size:32px;color:{hero_icon_color}"></i>'
        '<div>'
        f'<div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.05em;'
        f'color:{hero_title_color}">Salud del período</div>'
        f'<div style="font-size:26px;font-weight:700;color:{hero_value_color}">{salud_str}</div>'
        f'<div style="font-size:11px;color:var(--color-text-secondary)">'
        f'{_ar_num(len(cruzadas_ok))} de {_ar_num(total_reporte)} ventas del reporte '
        'cruzan correctamente con tu BD.</div>'
        '</div></div>'
    )

    # --- Cards de detalle del cruce (grid 2x2) ---
    ui.html(
        '<div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.04em;'
        'color:var(--color-text-tertiary);padding:0 14px 6px">Detalle del cruce</div>'
    )

    categorias = [
        (len(cruzadas_ok), "ti-check", "#EAF3DE", "#3B6D11", "Cruzadas OK",
         "Coinciden ML y tu BD, montos correctos", "ok"),
        (len(anuladas), "ti-refresh-alert", "#EAF3DE", "#3B6D11", "Anuladas",
         "Devueltas o canceladas, marcadas OK", "anuladas"),
        (len(solo_reporte), "ti-alert-triangle", "#FBF1DC", "#7A5A0E", "Solo en Reporte ML",
         "Ventas que ML facturó pero faltan en tu BD", "solo_reporte"),
        (len(con_diferencias), "ti-scale", "#FBF1DC", "#7A5A0E", "Con diferencias",
         "Montos o comisiones que no coinciden", "diff"),
    ]

    with ui.grid(columns=2).classes("w-full gap-[10px]").style("padding:0 14px 14px"):
        for count, icono, icon_bg, icon_color, label, desc, tipo in categorias:
            es_cero = count == 0
            with ui.row().classes(f"cruce-card{' is-zero' if es_cero else ''} items-center"):
                ui.html(
                    f'<div style="width:32px;height:32px;flex-shrink:0;border-radius:6px;'
                    f'background:{icon_bg};color:{icon_color};display:flex;align-items:center;'
                    f'justify-content:center"><i class="ti {icono}" style="font-size:17px"></i></div>'
                )
                with ui.column().classes("gap-0").style("flex:1"):
                    ui.label(_ar_num(count)).style("font-size:20px;font-weight:700;font-variant-numeric:tabular-nums")
                    ui.label(label).style("font-size:11px;color:var(--color-text-secondary);margin-top:2px")
                    ui.label(desc).style("font-size:10px;color:var(--color-text-tertiary);margin-top:3px;line-height:1.35")
                btn_dl = ui.button(on_click=lambda t=tipo: _descargar(t)).props(
                    "flat dense round icon=download size=sm"
                ).style(f"color:{_BLUE};opacity:{'0.35' if es_cero else '1'}")
                if es_cero:
                    btn_dl.disable()

    # --- Sub-sección: ventas fuera del período de facturación ML ---
    if solo_bd:
        try:
            hasta_dt = datetime.strptime(periodo_ml.get("hasta") or "", "%d/%m/%Y")
            inicio_cross = hasta_dt + timedelta(days=1)
            if inicio_cross.month == 12:
                fin_mes = datetime(inicio_cross.year, 12, 31)
            else:
                fin_mes = datetime(inicio_cross.year, inicio_cross.month + 1, 1) - timedelta(days=1)
            mes_siguiente = _MESES_ES[fin_mes.month + 1] if fin_mes.month < 12 else _MESES_ES[1]
            rango_txt = f"Ventas del {inicio_cross.strftime('%d/%m')} al {fin_mes.strftime('%d/%m')}"
            mes_txt = f"ML las va a facturar en el próximo Reporte ({mes_siguiente}). No es un error, es información."
        except ValueError:
            rango_txt = "Ventas fuera del período de facturación ML"
            mes_txt = "ML las va a facturar en el próximo Reporte. No es un error, es información."

        ui.html(
            '<div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.04em;'
            'color:var(--color-text-tertiary);padding:8px 14px 6px">'
            'Ventas fuera del período de facturación ML</div>'
        )
        with ui.row().classes("cruce-card items-center w-full").style("margin:0 14px 14px"):
            ui.html(
                '<div style="width:32px;height:32px;flex-shrink:0;border-radius:6px;'
                'background:#E8F1FA;color:#185FA5;display:flex;align-items:center;'
                'justify-content:center"><i class="ti ti-clock" style="font-size:17px"></i></div>'
            )
            with ui.column().classes("gap-0").style("flex:1"):
                ui.label(_ar_num(len(solo_bd))).style("font-size:20px;font-weight:700;font-variant-numeric:tabular-nums")
                ui.label(rango_txt).style("font-size:11px;color:var(--color-text-secondary);margin-top:2px")
                ui.label(mes_txt).style("font-size:10px;color:var(--color-text-tertiary);margin-top:3px;line-height:1.35")
            ui.button(on_click=lambda: _descargar("solo_bd")).props(
                "flat dense round icon=download size=sm"
            ).style(f"color:{_BLUE}")

    with ui.row().classes("w-full justify-center").style("padding:10px 14px 14px"):
        ui.button(
            "Descargar Excel completo",
            on_click=lambda: _descargar(None),
        ).props("icon=download no-caps").style(f"background:{_BLUE};color:white")


def _abrir_modal_consolidado(resultado: dict) -> None:
    periodo_fiscal, periodo_ml = _periodos_del_resultado(resultado)
    with ui.dialog() as dlg:
        with ui.card().style(
            "width:90vw;max-width:1400px;height:90vh;overflow:hidden;"
            "display:flex;flex-direction:column;padding:0"
        ):
            with ui.row().classes("items-center justify-between w-full px-4 py-3 flex-shrink-0").style(
                f"background:{_HDR_BG};border-bottom:1px solid {_HDR_BORDER}"
            ):
                with ui.column().classes("gap-0"):
                    ui.label(f"Análisis Consolidado del Período — {resultado['periodo']}").style(
                        f"color:{_HDR_COLOR};font-weight:700;font-size:16px"
                    )
                    ui.html(
                        '<div style="font-size:11px;color:var(--color-text-secondary);'
                        'padding-top:4px;line-height:1.5">'
                        f"Período fiscal: {periodo_fiscal['desde']} - {periodo_fiscal['hasta']} "
                        "(Impuestos, ARCA)<br>"
                        f"Período ML: {periodo_ml['desde']} - {periodo_ml['hasta']} "
                        "(Facturación, Cruce)</div>"
                    )
                ui.button(icon="close", on_click=dlg.close).props("flat round dense")

            with ui.column().classes("w-full flex-1").style("overflow-y:auto;min-height:0;padding:14px 0"):
                if periodo_ml.get("fallback"):
                    ui.html(
                        '<div style="background:#FBF1DC;color:#7A5A0E;border:1px solid #E2A93B;'
                        'border-radius:4px;padding:8px 14px;margin:0 14px 14px;font-size:12px">'
                        "⚠ No hay Reporte de Facturación ML de este período. Las secciones marcadas "
                        'como "Período ML" usarán el mes calendario como fallback.</div>'
                    )
                ui.html(_render_consolidado_html(resultado))
                _render_seccion_cruce_ventas(resultado)

            with ui.row().classes("w-full justify-end gap-2 px-4 py-3 flex-shrink-0").style(
                f"border-top:1px solid {_HDR_BORDER};background:{_HDR_BG}"
            ):
                ui.button(
                    "Exportar PDF",
                    on_click=lambda: ui.notify("Exportar PDF: próximamente", color="info"),
                ).props("flat no-caps")
                ui.button(
                    "Exportar Excel",
                    on_click=lambda: ui.notify("Exportar Excel: próximamente", color="info"),
                ).props("flat no-caps")
                ui.button("Cerrar", on_click=dlg.close).style(
                    f"background:{_BLUE};color:white"
                ).props("no-caps")
    dlg.open()


def procesar_archivo_con_gemini(
    archivo_path: str, seccion: str, prompt_custom: Optional[str] = None
) -> dict:
    """Envía el archivo a Gemini y retorna los datos extraídos."""
    try:
        from google import genai
        from google.genai import types as genai_types
    except ImportError:
        return {
            "success": False, "data": {}, "prompt_used": "",
            "error": "Instalar: pip install google-genai",
        }

    api_key = get_app_config("gemini_api_key")
    if not api_key:
        return {
            "success": False, "data": {}, "prompt_used": "",
            "error": "API Key de Gemini no configurada (ir a Config → Gemini)",
        }

    prompt = prompt_custom or _PROMPTS_DEFAULT.get(seccion, "Extraé los datos del documento en JSON puro.")
    path = Path(archivo_path)
    ext = path.suffix.lower()

    print(f"[DBG-GASTOS] seccion={seccion!r} path={archivo_path!r}", flush=True)
    try:
        sz = path.stat().st_size
    except Exception:
        sz = -1
    print(f"[DBG-GASTOS] tamaño={sz} bytes", flush=True)

    calc_percepciones = None

    try:
        _es_nc_ml = _es_notas_credito_ml(path.name)
        if seccion == "reportes_ml" and ext in (".xlsx", ".xls") and _es_nc_ml:
            filas_sin_hoja = [f for f in leer_excel_completo(path) if not f.startswith("=== HOJA:")]
            calc_nc_ml = _calcular_notas_credito_ml(filas_sin_hoja)
            if calc_nc_ml:
                print(f"[DBG-NOTAS-CREDITO-ML-CALC] {calc_nc_ml}", flush=True)
                return {
                    "success": True, "data": calc_nc_ml, "error": None,
                    "prompt_used": "(calculado en Python, sin Gemini)",
                }
            print("[DBG-NOTAS-CREDITO-ML-CALC] no se encontraron columnas fecha/valor del cargo", flush=True)

        _tipo_notas = (
            "credito" if _es_notas_credito(path.name) and not _es_nc_ml
            else "debito" if _es_notas_debito(path.name)
            else None
        )
        if seccion == "reportes_ml" and ext in (".xlsx", ".xls") and _tipo_notas:
            filas_sin_hoja = [f for f in leer_excel_completo(path) if not f.startswith("=== HOJA:")]
            calc_notas = _calcular_notas_bonificacion(filas_sin_hoja, _tipo_notas)
            tag = f"DBG-NOTAS-{_tipo_notas.upper()}-CALC"
            if calc_notas:
                print(f"[{tag}] {calc_notas}", flush=True)
                return {
                    "success": True, "data": calc_notas, "error": None,
                    "prompt_used": "(calculado en Python, sin Gemini)",
                }
            print(f"[{tag}] no se encontraron columnas fecha/valor", flush=True)

        if seccion == "reportes_ml" and ext in (".xlsx", ".xls") and _es_pagos_facturas(path.name):
            filas_sin_hoja = [f for f in leer_excel_completo(path) if not f.startswith("=== HOJA:")]
            calc_pagos = _calcular_pagos_facturas(filas_sin_hoja)
            if calc_pagos:
                print(
                    f"[DBG-PAGOS-FACTURAS-CALC] fecha_actualizacion={calc_pagos['fecha_actualizacion']} "
                    f"n_pagos={len(calc_pagos['pagos'])}",
                    flush=True,
                )
                return {
                    "success": True, "data": calc_pagos, "error": None,
                    "prompt_used": "(calculado en Python, sin Gemini)",
                }
            print("[DBG-PAGOS-FACTURAS-CALC] no se encontraron columnas fecha/estado/importe", flush=True)

        if seccion == "reportes_ml" and ext in (".xlsx", ".xls") and _es_facturacion_ml(path.name):
            calc_fact = analizar_facturacion_ml(path)
            if calc_fact:
                print(
                    f"[DBG-FACTURACION-ML-CALC] operaciones={calc_fact['cantidad_operaciones']} "
                    f"ingresos={calc_fact['total_ingresos']} neta={calc_fact['facturacion_neta']}",
                    flush=True,
                )
                return {
                    "success": True, "data": calc_fact, "error": None,
                    "prompt_used": "(calculado en Python, sin Gemini)",
                }
            print(
                "[DBG-FACTURACION-ML-CALC] no se encontró la hoja REPORT o no tiene filas de detalle",
                flush=True,
            )

        client = genai.Client(api_key=api_key)

        if ext == ".pdf":
            data = path.read_bytes()
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=[
                    genai_types.Part.from_bytes(data=data, mime_type="application/pdf"),
                    prompt,
                ],
            )
        elif ext in (".xlsx", ".xls"):
            filas_todas = leer_excel_completo(path)
            filas_sin_hoja = [f for f in filas_todas if not f.startswith("=== HOJA:")]

            if seccion == "percepciones":
                # Los totales se calculan en Python (sin límite de filas/tokens); a Gemini
                # solo se le manda la cabecera para que extraiga los campos de texto.
                calc_percepciones = _calcular_totales_percepciones(filas_sin_hoja)
                if calc_percepciones:
                    print(
                        f"[DBG-PERCEPCIONES-CALC] filas_tabla={calc_percepciones['filas_tabla']} "
                        f"base_imponible_calc=${calc_percepciones['base_imponible']} "
                        f"monto_percibido_calc=${calc_percepciones['monto_percibido']} "
                        f"fecha_desde_calc={calc_percepciones['fecha_desde']} "
                        f"fecha_hasta_calc={calc_percepciones['fecha_hasta']}",
                        flush=True,
                    )
                corte = _cortar_en_detalle(filas_sin_hoja)
                filas_texto = filas_sin_hoja[:corte] or filas_sin_hoja[:20]
            else:
                filas_texto = filas_todas[:500]

            print(f"[DBG-GASTOS] filas_totales={len(filas_todas)} filas_enviadas={len(filas_texto)}", flush=True)
            print(f"[DBG-GASTOS] primeras_3={filas_texto[:3]}", flush=True)
            print(f"[DBG-GASTOS] ultimas_3_enviadas={filas_texto[-3:]}", flush=True)
            if seccion == "percepciones":
                hojas = [
                    ln.replace("=== HOJA: ", "").replace(" ===", "")
                    for ln in filas_todas if ln.startswith("=== HOJA:")
                ]
                print(f"[DBG-PERCEPCIONES] path={str(path)!r}", flush=True)
                print(f"[DBG-PERCEPCIONES] filas_totales={len(filas_todas)}", flush=True)
                print(f"[DBG-PERCEPCIONES] hojas={hojas}", flush=True)
                print(f"[DBG-PERCEPCIONES] primeras_5={filas_todas[:5]}", flush=True)
                print(f"[DBG-PERCEPCIONES] ultimas_5={filas_todas[-5:]}", flush=True)
            response = client.models.generate_content(
                model="gemini-2.5-flash",
                contents=f"{prompt}\n\nDatos:\n" + "\n".join(filas_texto),
            )
        else:
            return {
                "success": False, "data": {}, "prompt_used": prompt,
                "error": f"Tipo no soportado: {ext}",
            }

        raw = response.text or ""
        print(f"[DBG-GASTOS] respuesta_gemini={raw[:400]!r}", flush=True)

        if seccion == "analisis_ml":
            return {
                "success": True, "data": {"texto_completo": raw},
                "error": None, "prompt_used": prompt,
            }

        m = re.search(r"\{.*\}", raw, re.DOTALL)
        data = json.loads(m.group()) if m else {"respuesta_raw": raw}

        if seccion == "percepciones":
            impuesto_calc = _extraer_impuesto_percepciones(path.name)
            if impuesto_calc:
                data["impuesto"] = impuesto_calc
            if calc_percepciones:
                data["base_imponible"] = calc_percepciones["base_imponible"]
                data["monto_percibido"] = calc_percepciones["monto_percibido"]
                if calc_percepciones["alicuota"]:
                    data["alicuota"] = calc_percepciones["alicuota"]
                if calc_percepciones["fecha_desde"]:
                    data["fecha_desde"] = calc_percepciones["fecha_desde"]
                if calc_percepciones["fecha_hasta"]:
                    data["fecha_hasta"] = calc_percepciones["fecha_hasta"]

        return {"success": True, "data": data, "error": None, "prompt_used": prompt}

    except Exception as exc:
        print(f"[DBG-GASTOS] excepcion={exc!r}", flush=True)
        return {"success": False, "data": {}, "error": str(exc), "prompt_used": prompt}


# ---------------------------------------------------------------------------
# Exported
# ---------------------------------------------------------------------------


def build_tab_gastos(container) -> None:
    user = _require_login()
    if not user:
        return
    with container:
        _build_gastos(user["id"])


# ---------------------------------------------------------------------------
# Main UI
# ---------------------------------------------------------------------------


def _build_gastos(user_id: int) -> None:
    now = datetime.now()

    # ── Barra superior ────────────────────────────────────────────────────────
    with ui.row().classes("w-full items-center gap-4 flex-wrap").style(
        "background:#f8fafc;border-bottom:1px solid #e0e0e0;padding:10px 16px"
    ):
        with ui.row().classes("items-center gap-2"):
            ui.label("Período:").classes("font-semibold text-sm text-gray-600")
            mes_sel = ui.select(options=_MESES, value=_MESES[now.month - 1]).style("width:148px")
            ano_sel = ui.select(
                options=[str(y) for y in range(now.year - 2, now.year + 2)],
                value=str(now.year),
            ).style("width:90px")
        progress_lbl = ui.label("").classes("text-sm text-gray-500 ml-2")
        size_lbl     = ui.label("").classes("text-sm text-gray-400")

    content = ui.column().classes("w-full p-4 gap-4")

    def _get_periodo() -> str:
        return f"{ano_sel.value}-{(_MESES.index(mes_sel.value) + 1):02d}"

    def _build_content() -> None:
        content.clear()
        periodo = _get_periodo()

        archivos_por_sec: Dict[str, list] = {}
        dot_refs:         Dict[str, Any]  = {}
        subtitle_ref:     list            = [None]
        final_btn_ref:    list            = [None]
        _seccion_botones: Dict[str, dict] = {}
        _secciones_visibles = [s for s in _SECCIONES if user_can_access_tab(user_id, s[0])]

        def _sec_state(sk: str) -> dict:
            return _seccion_botones.setdefault(sk, {"proc": None, "borrar": None, "aprobar": []})

        def _set_procesando(sk: str, procesando: bool) -> None:
            """Deshabilita/habilita Procesar Todo, Borrar Todos y los Aprobar de popups abiertos de esa sección."""
            st = _sec_state(sk)
            for key in ("proc", "borrar"):
                b = st.get(key)
                if b:
                    b.disable() if procesando else b.enable()
            for b in st.get("aprobar", []):
                b.disable() if procesando else b.enable()
                b.style(f"opacity:{'0.5' if procesando else '1'}")

        def _sec_verde(sk: str) -> bool:
            fs = archivos_por_sec.get(sk, [])
            return bool(fs) and all(f.get("extraction_status") == "procesado" for f in fs)

        def _count_verdes() -> int:
            return sum(1 for sk, *_ in _secciones_visibles if _sec_verde(sk))

        def _refresh_progress() -> None:
            n, total = _count_verdes(), len(_secciones_visibles)
            pct = int(n / total * 100) if total else 0
            progress_lbl.text = f"{n} de {total} secciones procesadas — {pct}%"
            tf = sum(len(v) for v in archivos_por_sec.values())
            tb = sum(sum(f.get("size_bytes", 0) for f in v) for v in archivos_por_sec.values())
            size_lbl.text = f"{tf} archivo(s) — {_fmt_size(tb)}" if tf else ""
            if subtitle_ref[0]:
                subtitle_ref[0].text = (
                    "Todas las secciones procesadas — listo para analizar"
                    if n == total
                    else f"Disponible cuando las {total} secciones estén procesadas — actualmente {n} de {total}"
                )
            if final_btn_ref[0]:
                final_btn_ref[0].enable() if n == total else final_btn_ref[0].disable()

        def _refresh_dot(sk: str) -> None:
            dot = dot_refs.get(sk)
            if dot:
                dot.style(_DOT.format(_semaforo_color(archivos_por_sec.get(sk, []))))

        # ── Modal visor de extracción IA ──────────────────────────────────────
        def _open_eye_modal(fa: dict, seccion: str, on_approve=None) -> None:
            path        = Path(fa["filepath"])
            ext         = path.suffix.lower()
            cur_status  = fa.get("extraction_status", "pendiente")
            prompt_init = get_gastos_prompt(user_id, seccion) or _PROMPTS_DEFAULT.get(seccion, "")
            try:
                data_dict = json.loads(fa.get("extracted_data") or "{}") or {}
            except Exception:
                data_dict = {}
            ya_procesado = cur_status == "procesado" or bool(data_dict)

            with ui.dialog() as dlg:
                with ui.card().style(
                    "width:90vw;max-width:1200px;max-height:90vh;overflow:hidden;"
                    "display:flex;flex-direction:column;padding:0"
                ):
                    # Header
                    with ui.row().classes("items-start justify-between w-full px-4 py-2 flex-shrink-0 gap-2").style(
                        f"background:{_HDR_BG};border-bottom:1px solid {_HDR_BORDER}"
                    ):
                        _fname_display = (
                            Path(fa["filename"]).stem.replace("_", " ").replace("-", " ")
                            + Path(fa["filename"]).suffix
                        )
                        ui.label(_fname_display).style(
                            f"color:{_HDR_COLOR};font-weight:700;font-size:15px;"
                            "white-space:normal;word-break:break-word;line-height:1.3;flex:1;min-width:0"
                        )
                        async def _cerrar():
                            dlg.close()
                        ui.button(icon="close", on_click=_cerrar).props("flat round dense").classes("flex-shrink-0")

                    # Body: dos columnas
                    with ui.row().classes("w-full flex-1").style("max-height:calc(90vh - 120px);overflow-y:auto;min-height:0"):

                        # IZQUIERDA — documento original
                        with ui.column().classes("p-3 gap-2 flex-1").style(
                            f"border-right:1px solid #e0e0e0;overflow-y:auto;min-width:0"
                        ):
                            ui.label("Documento original").style(
                                f"color:{_HDR_COLOR};font-weight:700;font-size:13px"
                            )
                            ui.label(f"Tipo: {ext.upper().lstrip('.')}").classes("text-xs text-gray-400")
                            ui.separator()
                            _desc_tipo = (
                                _tipo_reporte_ml_descripcion(fa["filename"])
                                if seccion == "reportes_ml" else None
                            )
                            _desc_texto = REPORTE_ML_DESCRIPCIONES.get(_desc_tipo) if _desc_tipo else None
                            if _desc_texto:
                                ui.html(
                                    '<div style="background:#EEF6FD;border-left:3px solid #2A7AC7;'
                                    'padding:10px 12px;border-radius:4px;font-size:11px;'
                                    'color:var(--color-text-secondary);line-height:1.5;margin-bottom:10px">'
                                    '<i class="ti ti-info-circle" style="margin-right:6px;'
                                    'vertical-align:middle"></i>'
                                    f'{_escape_prompt_html(_desc_texto).replace(chr(10), "<br>")}</div>'
                                )
                            _solo_fecha_actualizacion = seccion == "reportes_ml" and (
                                _es_pagos_facturas(fa["filename"]) or _es_facturacion_ml(fa["filename"])
                            )
                            if _solo_fecha_actualizacion and ext in (".xlsx", ".xls"):
                                try:
                                    _filas_izq = [
                                        f for f in leer_excel_completo(path) if not f.startswith("=== HOJA:")
                                    ]
                                    _fecha_act = _extraer_fecha_actualizacion(_filas_izq)
                                except Exception:
                                    _fecha_act = None
                                ui.label(f"Fecha de Actualización: {_fecha_act or '—'}").style(
                                    "font-size:12px;color:#333;font-weight:600;padding:4px 0"
                                )
                            elif seccion == "analisis_ml" and ext == ".pdf":
                                b64_pages = _pdf_all_pages_b64(path)
                                if b64_pages:
                                    for _b64_pg in b64_pages:
                                        ui.html(
                                            f'<img src="data:image/png;base64,{_b64_pg}" '
                                            'style="max-width:100%;border:1px solid #e0e0e0;'
                                            'border-radius:4px;margin-bottom:8px">'
                                        )
                                    ui.label(f"({len(b64_pages)} página(s))").classes(
                                        "text-xs text-gray-400 mt-1"
                                    )
                                else:
                                    ui.label("No se pudo renderizar el PDF").classes("text-xs text-red-500")
                            elif ext == ".pdf":
                                b64 = _pdf_first_page_b64(path)
                                if b64:
                                    ui.html(
                                        f'<img src="data:image/png;base64,{b64}" '
                                        'style="max-width:100%;border:1px solid #e0e0e0;border-radius:4px">'
                                    )
                                    ui.label("(Mostrando página 1)").classes("text-xs text-gray-400 mt-1")
                                else:
                                    ui.label("No se pudo renderizar el PDF").classes("text-xs text-red-500")
                            elif ext in (".xlsx", ".xls"):
                                ui.html(
                                    f'<div style="overflow-x:auto">{_excel_preview_html(path)}</div>'
                                )
                            else:
                                ui.label("Previsualización no disponible").classes("text-xs text-gray-400")

                        # DERECHA — datos extraídos
                        with ui.column().classes("p-3 gap-2 flex-1").style(
                            "overflow-y:auto;min-width:0"
                        ):
                            with ui.row().classes("items-center gap-2"):
                                ui.label("Datos extraídos por Gemini").style(
                                    f"color:{_HDR_COLOR};font-weight:700;font-size:13px"
                                )
                                ui.html(_badge_html(cur_status))

                            ui.separator()

                            kv_container = ui.column().classes("w-full gap-0")

                            def _render_kv(d: dict) -> None:
                                kv_container.clear()
                                with kv_container:
                                    if seccion == "analisis_ml":
                                        texto = (d or {}).get("texto_completo", "")
                                        if not texto:
                                            ui.label("Sin datos extraídos").classes(
                                                "text-xs text-gray-400 italic"
                                            )
                                            return
                                        _corte = re.search(
                                            r'\n\s*2\.\s*An[áa]lisis detallado', texto, re.IGNORECASE
                                        )
                                        if _corte:
                                            texto = texto[:_corte.start()]
                                        texto = texto.rstrip()
                                        ui.html(
                                            '<pre style="font-family:var(--font-mono, monospace);'
                                            'font-size:11px;line-height:1.5;white-space:pre-wrap;'
                                            'word-break:break-word;padding:12px;'
                                            'background:var(--color-background-secondary, #f5f5f5);'
                                            'border-radius:4px;max-height:calc(90vh - 200px);'
                                            f'overflow-y:auto">{_escape_prompt_html(texto)}</pre>'
                                        )
                                        return
                                    if seccion == "reportes_ml" and _es_facturacion_ml(fa.get("filename", "")):
                                        if not d or "total_ingresos" not in d:
                                            ui.label(
                                                "Sin datos extraídos (reprocesar archivo)"
                                            ).classes("text-xs text-gray-400 italic")
                                            return
                                        ui.html(_render_facturacion_ml_html(d))
                                        return
                                    if not d:
                                        ui.label("Sin datos extraídos").classes(
                                            "text-xs text-gray-400 italic"
                                        )
                                        return

                                    _MONEY_KEYS = {
                                        "monto", "total", "subtotal", "precio", "iva",
                                        "neto", "importe", "imponible", "valor", "suma",
                                        "cargo", "bonificacion", "percepcion", "retencion",
                                        "gravado", "honorario", "flete", "almacenaje",
                                    }

                                    def _is_money_key(k: str) -> bool:
                                        kl = k.lower()
                                        if kl.startswith("cantidad") or kl == "total_de_ventas":
                                            return False
                                        return any(mk in kl for mk in _MONEY_KEYS)

                                    def _fmt_money(val) -> str:
                                        try:
                                            n = float(val)
                                            s = f"{abs(n):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                                            return f"$ {'-' if n < 0 else ''}{s}"
                                        except Exception:
                                            return str(val)

                                    def _render_value(k: str, val, is_total: bool = False) -> None:
                                        _bw = ";font-weight:700" if is_total else ""
                                        # None / vacío
                                        if val is None or val == "" or val == []:
                                            ui.label("—").style("font-size:11px;color:#9e9e9e")
                                            return
                                        # Lista de dicts → mini-tabla
                                        if isinstance(val, list) and val and isinstance(val[0], dict):
                                            headers = list(val[0].keys())
                                            ths = "".join(
                                                f'<th style="border:1px solid #d0d0d0;padding:2px 8px;'
                                                f'background:#eef3f8;font-size:10px;font-weight:600;'
                                                f'color:{_HDR_COLOR};text-align:left">'
                                                f'{h.replace("_"," ").capitalize()}</th>'
                                                for h in headers
                                            )
                                            tbody = ""
                                            for row in val:
                                                tds = ""
                                                for h in headers:
                                                    cell = row.get(h)
                                                    if isinstance(cell, (int, float)) and _is_money_key(h):
                                                        cell_str = _fmt_money(cell)
                                                        td_style = (
                                                            f"color:{_BLUE};text-align:right;"
                                                            "font-variant-numeric:tabular-nums"
                                                        )
                                                    elif isinstance(cell, (int, float)):
                                                        cell_str = str(cell)
                                                        td_style = f"color:{_BLUE};text-align:right"
                                                    elif cell is None:
                                                        cell_str = "—"
                                                        td_style = "color:#9e9e9e"
                                                    else:
                                                        cell_str = str(cell)
                                                        td_style = "color:#333"
                                                    tds += (
                                                        f'<td style="border:1px solid #e8e8e8;'
                                                        f'padding:2px 8px;font-size:10px;{td_style}">'
                                                        f"{cell_str}</td>"
                                                    )
                                                tbody += f"<tr>{tds}</tr>"
                                            ui.html(
                                                f'<table style="border-collapse:collapse;width:100%;'
                                                f'background:#f9fbfe;border-radius:3px;margin-top:2px">'
                                                f"<thead><tr>{ths}</tr></thead>"
                                                f"<tbody>{tbody}</tbody></table>"
                                            )
                                            return
                                        # Lista de scalars → bullets
                                        if isinstance(val, list):
                                            with ui.column().classes("gap-0"):
                                                for item in val:
                                                    ui.label(f"• {item}").style("font-size:11px;color:#333")
                                            return
                                        # Dict → sub-tabla 2 columnas
                                        if isinstance(val, dict):
                                            with ui.column().classes("gap-0 w-full"):
                                                for dk, dv in val.items():
                                                    with ui.row().classes("items-start gap-1").style(
                                                        "border-bottom:1px solid #f5f5f5"
                                                    ):
                                                        ui.label(dk).style(
                                                            "font-size:10px;color:#9e9e9e;"
                                                            "width:100px;flex-shrink:0"
                                                        )
                                                        ui.label(str(dv) if dv is not None else "—").style(
                                                            "font-size:10px;color:#333"
                                                        )
                                            return
                                        # Número
                                        if isinstance(val, (int, float)):
                                            if _is_money_key(k):
                                                ui.label(_fmt_money(val)).style(
                                                    f"font-size:11px;color:{_BLUE};"
                                                    "font-variant-numeric:tabular-nums;"
                                                    "text-align:right;padding-right:12px;"
                                                    f"display:block;width:100%{_bw}"
                                                )
                                            else:
                                                _val_str = (
                                                    f"{val:,}".replace(",", ".")
                                                    if k == "total_de_ventas" else str(val)
                                                )
                                                ui.label(_val_str).style(
                                                    f"font-size:11px;color:{_BLUE};"
                                                    f"font-variant-numeric:tabular-nums{_bw}"
                                                )
                                            return
                                        # String / fallback
                                        val_str = str(val)
                                        if val_str.startswith("$"):
                                            ui.label(val_str).style(
                                                "font-size:11px;color:#333;"
                                                "text-align:right;padding-right:12px;"
                                                "font-variant-numeric:tabular-nums;"
                                                f"display:block;width:100%{_bw}"
                                            )
                                        else:
                                            ui.label(val_str).style(f"font-size:11px;color:#333{_bw}")

                                    def _concepto_to_key(concepto: str) -> str:
                                        s = str(concepto)
                                        prev = None
                                        while prev != s:
                                            prev = s
                                            s = re.sub(r'([A-Za-z])\.([A-Za-z])', r'\1\2', s)
                                        s = re.sub(r'[^\w\s]', ' ', s)
                                        s = re.sub(r'\s+', '_', s.lower().strip())
                                        s = re.sub(r'_+', '_', s).strip('_')
                                        return s or str(concepto).lower().replace(' ', '_')

                                    def prettify_key(key: str) -> str:
                                        _UPPER = {"iva", "iibb", "cae", "cuit", "rg", "sr", "srl", "caba"}
                                        _STOP = {
                                            'a', 'ante', 'bajo', 'con', 'contra', 'de', 'del', 'desde',
                                            'en', 'entre', 'hacia', 'hasta', 'para', 'por', 'según', 'sin',
                                            'sobre', 'tras', 'y', 'e', 'ni', 'o', 'u', 'el', 'la', 'los',
                                            'las', 'un', 'una', 'al',
                                        }
                                        words = key.replace("_", " ").split()
                                        result = []
                                        for i, w in enumerate(words):
                                            wl = w.lower()
                                            if wl in _UPPER:
                                                result.append(wl.upper())
                                            elif i > 0 and wl in _STOP:
                                                result.append(wl)
                                            else:
                                                result.append(w.capitalize())
                                        return " ".join(result)

                                    def _is_concepto_monto_list(val) -> bool:
                                        return (
                                            isinstance(val, list) and bool(val)
                                            and isinstance(val[0], dict)
                                            and "concepto" in val[0]
                                            and "monto" in val[0]
                                        )

                                    _ROW_STYLE = (
                                        "border-bottom:1px solid #f0f0f0;"
                                        "display:grid;"
                                        "grid-template-columns:minmax(230px,300px) 1fr;"
                                        "gap:0 20px;padding:2px 8px"
                                    )
                                    _KEY_STYLE = (
                                        "white-space:normal;word-break:break-word;"
                                        "padding-right:16px"
                                    )

                                    _HIDDEN_FIELDS = {
                                        'cae', 'cae_vto', 'cuit_receptor',
                                        'receptor', 'emisor', 'cuit_emisor',
                                    }
                                    if seccion == "pagos_arca":
                                        _HIDDEN_FIELDS = _HIDDEN_FIELDS | {
                                            'cuit', 'organismo_recaudador', 'tipo_pago', 'concepto',
                                            'monto_pagado', 'banco', 'fecha_vencimiento', 'subconcepto',
                                            'descripcion_reducida', 'codigo_jurisdiccion', 'nro_inscripcion',
                                            'formulario_origen', 'generado_por_usuario', 'fecha_generacion',
                                            'dia_expiracion', 'generado_desde_presentacion_de_dj_nro',
                                            'denominacion', 'secuencia', 'fecha_presentacion',
                                            'nro_transaccion', 'codigo_identificacion_presentacion_md5',
                                        }
                                    if seccion == "retenciones":
                                        _HIDDEN_FIELDS = _HIDDEN_FIELDS | {
                                            'cuit_agente', 'monto_retenido',
                                            'numero_comprobante', 'agente_retencion',
                                        }
                                    if seccion == "percepciones":
                                        _HIDDEN_FIELDS = _HIDDEN_FIELDS | {
                                            'agente_percepcion', 'agente_percepciones',
                                            'cuit_agente', 'numero_comprobante', 'monto',
                                            'fecha_percepcion', 'agente',
                                        }
                                    if seccion == "reportes_ml" and _es_notas_credito(fa.get("filename", "")):
                                        _CAMPOS_NOTAS_CREDITO = {
                                            'fecha_desde', 'fecha_hasta',
                                            'total_bonificado', 'cantidad_bonificaciones',
                                        }
                                        _HIDDEN_FIELDS = {k for k in d.keys() if k not in _CAMPOS_NOTAS_CREDITO}
                                    # Más específico que el bloque anterior: pisa _HIDDEN_FIELDS con el
                                    # set correcto de campos para la variante MercadoLibre (no EnvíosFlex).
                                    if seccion == "reportes_ml" and _es_notas_credito_ml(fa.get("filename", "")):
                                        _CAMPOS_NC_ML = {'fecha_desde', 'fecha_hasta', 'total'}
                                        _HIDDEN_FIELDS = {k for k in d.keys() if k not in _CAMPOS_NC_ML}
                                    if seccion == "reportes_ml" and _es_notas_debito(fa.get("filename", "")):
                                        _CAMPOS_NOTAS_DEBITO = {
                                            'fecha_desde', 'fecha_hasta',
                                            'total_debitado', 'cantidad_debitos',
                                        }
                                        _HIDDEN_FIELDS = {k for k in d.keys() if k not in _CAMPOS_NOTAS_DEBITO}
                                    if seccion == "reportes_ml" and _es_pagos_facturas(fa.get("filename", "")):
                                        _CAMPOS_PAGOS_FACTURAS = {'fecha_actualizacion', 'pagos'}
                                        _HIDDEN_FIELDS = {k for k in d.keys() if k not in _CAMPOS_PAGOS_FACTURAS}
                                    _SECTION_LABELS = {
                                        "lineas_convenio_multilateral": "Convenio Multilateral",
                                        "determinacion_del_impuesto": "Determinación del Impuesto",
                                        "determinacion_posicion_mensual": "Determinación de la Posición Mensual",
                                    }
                                    _SEP_STYLE = (
                                        f"width:100%;font-size:10px;font-weight:600;"
                                        f"text-transform:uppercase;letter-spacing:0.04em;"
                                        f"color:{_HDR_COLOR};background:{_HDR_BG};"
                                        f"padding:8px 12px 4px;"
                                        f"border-bottom:0.5px solid {_HDR_BORDER};"
                                        f"margin-top:8px;display:block"
                                    )
                                    for k, v in d.items():
                                        if k in _HIDDEN_FIELDS:
                                            continue
                                        # pagos (reportes_ml/pagos_facturas) → separador + fila por pago/NC
                                        if k == "pagos" and seccion == "reportes_ml" and isinstance(v, list):
                                            if v:
                                                ui.html(f'<div style="{_SEP_STYLE}">Pagos</div>')
                                            for pago in v:
                                                fecha_p = pago.get("fecha_pago") or "—"
                                                estado_p = pago.get("estado") or "—"
                                                nc_p = pago.get("nota_credito")
                                                label = f"{fecha_p} · {estado_p}"
                                                if nc_p:
                                                    label += f" ({nc_p})"
                                                importe_p = pago.get("importe_total")
                                                with ui.row().classes("w-full items-start py-1").style(_ROW_STYLE):
                                                    ui.label(label).classes(
                                                        "text-xs text-gray-500 font-medium"
                                                    ).style(_KEY_STYLE)
                                                    if isinstance(importe_p, (int, float)):
                                                        ui.label(_fmt_money(importe_p)).style(
                                                            f"font-size:11px;color:{_BLUE};"
                                                            "font-variant-numeric:tabular-nums;"
                                                            "text-align:right;padding-right:12px;"
                                                            "display:block;width:100%"
                                                        )
                                                    else:
                                                        ui.label("—").style("font-size:11px;color:#9e9e9e")
                                            continue
                                        # lineas_convenio_multilateral → separador + filas CM
                                        if k == "lineas_convenio_multilateral" and isinstance(v, list):
                                            if seccion == "pagos_arca" and v:
                                                ui.html(
                                                    f'<div style="{_SEP_STYLE}">'
                                                    f'{_SECTION_LABELS.get(k, prettify_key(k))}</div>'
                                                )
                                            for cm_item in v:
                                                juris = cm_item.get("jurisdiccion", "")
                                                code  = cm_item.get("codigo", "")
                                                cm_val = cm_item.get("monto")
                                                cm_key = f"CM {juris} ({code})" if code else f"CM {juris}"
                                                _is_tot = 'total' in cm_key.lower()
                                                _kw = ";font-weight:700" if _is_tot else ""
                                                with ui.row().classes("w-full items-start py-1").style(_ROW_STYLE):
                                                    ui.label(cm_key).classes(
                                                        "text-xs text-gray-500 font-medium"
                                                    ).style(_KEY_STYLE + _kw)
                                                    if isinstance(cm_val, (int, float)):
                                                        ui.label(_fmt_money(cm_val)).style(
                                                            f"font-size:11px;color:{_BLUE};"
                                                            "font-variant-numeric:tabular-nums;"
                                                            "white-space:nowrap;"
                                                            "text-align:right;padding-right:12px;"
                                                            f"display:block;width:100%{_kw}"
                                                        )
                                                    else:
                                                        ui.label(str(cm_val) if cm_val is not None else "—").style(
                                                            f"font-size:11px;color:#333{_kw}"
                                                        )
                                            continue
                                        # concepto+monto list → separador (pagos_arca) + filas con bold
                                        if _is_concepto_monto_list(v):
                                            if seccion == "pagos_arca" and k in _SECTION_LABELS:
                                                ui.html(
                                                    f'<div style="{_SEP_STYLE}">'
                                                    f'{_SECTION_LABELS[k]}</div>'
                                                )
                                            for item in v:
                                                concepto_raw = str(item.get("concepto", k))
                                                row_key = _concepto_to_key(concepto_raw)
                                                row_val = item.get("monto")
                                                _is_tot = 'total' in concepto_raw.lower()
                                                _kw = ";font-weight:700" if _is_tot else ""
                                                with ui.row().classes("w-full items-start py-1").style(_ROW_STYLE):
                                                    ui.label(prettify_key(row_key)).classes(
                                                        "text-xs text-gray-500 font-medium"
                                                    ).style(_KEY_STYLE + _kw)
                                                    if isinstance(row_val, (int, float)):
                                                        ui.label(_fmt_money(row_val)).style(
                                                            f"font-size:11px;color:{_BLUE};"
                                                            "font-variant-numeric:tabular-nums;"
                                                            "white-space:nowrap;"
                                                            "text-align:right;padding-right:12px;"
                                                            f"display:block;width:100%{_kw}"
                                                        )
                                                    else:
                                                        ui.label(str(row_val) if row_val is not None else "—").style(
                                                            f"font-size:11px;color:#333;white-space:nowrap{_kw}"
                                                        )
                                        else:
                                            _is_tot = 'total' in k.lower()
                                            _kw = ";font-weight:700" if _is_tot else ""
                                            with ui.row().classes("w-full items-start py-1").style(_ROW_STYLE):
                                                ui.label(prettify_key(k)).classes(
                                                    "text-xs text-gray-500 font-medium"
                                                ).style(_KEY_STYLE + _kw)
                                                _render_value(k, v, is_total=_is_tot)

                            _render_kv(data_dict)

                            _prompt_open = [False]
                            with ui.column().classes("w-full mt-3 gap-0"):
                                with ui.row().classes("items-center gap-1 py-1").style(
                                    "cursor:pointer;user-select:none"
                                ) as _prompt_hdr:
                                    _prompt_chev = ui.element("i").classes("ti ti-chevron-right").style(
                                        f"color:{_HDR_COLOR};font-size:14px"
                                    )
                                    ui.label("Prompt usado").classes("text-xs font-semibold text-gray-600")

                                prompt_display = ui.html(
                                    f'<pre style="{_PROMPT_PRE_STYLE}">'
                                    f'{_escape_prompt_html(prompt_init)}</pre>'
                                ).classes("w-full hidden")

                                prompt_ta = (
                                    ui.textarea(value=prompt_init)
                                    .props("outlined autogrow")
                                    .classes("w-full hidden")
                                    .style("font-size:11px;line-height:1.5")
                                )

                                def _toggle_prompt():
                                    _prompt_open[0] = not _prompt_open[0]
                                    if _prompt_open[0]:
                                        prompt_display.classes(remove="hidden")
                                        _prompt_chev.classes(remove="ti-chevron-right")
                                        _prompt_chev.classes(add="ti-chevron-down")
                                    else:
                                        prompt_display.classes(add="hidden")
                                        _prompt_chev.classes(remove="ti-chevron-down")
                                        _prompt_chev.classes(add="ti-chevron-right")

                                _prompt_hdr.on("click", _toggle_prompt)

                            reproc_lbl = ui.label("").classes("text-xs text-gray-500 mt-1")
                            reproc_btn_ref: list = [None]

                            async def _reprocesar() -> None:
                                _set_procesando(seccion, True)
                                if reproc_btn_ref[0]:
                                    reproc_btn_ref[0].disable()
                                    reproc_btn_ref[0].text = "Procesando..."
                                reproc_lbl.text = "Procesando con Gemini..."
                                result = await run.io_bound(
                                    procesar_archivo_con_gemini,
                                    fa["filepath"], seccion, prompt_ta.value,
                                )
                                new_status = "procesado" if result["success"] else "error"
                                update_gastos_extraccion(
                                    fa["id"],
                                    extracted_data=json.dumps(result["data"]) if result["success"] else None,
                                    prompt_used=result["prompt_used"],
                                    extraction_status=new_status,
                                    extraction_error=result.get("error"),
                                )
                                archivos_por_sec[seccion] = get_gastos_archivos(user_id, periodo, seccion)
                                _refresh_dot(seccion)
                                _refresh_progress()
                                if result["success"]:
                                    _render_kv(result["data"])
                                    reproc_lbl.text = "Re-procesado correctamente"
                                else:
                                    reproc_lbl.text = f"Error: {result['error']}"
                                if reproc_btn_ref[0]:
                                    reproc_btn_ref[0].enable()
                                    reproc_btn_ref[0].text = "Re-procesar"
                                _set_procesando(seccion, False)

                            async def _aprobar() -> None:
                                mark_gastos_procesado(fa["id"])
                                archivos_por_sec[seccion] = get_gastos_archivos(user_id, periodo, seccion)
                                _refresh_dot(seccion)
                                _refresh_progress()
                                if on_approve:
                                    on_approve(seccion)
                                ui.notify("Archivo aprobado", color="positive")
                                dlg.close()

                            with ui.row().classes("gap-2 mt-2 flex-wrap items-center"):
                                edit_btn = ui.button("Editar prompt").props("outline dense").style(
                                    f"color:{_BLUE};border-color:{_BLUE};font-size:12px"
                                )
                                guardar_btn = ui.button("Guardar").props("dense").style(
                                    f"background:{_GREEN};color:white;font-size:12px"
                                ).classes("hidden")
                                cancelar_btn = ui.button("Cancelar").props("outline dense").style(
                                    "font-size:12px"
                                ).classes("hidden")
                                _rb = ui.button("Re-procesar" if ya_procesado else "Procesar", on_click=_reprocesar).style(
                                    f"background:{_YELLOW};color:white;font-size:12px"
                                ).props("dense")
                                reproc_btn_ref[0] = _rb
                                _ab = ui.button("Aprobar", on_click=_aprobar).style(
                                    f"background:{_GREEN};color:white;font-size:12px"
                                ).props("dense")
                                _sec_state(seccion)["aprobar"].append(_ab)

                                def _unregister_aprobar(_b=_ab, _sk=seccion):
                                    _lst = _sec_state(_sk)["aprobar"]
                                    if _b in _lst:
                                        _lst.remove(_b)
                                dlg.on("hide", _unregister_aprobar)

                            # Prompt desactualizado → dialog Sí/No para actualizar
                            if seccion == "facturas_ml" and "lineas_antes_subtotal" not in prompt_init:
                                reproc_lbl.text = (
                                    "⚠ El prompt guardado no pide líneas arriba del subtotal."
                                )
                                with ui.dialog() as _upd_dlg:
                                    with ui.card().classes("p-4 gap-2").style("min-width:340px"):
                                        ui.label("Prompt desactualizado").classes("font-semibold text-sm")
                                        ui.label(
                                            "El prompt guardado no extrae lineas_antes_subtotal "
                                            "(bonificaciones y cargos antes del subtotal). "
                                            "¿Actualizar al prompt default nuevo?"
                                        ).classes("text-xs text-gray-600").style("line-height:1.5")
                                        with ui.row().classes("gap-2 justify-end w-full mt-2"):
                                            ui.button("No", on_click=_upd_dlg.close).props("flat dense")
                                            async def _actualizar_prompt(_d=_upd_dlg):
                                                new_p = _PROMPTS_DEFAULT["facturas_ml"]
                                                upsert_gastos_prompt(user_id, seccion, new_p)
                                                prompt_ta.value = new_p
                                                prompt_display.content = (
                                                    f'<pre style="{_PROMPT_PRE_STYLE}">'
                                                    f'{_escape_prompt_html(new_p)}</pre>'
                                                )
                                                reproc_lbl.text = (
                                                    "Prompt actualizado. Reprocesá todos para obtener datos completos."
                                                )
                                                _d.close()
                                            ui.button("Sí, actualizar", on_click=_actualizar_prompt).style(
                                                f"background:{_BLUE};color:white"
                                            ).props("dense")
                                _upd_dlg.open()

                            # Percepciones: datos extraídos con el prompt viejo (sin agregados) → ofrecer reprocesar
                            if seccion == "percepciones" and data_dict and "base_imponible" not in data_dict:
                                with ui.dialog() as _upd_dlg_perc:
                                    with ui.card().classes("p-4 gap-2").style("min-width:340px"):
                                        ui.label("Prompt actualizado").classes("font-semibold text-sm")
                                        ui.label(
                                            "El prompt fue actualizado. "
                                            "¿Reprocesar archivos existentes de esta sección?"
                                        ).classes("text-xs text-gray-600").style("line-height:1.5")
                                        with ui.row().classes("gap-2 justify-end w-full mt-2"):
                                            ui.button("No", on_click=_upd_dlg_perc.close).props("flat dense")
                                            async def _reprocesar_desde_dialog(_d=_upd_dlg_perc):
                                                _d.close()
                                                await _reprocesar()
                                            ui.button("Sí, reprocesar", on_click=_reprocesar_desde_dialog).style(
                                                f"background:{_BLUE};color:white"
                                            ).props("dense")
                                _upd_dlg_perc.open()

                            def _start_edit() -> None:
                                if not _prompt_open[0]:
                                    _toggle_prompt()
                                prompt_display.classes(add="hidden")
                                prompt_ta.classes(remove="hidden")
                                edit_btn.classes(add="hidden")
                                guardar_btn.classes(remove="hidden")
                                cancelar_btn.classes(remove="hidden")

                            async def _guardar_prompt() -> None:
                                new_val = prompt_ta.value
                                upsert_gastos_prompt(user_id, seccion, new_val)
                                prompt_ta.classes(add="hidden")
                                prompt_display.content = (
                                    f'<pre style="{_PROMPT_PRE_STYLE}">'
                                    f'{_escape_prompt_html(new_val)}</pre>'
                                )
                                prompt_display.classes(remove="hidden")
                                edit_btn.classes(remove="hidden")
                                guardar_btn.classes(add="hidden")
                                cancelar_btn.classes(add="hidden")
                                ui.notify("Prompt guardado", color="positive")

                            def _cancelar_edit() -> None:
                                prompt_ta.value = prompt_init
                                prompt_ta.classes(add="hidden")
                                prompt_display.content = (
                                    f'<pre style="{_PROMPT_PRE_STYLE}">'
                                    f'{_escape_prompt_html(prompt_init)}</pre>'
                                )
                                prompt_display.classes(remove="hidden")
                                edit_btn.classes(remove="hidden")
                                guardar_btn.classes(add="hidden")
                                cancelar_btn.classes(add="hidden")

                            edit_btn.on("click", _start_edit)
                            guardar_btn.on("click", _guardar_prompt)
                            cancelar_btn.on("click", _cancelar_edit)

            dlg.open()

        # ── Tarjeta de sección ────────────────────────────────────────────────
        def _build_section_card(sk: str, lbl: str, ext: str, multiple: bool, icon: str) -> None:
            rows = get_gastos_archivos(user_id, periodo, sk)
            archivos_por_sec[sk] = rows
            footer_lbl_ref:    list = [None]
            proc_lbl_ref:      list = [None]
            borrar_btn_ref:    list = [None]
            upload_ref:        list = [None]
            proc_btn_ref:      list = [None]
            proc_btn_html_ref: list = [None]
            is_proc_ref:       list = [False]

            def _proc_btn_html(txt, icon=None, spinning=False):
                if not icon:
                    return f'<span style="font-size:12px;font-weight:500">{txt}</span>'
                spin = "animation:spin 1s linear infinite;" if spinning else ""
                return (
                    f'<i class="ti {icon}" style="font-size:12px;margin-right:4px;'
                    f'vertical-align:middle;{spin}"></i>'
                    f'<span style="font-size:12px;font-weight:500;vertical-align:middle">{txt}</span>'
                )

            def _refresh_proc_btn(sk_=sk) -> None:
                btn = proc_btn_ref[0]
                if btn is None:
                    return
                fs = archivos_por_sec.get(sk_, [])
                _isp = is_proc_ref[0]
                can = bool(fs) if _isp else any(
                    f.get("extraction_status") != "procesado" for f in fs
                )
                btn.enable() if can else btn.disable()

            with ui.card().classes("w-full").style("border:1px solid #e0e0e0"):
                # Header
                with ui.row().classes("items-center gap-2 w-full px-3 py-2").style(
                    f"background:{_HDR_BG};border-bottom:1px solid {_HDR_BORDER};"
                    "border-radius:4px 4px 0 0"
                ):
                    dot = ui.element("span").style(_DOT.format(_semaforo_color(rows)))
                    dot_refs[sk] = dot
                    ui.element("i").classes(f"ti {icon}").style(f"color:{_HDR_COLOR};font-size:16px")
                    ui.label(lbl).style(f"color:{_HDR_COLOR};font-weight:600;font-size:14px")
                    if not multiple:
                        ui.label("(máx. 1 archivo)").classes("text-xs text-gray-400 ml-1")

                # Lista de archivos
                file_list = ui.column().classes("w-full px-3 pt-2 gap-1 min-h-[32px]")

                def _render_list(sk_=sk) -> None:
                    file_list.clear()
                    with file_list:
                        fs = archivos_por_sec.get(sk_, [])
                        if not fs:
                            ui.label("Sin archivos").classes("text-xs text-gray-400 italic")
                        else:
                            for fa in fs:
                                _fa = dict(fa)
                                status = _fa.get("extraction_status", "pendiente")
                                with ui.row().classes("items-center gap-1 w-full flex-nowrap"):
                                    ui.label(
                                        f"{_fa['filename']}  ({_fmt_size(_fa.get('size_bytes', 0))})"
                                    ).classes("text-xs flex-1 truncate text-gray-700")
                                    ui.html(_badge_html(status))

                                    # Ícono ojo
                                    ui.element("i").classes("ti ti-eye").style(
                                        f"color:{_BLUE};font-size:13px;cursor:pointer;flex-shrink:0"
                                    ).tooltip("Ver extracción IA").on(
                                        "click",
                                        lambda _fa_=_fa, sk2=sk_: _open_eye_modal(_fa_, sk2, on_approve=_render_list),
                                    )

                                    # Ícono tacho con confirmación
                                    def _confirm_del(fa_=_fa, sk2=sk_) -> None:
                                        with ui.dialog() as conf, ui.card().classes("p-4"):
                                            ui.label("¿Eliminar este archivo?").classes(
                                                "font-semibold text-sm mb-1"
                                            )
                                            ui.label(fa_["filename"]).classes(
                                                "text-xs text-gray-500 mb-4 truncate"
                                            ).style("max-width:280px")
                                            with ui.row().classes("gap-2 justify-end w-full"):
                                                ui.button("Cancelar", on_click=conf.close).props("flat dense")

                                                def _do_del(fa__=fa_, sk3=sk2, _c=conf) -> None:
                                                    try:
                                                        Path(fa__["filepath"]).unlink(missing_ok=True)
                                                    except Exception:
                                                        pass
                                                    delete_gastos_archivo(fa__["id"])
                                                    archivos_por_sec[sk3] = get_gastos_archivos(
                                                        user_id, periodo, sk3
                                                    )
                                                    _c.close()
                                                    _render_list(sk3)
                                                    _refresh_dot(sk3)
                                                    _refresh_progress()
                                                    ui.notify("Archivo eliminado", color="positive")

                                                ui.button("Eliminar", on_click=_do_del).style(
                                                    f"background:{_RED};color:white"
                                                ).props("dense")
                                        conf.open()

                                    ui.element("i").classes("ti ti-trash").style(
                                        f"color:{_RED};font-size:13px;cursor:pointer;flex-shrink:0"
                                    ).tooltip("Eliminar").on("click", _confirm_del)

                    # Actualizar contador del footer
                    cnt = len(archivos_por_sec.get(sk_, []))
                    if footer_lbl_ref[0]:
                        footer_lbl_ref[0].text = f"{cnt} archivo(s)" if cnt else "Sin archivos"
                    if borrar_btn_ref[0]:
                        borrar_btn_ref[0].enable() if cnt else borrar_btn_ref[0].disable()
                    _refresh_proc_btn(sk_)

                _render_list(sk)

                # Zona de upload
                with ui.element("div").classes("px-3 pb-2 pt-2"):
                    def _on_upload(e, sk_=sk, mul=multiple) -> None:
                        if not mul and archivos_por_sec.get(sk_):
                            ui.notify(
                                "Esta sección admite solo 1 archivo. Eliminá el existente primero.",
                                color="warning",
                            )
                            return
                        dest_dir = _BASE_PATH / str(user_id) / periodo[:4] / periodo[5:] / sk_
                        dest_dir.mkdir(parents=True, exist_ok=True)
                        dest = dest_dir / e.name
                        data = e.content.read()
                        dest.write_bytes(data)
                        insert_gastos_archivo(
                            user_id=user_id, periodo=periodo, seccion=sk_,
                            filename=e.name, filepath=str(dest), size_bytes=len(data),
                        )
                        archivos_por_sec[sk_] = get_gastos_archivos(user_id, periodo, sk_)
                        _render_list(sk_)
                        _refresh_dot(sk_)
                        _refresh_progress()
                        ui.notify(f"'{e.name}' subido", color="positive")

                    upload_ref[0] = ui.upload(
                        multiple=multiple, auto_upload=True, on_upload=_on_upload,
                        label="Arrastrá archivos aquí o hacé clic",
                    ).props(
                        f'accept="{ext}" flat hide-upload-btn color="primary"'
                    ).classes("w-full").style(
                        "border:2px dashed #85B7EB;border-radius:6px;background:#f9fbfe;min-height:56px"
                    )

                # Footer
                with ui.row().classes("items-center gap-2 px-3 pb-3 pt-1 flex-wrap"):
                    cnt0 = len(rows)
                    footer_lbl_ref[0] = ui.label(
                        f"{cnt0} archivo(s)" if cnt0 else "Sin archivos"
                    ).classes("text-xs text-gray-500 flex-1")
                    proc_lbl_ref[0] = ui.label("").classes("text-xs text-gray-400")

                    is_proc   = _sec_verde(sk)
                    is_proc_ref[0] = is_proc
                    btn_color = _GREEN if is_proc else _BLUE
                    btn_icon  = "ti-refresh" if is_proc else None
                    btn_txt   = "Reprocesar Todos" if is_proc else "Procesar Todo"

                    async def _procesar(sk_=sk) -> None:
                        _set_procesando(sk_, True)
                        if proc_btn_html_ref[0]:
                            proc_btn_html_ref[0].content = _proc_btn_html(
                                "Procesando...", "ti-loader-2"
                            )
                        pl = proc_lbl_ref[0]
                        fs = archivos_por_sec.get(sk_, [])
                        if not fs:
                            ui.notify("No hay archivos para procesar", color="warning")
                            if proc_btn_html_ref[0]:
                                proc_btn_html_ref[0].content = _proc_btn_html(btn_txt, btn_icon)
                            _set_procesando(sk_, False)
                            _refresh_proc_btn(sk_)
                            if borrar_btn_ref[0]:
                                borrar_btn_ref[0].enable() if archivos_por_sec.get(sk_) else borrar_btn_ref[0].disable()
                            return
                        pendientes = [f for f in fs if f.get("extraction_status") != "procesado"] or fs
                        if pl:
                            pl.text = f"Procesando 0 de {len(pendientes)}..."
                        _prompt_custom = get_gastos_prompt(user_id, sk_)
                        for i, fa in enumerate(pendientes, 1):
                            if pl:
                                pl.text = f"Procesando {i} de {len(pendientes)}..."
                            result = await run.io_bound(
                                procesar_archivo_con_gemini, fa["filepath"], sk_, _prompt_custom
                            )
                            new_status = "procesado" if result["success"] else "error"
                            update_gastos_extraccion(
                                fa["id"],
                                extracted_data=json.dumps(result["data"]) if result["success"] else None,
                                prompt_used=result["prompt_used"],
                                extraction_status=new_status,
                                extraction_error=result.get("error"),
                            )
                        archivos_por_sec[sk_] = get_gastos_archivos(user_id, periodo, sk_)
                        _render_list(sk_)
                        _refresh_dot(sk_)
                        _refresh_progress()
                        ok    = sum(1 for f in archivos_por_sec[sk_] if f.get("extraction_status") == "procesado")
                        total = len(archivos_por_sec[sk_])
                        if pl:
                            pl.text = f"{ok}/{total} procesados"
                        if ok == total:
                            ui.notify("Procesado correctamente", color="positive")
                        else:
                            ui.notify(f"{total - ok} archivo(s) con error — revisá el ícono ojo", color="warning")
                        if upload_ref[0] is not None:
                            await upload_ref[0].run_method("reset")
                        if proc_btn_html_ref[0]:
                            proc_btn_html_ref[0].content = _proc_btn_html(btn_txt, btn_icon)
                        _set_procesando(sk_, False)
                        _refresh_proc_btn(sk_)
                        if borrar_btn_ref[0]:
                            cnt_now = len(archivos_por_sec.get(sk_, []))
                            borrar_btn_ref[0].enable() if cnt_now else borrar_btn_ref[0].disable()

                    with ui.button(on_click=_procesar).style(
                        f"background:{btn_color};color:white;font-size:12px;"
                        "padding:4px 14px;border-radius:4px"
                    ).props("dense no-caps") as _proc_btn:
                        _html_el = ui.html(_proc_btn_html(btn_txt, btn_icon))
                        proc_btn_html_ref[0] = _html_el
                    proc_btn_ref[0] = _proc_btn
                    _sec_state(sk)["proc"] = _proc_btn
                    _refresh_proc_btn(sk)

                    def _confirm_borrar_todos(sk_=sk) -> None:
                        fs = archivos_por_sec.get(sk_, [])
                        n = len(fs)
                        if not n:
                            return
                        with ui.dialog() as conf_del, ui.card().classes("p-4"):
                            ui.label(
                                f"¿Eliminar los {n} archivos de esta sección?"
                            ).classes("font-semibold text-sm mb-1")
                            ui.label(
                                "Esta acción borra los archivos del disco Y las "
                                "extracciones guardadas. Es irreversible."
                            ).classes("text-xs text-gray-500 mb-4")
                            with ui.row().classes("gap-2 justify-end w-full"):
                                ui.button("Cancelar", on_click=conf_del.close).props("flat dense")
                                def _do_borrar_todos(sk2=sk_, _c=conf_del) -> None:
                                    fs2 = archivos_por_sec.get(sk2, [])
                                    count = len(fs2)
                                    for fa_item in fs2:
                                        try:
                                            Path(fa_item["filepath"]).unlink(missing_ok=True)
                                        except Exception:
                                            pass
                                        delete_gastos_archivo(fa_item["id"])
                                    archivos_por_sec[sk2] = get_gastos_archivos(
                                        user_id, periodo, sk2
                                    )
                                    _c.close()
                                    _render_list(sk2)
                                    _refresh_dot(sk2)
                                    _refresh_progress()
                                    ui.notify(f"{count} archivos eliminados", color="warning")
                                ui.button("Sí, borrar todos", on_click=_do_borrar_todos).style(
                                    f"background:{_RED};color:white"
                                ).props("dense")
                        conf_del.open()

                    with ui.button(on_click=_confirm_borrar_todos).style(
                        f"background:{_RED};color:white;height:30px;padding:0 12px;border-radius:4px"
                    ).props("dense no-caps") as _borrar_btn:
                        ui.html(
                            '<i class="ti ti-trash" style="font-size:12px;margin-right:6px;'
                            'vertical-align:middle"></i>'
                            '<span style="font-size:11px;font-weight:500;vertical-align:middle">'
                            'Borrar todos</span>'
                        )
                    if not cnt0:
                        _borrar_btn.disable()
                    borrar_btn_ref[0] = _borrar_btn
                    _sec_state(sk)["borrar"] = _borrar_btn

        # ── Grid de tarjetas ──────────────────────────────────────────────────
        with content:
            with ui.grid(columns=2).classes("w-full gap-4"):
                for sk, lbl, ext, mul, icon in _secciones_visibles:
                    _build_section_card(sk, lbl, ext, mul, icon)

                # Card análisis final — fila propia, ancho completo
                with ui.card().classes("w-full").style(
                    f"border:2px solid {_BLUE};background:{_BLUE_BG};border-radius:8px;"
                    "grid-column:1 / -1"
                ):
                    with ui.column().classes("p-4 gap-1"):
                        ui.label("Análisis consolidado del período").style(
                            f"color:{_BLUE};font-size:16px;font-weight:700"
                        )
                        subtitle = ui.label("").classes("text-sm text-gray-500")
                        subtitle_ref[0] = subtitle

                        _consolidado_previo = get_gastos_consolidado(user_id, periodo)
                        ultimo_lbl = ui.label(
                            f"Último análisis: {_consolidado_previo['_generado_at']}"
                            if _consolidado_previo else ""
                        ).classes("text-xs text-gray-400")

                        async def _final_procesar() -> None:
                            fb = final_btn_ref[0]
                            periodo_actual = _get_periodo()
                            if fb:
                                fb.disable()
                                fb.text = "Analizando 6 secciones..."
                            try:
                                resultado = await run.io_bound(
                                    analizar_periodo_consolidado, user_id, periodo_actual
                                )
                                save_gastos_consolidado(user_id, periodo_actual, resultado)
                                ultimo_lbl.text = f"Último análisis: {datetime.now().isoformat(timespec='seconds')}"
                                _abrir_modal_consolidado(resultado)
                            finally:
                                if fb:
                                    fb.text = "Procesar análisis final"
                                    fb.enable()

                        with ui.row().classes("items-center gap-2 mt-2"):
                            fb = ui.button("Procesar análisis final", on_click=_final_procesar).style(
                                f"background:{_BLUE};color:white;font-size:14px;"
                                "font-weight:600;padding:10px 24px"
                            )
                            final_btn_ref[0] = fb
                            if _consolidado_previo:
                                ui.button(
                                    "Ver último análisis",
                                    on_click=lambda r=_consolidado_previo: _abrir_modal_consolidado(r),
                                ).props("flat no-caps").style(f"color:{_BLUE}")

            _refresh_progress()

    mes_sel.on("update:model-value", lambda _: _build_content())
    ano_sel.on("update:model-value", lambda _: _build_content())
    _build_content()
